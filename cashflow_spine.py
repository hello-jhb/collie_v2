"""
cashflow_spine.py — the universal cash-flow engine finder.

The foundation of the rebuild. It answers ONE question for any workbook, with
zero reliance on sheet names, row labels, tab roles, units, or layout:

    Which dated cash-flow stream is the deal's true engine?

The only universal fact about a DCF real-estate model: it states a headline IRR
somewhere, and that IRR is produced by a dated cash-flow stream that goes
outflow -> inflows. So:

    1. find every cell that states an IRR (label-anchored),
    2. find every dated, sign-changing numeric series anywhere in the workbook,
       in EITHER orientation (periods across columns, or periods down rows),
    3. recompute each series' IRR (XIRR on real dates),
    4. the series whose recomputed IRR reproduces a stated IRR IS the engine.

This is scale-invariant (IRR ignores a thousands-vs-dollars mix), label-invariant
("USD" or "Levered Cash Flow" — irrelevant), role-invariant (works when sheet
classification is wrong), and layout-invariant. It is NOT tuned to any file:
the model's own stated IRR is the ground truth, so the same code self-validates
on every model.

`find_spine(path)` returns `ok=True` only when a stream validates against a
stated IRR. Downstream (DealTruth) must not claim "cash-flow reconstructed"
unless `ok` is True.

Run `python cashflow_spine.py <dir-or-files>` to validate across a whole corpus
at once: for each file it reports whether a stream reproduced that file's own
stated IRR (no per-file ground truth needed).
"""
from __future__ import annotations

import datetime as _dt
import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

log = logging.getLogger("fb.spine")
if not log.handlers:
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("[fb.spine] %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

SPINE_VERSION = "2026-06-15.1"

_MIN_PERIODS = 6          # an investment stream needs at least this many points
_IRR_MATCH_BPS = 75       # a recomputed IRR within this of a stated IRR = match
_MAX_ROWS = 400
_MAX_COLS = 400


# ---------------------------------------------------------------------------
# IRR
# ---------------------------------------------------------------------------

def _xnpv(rate: float, flows: list[tuple[_dt.date, float]]) -> float:
    t0 = flows[0][0]
    return sum(v / (1.0 + rate) ** ((d - t0).days / 365.0) for d, v in flows)


def xirr(flows: list[tuple[_dt.date, float]]) -> float | None:
    """Annualized IRR of dated flows; bracketed bisection so it can't diverge.
    Requires a sign change. Returns None when no root exists in a sane band."""
    if len(flows) < 2:
        return None
    if not (any(v > 0 for _, v in flows) and any(v < 0 for _, v in flows)):
        return None
    lo, hi = -0.95, 10.0
    flo, fhi = _xnpv(lo, flows), _xnpv(hi, flows)
    if flo == 0:
        return lo
    if flo * fhi > 0:
        return None
    for _ in range(200):
        mid = (lo + hi) / 2.0
        fm = _xnpv(mid, flows)
        if abs(fm) < 1.0:
            return mid
        if flo * fm < 0:
            hi = mid
        else:
            lo, flo = mid, fm
    return (lo + hi) / 2.0


# ---------------------------------------------------------------------------
# Cells / dates
# ---------------------------------------------------------------------------

_MONTHS = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"])}
import re as _re
_RE_MON_Y = _re.compile(r"^([A-Za-z]{3,9})[\s\-/.]+(\d{2,4})$")
_RE_MDY = _re.compile(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$")
_RE_YMD = _re.compile(r"^(\d{4})[-/.](\d{1,2})(?:[-/.](\d{1,2}))?$")


def _parse_textdate(s: str) -> _dt.date | None:
    """Parse common TEXT period headers into a date: 'Jul-2023', 'Jul 23',
    '1/31/2023', '2023-01-31', '2023-07'. Day defaults to 1 — spacing, not the
    day, is what matters for the axis and XIRR. Generic, not per-file."""
    s = s.strip()
    if not s or len(s) > 12:
        return None
    m = _RE_MON_Y.match(s)
    if m:
        mo = _MONTHS.get(m.group(1)[:3].lower())
        if mo:
            y = int(m.group(2))
            return _dt.date(2000 + y if y < 100 else y, mo, 1)
    m = _RE_MDY.match(s)
    if m:
        mo, da, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= da <= 31:
            try:
                return _dt.date(2000 + y if y < 100 else y, mo, da)
            except ValueError:
                return None
    m = _RE_YMD.match(s)
    if m:
        try:
            return _dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3) or 1))
        except ValueError:
            return None
    return None


def _as_date(v: Any) -> _dt.date | None:
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    return None


def _coerce_date(v: Any) -> _dt.date | None:
    """A date from a real date, a bare year-integer, or a text header."""
    d = _as_date(v)
    if d is not None:
        return d
    n = _num(v)
    if n is not None and 1900 <= n <= 2100 and float(n).is_integer():
        return _dt.date(int(n), 12, 31)
    if isinstance(v, str):
        return _parse_textdate(v)
    return None


def _num(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _date_run(seq: list[Any]) -> list[tuple[int, _dt.date]] | None:
    """The longest strictly-increasing run of dates / year-integers in a 0-indexed
    sequence of cells. Year-ints (1990..2100, step 1) become Dec-31 of that year.
    Returns [(index, date), ...] or None if shorter than _MIN_PERIODS."""
    best: list[tuple[int, _dt.date]] = []
    cur: list[tuple[int, _dt.date]] = []

    def flush():
        nonlocal best, cur
        if len(cur) > len(best):
            best = cur
        cur = []

    prev: _dt.date | None = None
    for i, v in enumerate(seq):
        d = _coerce_date(v)
        if d is not None and (prev is None or d > prev):
            cur.append((i, d))
            prev = d
        else:
            flush()
            prev = None
            if d is not None:
                cur = [(i, d)]
                prev = d
    flush()
    return best if len(best) >= _MIN_PERIODS else None


# ---------------------------------------------------------------------------
# Workbook -> grids
# ---------------------------------------------------------------------------

def _load_grids(file_path: Path) -> dict[str, list[tuple]]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    grids: dict[str, list[tuple]] = {}
    for s in wb.sheetnames:
        try:
            ws = wb[s]
            if not ws.max_row:
                continue
            grids[s] = [row for row in ws.iter_rows(
                min_row=1, max_row=min(ws.max_row, _MAX_ROWS),
                max_col=min(ws.max_column or 1, _MAX_COLS), values_only=True)]
        except Exception:
            continue
    try:
        wb.close()
    except Exception:
        pass
    return grids


def _cell(grid: list[tuple], r: int, c: int) -> Any:
    """0-indexed."""
    if 0 <= r < len(grid):
        row = grid[r]
        if 0 <= c < len(row):
            return row[c]
    return None


def _a1(r: int, c: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(c + 1)}{r + 1}"


# ---------------------------------------------------------------------------
# Stated IRRs (label-anchored; the targets to validate against)
# ---------------------------------------------------------------------------

def _leg_of(label: str) -> str:
    l = label.lower()
    if "unlever" in l or "un-lever" in l or "unleveraged" in l:
        return "unlevered"
    if "lever" in l or "equity" in l or "project" in l:
        return "levered"
    return "unknown"


def _norm_irr(v: float) -> float | None:
    """Normalize to a fraction; accept fraction (0.18) or percent (18.0) form.
    Reject values outside a sane IRR band."""
    f = v / 100.0 if abs(v) > 1.5 else v
    # Reject a degenerate ~0% "IRR" (a blank/placeholder near an IRR label is not
    # a real headline return and would match flat debt/expense rows).
    if abs(f) < 0.005:
        return None
    return f if -0.6 <= f <= 2.0 else None


_RETURN_PHRASES = ("levered return", "unlevered return", "leveraged return",
                   "unleveraged return", "equity return", "project return",
                   "deal return", "partnership return")


def _is_return_label(label: str) -> bool:
    """A label that names a deal IRR. Includes 'IRR'/'internal rate' and the
    common '<leg> Return' phrasings — but NOT bare 'return' (which catches
    return-on-cost, cash-on-cash, etc.). False positives are harmless: a value
    with no matching cash-flow stream is dropped."""
    l = label.lower()
    if "irr" in l or "internal rate" in l:
        return True
    return any(p in l for p in _RETURN_PHRASES)


def find_stated_irrs(grids: dict[str, list[tuple]]) -> list[dict]:
    """Every cell that states an IRR: a number in IRR range whose nearest label
    (left in its row, or above in its column) names an IRR. Deduped per value+leg."""
    out: list[dict] = []
    seen: set = set()
    for sheet, grid in grids.items():
        for r, row in enumerate(grid):
            for c, v in enumerate(row):
                n = _num(v)
                if n is None:
                    continue
                f = _norm_irr(n)
                if f is None:
                    continue
                label = _label_left(grid, r, c) or _label_above(grid, r, c) or ""
                if not _is_return_label(label):
                    continue
                leg = _leg_of(label)
                key = (round(f, 4), leg)
                if key in seen:
                    continue
                seen.add(key)
                out.append({"value": f, "leg": leg, "sheet": sheet,
                            "cell": _a1(r, c), "label": label.strip()[:40]})
    return out


def _label_left(grid: list[tuple], r: int, c: int, reach: int = 12) -> str | None:
    for cc in range(c - 1, max(-1, c - 1 - reach), -1):
        v = _cell(grid, r, cc)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def _label_above(grid: list[tuple], r: int, c: int, reach: int = 4) -> str | None:
    for rr in range(r - 1, max(-1, r - 1 - reach), -1):
        v = _cell(grid, rr, c)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


# ---------------------------------------------------------------------------
# Candidate streams (BOTH orientations)
# ---------------------------------------------------------------------------

@dataclass
class Stream:
    sheet: str
    orientation: str                 # 'row' (periods across cols) | 'col' (down rows)
    index: int                       # row index (row-stream) or col index (col-stream)
    label: str
    flows: list[tuple[_dt.date, float]]
    irr: float
    em: float | None
    initial_outflow: float
    n: int
    loc: str                         # e.g. "Sheet!row12" / "Sheet!col F"


def periodicity_of(flows: list[tuple[_dt.date, float]]) -> str:
    """Coarse periodicity from the median spacing of a dated stream."""
    if len(flows) < 2:
        return "unknown"
    import statistics
    gaps = [(flows[i + 1][0] - flows[i][0]).days for i in range(len(flows) - 1)]
    med = statistics.median(gaps) if gaps else 0
    if med <= 0:
        return "unknown"
    return "monthly" if med <= 45 else "quarterly" if med <= 135 else "annual"


def _stream_stats(flows: list[tuple[_dt.date, float]]):
    irr = xirr(flows)
    if irr is None:
        return None
    infl = sum(v for _, v in flows if v > 0)
    outf = sum(v for _, v in flows if v < 0)
    em = (infl / abs(outf)) if outf else None
    return irr, em, min(v for _, v in flows)


def find_streams(grids: dict[str, list[tuple]]) -> list[Stream]:
    """Every dated, sign-changing numeric series in the workbook, both
    orientations. Permissive on shape — the IRR match is the real filter, so we
    don't pre-judge which series 'looks like' a cash flow."""
    streams: list[Stream] = []
    for sheet, grid in grids.items():
        nrows = len(grid)
        ncols = max((len(r) for r in grid), default=0)

        # --- horizontal axes: a row of dates; each other row is a series -------
        for ar in range(nrows):
            run = _date_run(list(grid[ar]))
            if not run:
                continue
            cols = [i for i, _ in run]
            dates = [d for _, d in run]
            for r in range(nrows):
                if r == ar:
                    continue
                vals = [_num(_cell(grid, r, c)) for c in cols]
                flows = [(dates[i], vals[i]) for i in range(len(cols)) if vals[i] is not None]
                if len(flows) < _MIN_PERIODS:
                    continue
                st = _stream_stats(flows)
                if st is None:
                    continue
                irr, em, init = st
                label = _label_left(grid, r, cols[0]) or ""
                streams.append(Stream(sheet, "row", r, label.strip()[:40], flows,
                                      irr, em, init, len(flows), f"{sheet}!row{r + 1}"))

        # --- vertical axes: a column of dates; each other column is a series ----
        for ac in range(ncols):
            colcells = [_cell(grid, r, ac) for r in range(nrows)]
            run = _date_run(colcells)
            if not run:
                continue
            rows = [i for i, _ in run]
            dates = [d for _, d in run]
            for c in range(ncols):
                if c == ac:
                    continue
                vals = [_num(_cell(grid, r, c)) for r in rows]
                flows = [(dates[i], vals[i]) for i in range(len(rows)) if vals[i] is not None]
                if len(flows) < _MIN_PERIODS:
                    continue
                st = _stream_stats(flows)
                if st is None:
                    continue
                irr, em, init = st
                label = _label_above(grid, rows[0], c) or ""
                from openpyxl.utils import get_column_letter
                streams.append(Stream(sheet, "col", c, label.strip()[:40], flows,
                                      irr, em, init, len(flows),
                                      f"{sheet}!col {get_column_letter(c + 1)}"))
    return streams


# ---------------------------------------------------------------------------
# The spine: match streams to stated IRRs
# ---------------------------------------------------------------------------

@dataclass
class Spine:
    ok: bool
    file: str
    stated_irrs: list[dict] = field(default_factory=list)
    matched: dict[str, dict] = field(default_factory=dict)   # leg -> matched stream
    n_candidate_streams: int = 0
    diagnostics: dict = field(default_factory=dict)


def find_spine(file_path: str | Path) -> Spine:
    file_path = Path(file_path)
    grids = _load_grids(file_path)
    n_numeric = sum(1 for g in grids.values() for row in g for v in row
                    if isinstance(v, (int, float)) and not isinstance(v, bool))
    if not grids or n_numeric < 50:
        # Formulas saved without computed results read as blank under data_only.
        return Spine(ok=False, file=file_path.name,
                     diagnostics={"reason": "no cached values — re-save the file in Excel "
                                            "so formula results are stored",
                                  "n_sheets": len(grids), "n_numeric_cells": n_numeric})
    stated = find_stated_irrs(grids)
    streams = find_streams(grids)

    tol = _IRR_MATCH_BPS / 10000.0
    # For each stated IRR, the stream whose recomputed IRR is closest within tol.
    matched: dict[str, dict] = {}
    used_keys: set = set()
    for s in sorted(stated, key=lambda x: -abs(x["value"])):
        best, best_err = None, None
        for st in streams:
            err = abs(st.irr - s["value"])
            if best is None or err < best_err:
                best, best_err = st, err
        if best is None or best_err > tol:
            continue
        leg = s["leg"] if s["leg"] != "unknown" else "primary"
        # don't overwrite a leg already matched by a closer stated value
        if leg in matched and matched[leg]["match_err_bps"] <= best_err * 10000:
            continue
        key = (best.sheet, best.orientation, best.index)
        matched[leg] = {
            "leg": leg, "sheet": best.sheet, "orientation": best.orientation,
            "loc": best.loc, "label": best.label,
            "recomputed_irr": round(best.irr, 4), "recomputed_em": round(best.em, 2) if best.em else None,
            "stated_irr": round(s["value"], 4), "stated_cell": f"{s['sheet']}!{s['cell']}",
            "match_err_bps": round(best_err * 10000, 1),
            "n_periods": best.n, "initial_outflow": best.initial_outflow,
            "period_start": best.flows[0][0].isoformat(), "period_end": best.flows[-1][0].isoformat(),
            "periodicity": periodicity_of(best.flows),
            "flows": best.flows,          # retained in-process for the deal-truth anchor
            "_key": key,
        }
        used_keys.add(key)

    # Drop the unknown-leg "primary" entry when it resolves to the same physical
    # series as a labelled leg (it's a duplicate of levered/unlevered).
    if "primary" in matched:
        pk = matched["primary"]["_key"]
        if any(matched[l]["_key"] == pk for l in ("levered", "unlevered") if l in matched):
            matched.pop("primary")
    # If two legs resolved to the same physical series, demote the duplicate.
    if {"levered", "unlevered"} <= set(matched):
        if matched["levered"]["_key"] == matched["unlevered"]["_key"]:
            matched.pop("unlevered")
    ok = bool(matched)

    spine = Spine(
        ok=ok, file=file_path.name, stated_irrs=stated, matched=matched,
        n_candidate_streams=len(streams),
        diagnostics={
            "n_sheets": len(grids),
            "n_stated_irrs": len(stated),
            "anchor_sheets": sorted({m["sheet"] for m in matched.values()}),
        },
    )
    log.info("SPINE %s — ok=%s, anchor=%s, %d stated IRR(s), %d candidate stream(s)",
             file_path.name, ok, spine.diagnostics["anchor_sheets"],
             len(stated), len(streams))
    return spine


def render_spine(sp: Spine) -> str:
    L = [f"SPINE — {sp.file}   ok={sp.ok}"]
    L.append(f"  stated IRRs: " + (", ".join(f"{s['value']*100:.2f}% [{s['leg']}] ({s['sheet']}!{s['cell']})"
                                              for s in sp.stated_irrs[:6]) or "none"))
    L.append(f"  candidate streams scanned: {sp.n_candidate_streams}")
    if sp.matched:
        for leg, m in sp.matched.items():
            ok = "✓" if m["match_err_bps"] <= _IRR_MATCH_BPS else "✗"
            L.append(f"  {ok} {leg:<9} {m['loc']:<22} '{m['label']}'  "
                     f"XIRR {m['recomputed_irr']*100:.2f}% = stated {m['stated_irr']*100:.2f}% "
                     f"({m['match_err_bps']} bps)  EM {m['recomputed_em']}  "
                     f"[{m['period_start']}→{m['period_end']}, n={m['n_periods']}]")
    else:
        L.append("  NO STREAM VALIDATED — cash flow NOT reconstructed (ok=False)")
    return "\n".join(L)


# ---------------------------------------------------------------------------
# Corpus harness — validate across all models at once (self-consistency)
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    args = sys.argv[1:]
    targets: list[Path] = []
    for a in args:
        p = Path(a)
        if p.is_dir():
            targets += sorted(p.glob("*.xlsx")) + sorted(p.glob("*.xlsm"))
        elif p.exists():
            targets.append(p)
    targets = [t for t in targets if not t.name.startswith("~$")]   # skip Excel lock files
    if not targets:
        print("usage: python cashflow_spine.py <dir-or-files>")
        raise SystemExit(1)

    n_ok = 0
    rows = []
    for t in targets:
        try:
            sp = find_spine(t)
        except Exception as e:
            rows.append((t.name, "ERROR", str(e)[:50], ""))
            continue
        n_ok += 1 if sp.ok else 0
        if sp.ok:
            legs = " | ".join(
                f"{leg}:{m['recomputed_irr']*100:.1f}%vs{m['stated_irr']*100:.1f}%({m['match_err_bps']:.0f}bp)@{m['sheet']}"
                for leg, m in sp.matched.items())
            rows.append((t.name, "OK", legs, ""))
        else:
            why = sp.diagnostics.get("reason") or (
                f"{len(sp.stated_irrs)} stated IRR(s), {sp.n_candidate_streams} streams, none matched")
            rows.append((t.name, "MISS", why, ""))

    print("\n" + "=" * 100)
    print(f"CORPUS SELF-CONSISTENCY: {n_ok}/{len(targets)} files had a stream reproduce their own stated IRR")
    print("=" * 100)
    for name, status, detail, _ in rows:
        print(f"  [{status:<5}] {name[:46]:<46} {detail}")
