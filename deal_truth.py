"""
deal_truth.py — Slice 2 of the deal-reconstruction rebuild (2026-06-15).

Consumes the deterministic workbook map (workbook_map.py) and turns candidate
facts into ONE canonical deal truth — still no GPT, still no narrative. The job:

    1. CASH-FLOW ORACLE — detect the levered & unlevered cash-flow streams,
       recompute IRR (XIRR) and equity multiple, and VALIDATE them against the
       returns the model itself prints. The stream that reproduces the stated
       IRR is the canonical one, regardless of its label — "the model grades the
       parse." This binds levered/unlevered authoritatively (fixing the matrix-
       header ambiguity the map leaves open) and is the integrity check on the
       whole read.

    2. CANONICAL RECONCILIATION — for each non-negotiable concept, pick the value
       from the source APPROPRIATE TO THAT CONCEPT (inputs for the capital stack,
       the operating pro forma for NOI, the validated stream for returns/exit
       proceeds), using formula provenance so a summary cell that merely DISPLAYS
       a number collapses into corroboration, not a competing truth. Genuine
       cross-source disagreements are recorded as conflicts with a winner.

    3. IDENTITY CHECKS — the deal must hang together: Debt+Equity≈Cost,
       ExitValue≈ForwardNOI/ExitCap, recomputed IRR≈stated IRR.

    4. UNSUPPORTED-CONCLUSION GUARDRAILS — generated DYNAMICALLY from the above
       validation results (missing metric / source conflict / display-vs-source /
       label-vs-formula / unverified exit / non-standard multiple / deal type /
       missing debt schedule / unvalidated returns). Never hard-coded per file:
       a guardrail fires only when its evidence is detected in THIS workbook.
       This is the contract a later GPT layer must obey.

No per-file logic. Judged on whether it survives acquisition / development /
refi / value-add / mixed-use / weird-name models — 1425 is just one test.
"""
from __future__ import annotations

import datetime as _dt
import logging
import sys
from pathlib import Path
from typing import Any

from workbook_map import build_workbook_map, _passes_domain, _concept_of
from financial_model_parser import parse_workbook_tables_cached
from cashflow_spine import find_spine

log = logging.getLogger("fb.dealtruth")
if not log.handlers:
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("[fb.truth] %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

DEAL_TRUTH_VERSION = "2026-06-15.1"

_IRR_TOL = 0.005          # 50 bps: recomputed XIRR vs stated IRR
_REL_TOL = 0.05           # 5% relative tolerance for value/identity reconciliation


# ---------------------------------------------------------------------------
# XIRR (deterministic; bisection with a Newton refine — never diverges silently)
# ---------------------------------------------------------------------------

def _xnpv(rate: float, flows: list[tuple[_dt.date, float]]) -> float:
    t0 = flows[0][0]
    return sum(f / (1.0 + rate) ** ((d - t0).days / 365.0) for d, f in flows)


def xirr(flows: list[tuple[_dt.date, float]]) -> float | None:
    """Annualized IRR of dated flows. Requires a sign change; bracketed so it
    can't run away. Returns None when no real root exists in a sane band."""
    if len(flows) < 2:
        return None
    pos = any(f > 0 for _, f in flows)
    neg = any(f < 0 for _, f in flows)
    if not (pos and neg):
        return None
    lo, hi = -0.95, 5.0
    flo, fhi = _xnpv(lo, flows), _xnpv(hi, flows)
    if flo == 0:
        return lo
    if flo * fhi > 0:
        return None                      # no sign change in the band
    for _ in range(200):
        mid = (lo + hi) / 2.0
        fm = _xnpv(mid, flows)
        if abs(fm) < 1.0:
            return mid
        if flo * fm < 0:
            hi = mid
        else:
            lo, flo = mid, fm
    return (lo + hi) / 2.0


def _iso(s: Any) -> _dt.date | None:
    try:
        return _dt.date.fromisoformat(str(s))
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 1. Cash-flow oracle
# ---------------------------------------------------------------------------

def _candidate_streams(tables: list[dict]) -> list[dict]:
    """Every cash-flow-like row across cashflow tables, as dated flow streams.
    A candidate must be a real investment stream: enough periods AND a sign
    change (an outflow then inflows)."""
    streams: list[dict] = []
    for t in tables:
        if t.get("table_type") not in ("cashflow_rollup", "monthly_table", "annual_table"):
            continue
        for r in t.get("rows", []):
            label = (r.get("label") or "").strip()
            concept = _concept_of(label)
            looks_cf = concept in ("levered_cf", "unlevered_cf") or "cash flow" in label.lower()
            if not looks_cf:
                continue
            vbp = r.get("values_by_period") or {}
            flows = sorted((_iso(d), v) for d, v in vbp.items()
                           if _iso(d) and isinstance(v, (int, float)) and not isinstance(v, bool))
            if len(flows) < 4:
                continue
            if not (any(f > 0 for _, f in flows) and any(f < 0 for _, f in flows)):
                continue
            x = xirr(flows)
            if x is None:
                continue
            infl = sum(v for _, v in flows if v > 0)
            outf = sum(v for _, v in flows if v < 0)
            streams.append({
                "sheet": t["sheet"], "row": r.get("row"), "label": label,
                "concept": concept, "periodicity": t.get("periodicity"),
                "flows": flows, "xirr": x,
                "em": (infl / abs(outf)) if outf else None,
                "initial_outflow": min((f for _, f in flows), default=0.0),
                "n_periods": len(flows),
            })
    return streams


def _leg_from_label(label: str) -> str | None:
    """Levered / unlevered from a label's own words (reliable at the row level)."""
    l = label.lower()
    if "unlever" in l or "un-lever" in l or "unleveraged" in l:
        return "unlevered"
    if "lever" in l or "equity" in l:        # 'leveraged', 'equity net cash flow'
        return "levered"
    return None


def run_oracle(m: dict, tables: list[dict]) -> dict:
    """Validate returns by recomputing them from the cash-flow streams.

    Returns {"levered": {...}|None, "unlevered": {...}|None, "candidates": [...]}.
    For each leg: the canonical stream (closest XIRR to the stated IRR), the
    recomputed irr/em, the stated value it matched, and `validated`."""
    streams = _candidate_streams(tables)

    # All stated IRRs the model prints, leg-AGNOSTIC (fraction form, no ~0 noise).
    # We validate against these but take the LEG from each stream's own row label,
    # which is reliable even when a summary's IRR cell was mis-binned upstream.
    stated_vals: list[float] = []
    for concept in ("levered_irr", "unlevered_irr"):
        for e in m["candidates"].get(concept, []):
            v = float(e["value"])
            if -0.5 <= v <= 1.5 and abs(v) >= 1e-4:
                stated_vals.append(v)

    out: dict[str, Any] = {"candidates": streams, "levered": None, "unlevered": None}
    for leg in ("levered", "unlevered"):
        leg_streams = [s for s in streams if _leg_from_label(s["label"]) == leg]
        if not leg_streams or not stated_vals:
            continue
        # the leg-stream whose XIRR best matches ANY stated IRR
        best, best_err, target = None, None, None
        for s in leg_streams:
            closest = min(stated_vals, key=lambda t: abs(s["xirr"] - t))
            err = abs(s["xirr"] - closest)
            if best is None or err < best_err:
                best, best_err, target = s, err, closest
        out[leg] = {
            "sheet": best["sheet"], "row": best["row"], "label": best["label"],
            "recomputed_irr": round(best["xirr"], 4),
            "recomputed_em": round(best["em"], 2) if best["em"] is not None else None,
            "stated_irr": round(target, 4),
            "validated": best_err <= _IRR_TOL,
            "match_err_bps": round(best_err * 10000, 1),
            "initial_outflow": best["initial_outflow"],
            "flows": best["flows"],
            "periodicity": best["periodicity"],
        }

    # Consistency: the unlevered stream invests more equity-equivalent capital
    # (no debt), so |unlevered initial outflow| should exceed |levered|.
    lev, unlev = out.get("levered"), out.get("unlevered")
    if lev and unlev:
        out["leg_order_ok"] = abs(unlev["initial_outflow"]) >= abs(lev["initial_outflow"])
    return out


# ---------------------------------------------------------------------------
# 2. Hold period — from the ACTUAL sale event, not the model length
# ---------------------------------------------------------------------------

def _months_between(a: _dt.date, b: _dt.date) -> int:
    return (b.year - a.year) * 12 + (b.month - a.month)


def detect_hold(oracle: dict) -> dict | None:
    """Hold = month 0 → the period the sale is logged, read from the cash flow.
    NOT the model's total length: a 120-month model may sell at month 61. The
    sale is the terminal-scale inflow (it dwarfs operating flows); we take the
    LAST such period (a refi inflow earlier won't be mistaken for the sale)."""
    stream = oracle.get("levered") or oracle.get("unlevered")
    if not stream or not stream.get("flows"):
        return None
    flows = stream["flows"]
    start = flows[0][0]
    pos = [v for _, v in flows if v > 0]
    if not pos:
        return None
    import statistics
    big, med = max(pos), (statistics.median(pos) if pos else 0.0)
    threshold = max(big * 0.5, med * 5)          # terminal-scale inflow
    sale_idx = next((i for i in range(len(flows) - 1, -1, -1)
                     if flows[i][1] >= threshold), None)
    if sale_idx is None:
        sale_idx = next((i for i in range(len(flows) - 1, -1, -1)
                         if abs(flows[i][1]) > 1.0), None)
    if sale_idx is None:
        return None
    sale_date = flows[sale_idx][0]
    months = _months_between(start, sale_date)
    model_months = _months_between(start, flows[-1][0])
    return {
        "months": months, "years": round(months / 12.0, 2),
        "sale_date": sale_date.isoformat(), "start_date": start.isoformat(),
        "model_months": model_months,
        "sells_before_model_end": months < model_months - 1,
        "source": f"{stream.get('loc') or stream.get('sheet')} (sale period detected from cash flow)",
    }


# ---------------------------------------------------------------------------
# 3. Operating trajectory (NOI / revenue / opex) — tiered, with provenance
#    Tier 1  operating-model rows (reject all-zero) → "operating_model"
#    Tier 1b derived from the validated unlevered stream → "unlevered_stream"
#    Tier 2  summary point, formula-traced + identity cross-checked → "summary_*"
# ---------------------------------------------------------------------------

def _operating_tier1_tables(m: dict, tables: list[dict]) -> dict:
    """Operating rows from the PARSED tables. Sums component rows within one table
    (mixed-use). Rejects all-zero rows — the reason the obvious rows fail on some
    models (cached as zeros / misaligned with the cash-flow date axis)."""
    engine = m.get("cashflow_engine")
    model_sheets = {s for s, i in m["sheets"].items() if i["role"] == "model"}

    def trank(t: dict) -> tuple:
        return (0 if t["sheet"] == engine else 1,
                0 if t.get("periodicity") == "monthly" else 1)

    cand = sorted([t for t in tables if t["sheet"] == engine or t["sheet"] in model_sheets],
                  key=trank)
    out: dict[str, Any] = {}
    for concept in ("noi", "revenue", "opex"):
        for t in cand:
            rows = [r for r in t.get("rows", []) if _concept_of(r.get("label") or "") == concept]
            if not rows:
                continue
            dmap: dict[_dt.date, float] = {}
            for r in rows:
                for d, v in (r.get("values_by_period") or {}).items():
                    dd = _iso(d)
                    if dd and isinstance(v, (int, float)) and not isinstance(v, bool):
                        dmap[dd] = dmap.get(dd, 0.0) + v
            by_year: dict[int, float] = {}
            for d, v in dmap.items():
                by_year[d.year] = by_year.get(d.year, 0.0) + v
            bylist = sorted(by_year.items())
            nz = [(y, v) for y, v in bylist if abs(v) > 1.0]
            if not nz:                            # reject all-zero rows
                continue
            out[concept] = {
                "by_year": bylist, "going_in": nz[0][1], "terminal": nz[-1][1],
                "stabilized": max(v for _, v in nz),
                "source": f"{t['sheet']}!rows {[r.get('row') for r in rows]}",
                "provenance": "operating_model",
            }
            break
    return out


def _scan_model_sheets_for_ops(m: dict, file_path: Path) -> dict:
    """Gap-immune Tier 1: read operating rows DIRECTLY from model-role sheets the
    generic table parser misses (formatted pro formas with section gaps stop the
    parser's body scan). Finds the period header, then scans the whole sheet for
    a CONSOLIDATED concept row — picking the single best row (most periods, then
    largest magnitude: a total dominates its components), never summing (which
    would double-count 'before/after' stages and component+total rows)."""
    import openpyxl
    from financial_model_parser import _detect_header, _gcell
    engine = m.get("cashflow_engine")
    sheets = [s for s, i in m["sheets"].items() if i["role"] == "model"]
    if engine in sheets:                          # engine first
        sheets = [engine] + [s for s in sheets if s != engine]
    out: dict[str, Any] = {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return out
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        if not ws.max_row or ws.max_row > 600:
            continue
        grid = [r for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200),
                                        max_col=min(ws.max_column or 1, 140), values_only=True)]
        hdr = hrow = None
        for r in range(1, len(grid) + 1):
            h = _detect_header(grid, r)
            if h:
                hdr, hrow = h, r
                break
        if not hdr:
            continue
        pcols, dh = hdr["period_cols"], hdr["date_headers"]
        rows_by: dict[str, list] = {}
        for r in range(hrow + 1, len(grid) + 1):
            label = next((str(_gcell(grid, r, c)).strip() for c in range(1, pcols[0])
                          if isinstance(_gcell(grid, r, c), str) and _gcell(grid, r, c).strip()),
                         None)
            if not label:
                continue
            concept = _concept_of(label)
            if concept not in ("noi", "revenue", "opex"):
                continue
            yr: dict[int, float] = {}
            for i, c in enumerate(pcols):
                v = _gcell(grid, r, c)
                if isinstance(v, (int, float)) and not isinstance(v, bool) and i < len(dh):
                    dd = _iso(dh[i])
                    if dd:
                        yr[dd.year] = yr.get(dd.year, 0.0) + v
            nz = {y: v for y, v in yr.items() if abs(v) > 1.0}
            if nz:
                rows_by.setdefault(concept, []).append(
                    (label, r, yr, len(nz), sum(abs(v) for v in nz.values())))
        for concept, rows in rows_by.items():
            if concept in out:
                continue
            label, r, yr, _, _ = max(rows, key=lambda t: (t[3], t[4]))
            bylist = sorted(yr.items())
            nz = [(y, v) for y, v in bylist if abs(v) > 1.0]
            out[concept] = {
                "by_year": bylist, "going_in": nz[0][1], "terminal": nz[-1][1],
                "stabilized": max(v for _, v in nz),
                "source": f"{sheet}!row{r} '{label[:30]}'", "provenance": "operating_model",
            }
    try:
        wb.close()
    except Exception:
        pass
    return out


def _operating_tier1(m: dict, tables: list[dict], file_path: Path) -> dict:
    """Tier 1 = parsed tables first, then a gap-immune direct sheet scan for any
    concept the parser missed (catches formatted pro formas)."""
    out = _operating_tier1_tables(m, tables)
    for concept, traj in _scan_model_sheets_for_ops(m, file_path).items():
        out.setdefault(concept, traj)
    return out


def _noi_reconciles(noi_terminal, canonical: dict) -> bool | None:
    """Does a terminal NOI tie to the deal exit (NOI / exit_cap ≈ exit value)?
    Used to accept a Tier-1 operating-model NOI only when it's the CONSOLIDATED
    deal NOI — a component-only row (e.g. one asset's income) won't reconcile."""
    ec, ev = canonical.get("exit_cap"), canonical.get("exit_value")
    if not (ec and ev and noi_terminal):
        return None
    c, v = float(ec["value"]), float(ev["value"])
    cf = c / 100.0 if c > 1.5 else c
    if cf <= 0:
        return None
    return abs(noi_terminal / cf - v) / max(abs(v), 1e-9) <= 0.15


def _operating_from_stream(oracle: dict, canonical: dict, hold: dict | None) -> dict | None:
    """Tier 1b: derive the operating trajectory from the UNLEVERED stream — strip
    the construction draws (early negatives) and back out the net sale proceeds
    from the disposition year. The remainder is operating cash flow (≈ NOI net of
    reserves), genuinely cash-flow-sourced. Prefers an ANNUAL stream — full-year
    values avoid the partial-calendar-year skew a monthly stream introduces."""
    cands = [s for s in (oracle.get("candidates") or [])
             if _leg_from_label(s.get("label", "")) == "unlevered"]
    annual = [s for s in cands if s.get("periodicity") == "annual"]
    stream = (annual or cands)[0] if (annual or cands) else oracle.get("unlevered")
    if not stream or not stream.get("flows"):
        return None

    by_year: dict[int, float] = {}
    for d, v in stream["flows"]:
        by_year[d.year] = by_year.get(d.year, 0.0) + v
    bylist = sorted(by_year.items())

    # Back out the sale from its year. The unlevered stream carries the proceeds
    # net of selling costs (no debt), so subtract the NET exit value.
    sale_year = int(hold["sale_date"][:4]) if hold and hold.get("sale_date") else None
    sale_px = None
    for key in ("exit_value", "sale_price"):
        if key in canonical:
            try:
                sale_px = float(canonical[key]["value"])
            except (TypeError, ValueError):
                pass
            break
    op = [(y, (v - sale_px if (sale_year and y == sale_year and sale_px) else v))
          for y, v in bylist]
    # Use only FULL operating years for the headline figures — a partial first/
    # last calendar year (a mid-year start or sale) understates NOI. (For an
    # annual stream every year is full, so this is a no-op there.)
    yr_count: dict[int, int] = {}
    for d, _ in stream["flows"]:
        yr_count[d.year] = yr_count.get(d.year, 0) + 1
    mx = max(yr_count.values(), default=1)
    full = {y for y, c in yr_count.items() if c >= max(1, mx - 2)}
    pos = [(y, v) for y, v in op if v > 0 and y in full]   # full operating years
    if not pos:
        pos = [(y, v) for y, v in op if v > 0]
    if not pos:
        return None
    return {
        "by_year": op, "going_in": pos[0][1], "terminal": pos[-1][1],
        "stabilized": max(v for _, v in pos),
        "source": f"{stream['sheet']}!row{stream.get('row')} "
                  f"(unlevered {stream.get('periodicity')} stream, sale backed out ≈ NOI net of reserves)",
        "provenance": "unlevered_stream",
    }


def operating_trajectory(m: dict, tables: list[dict], oracle: dict,
                         canonical: dict, hold: dict | None, file_path: Path) -> dict:
    """Operating NOI/revenue/opex with explicit provenance per the tier policy:
    operating model first; else derive from the unlevered stream; else fall back
    to the summary point (formula-traced + reconciled), clearly labelled."""
    out = _operating_tier1(m, tables, file_path)
    # Reconcile guard: a Tier-1 NOI is only trusted if it ties to the deal exit.
    # A component-only NOI (one asset of a mixed-use deal) won't reconcile — drop
    # it so the consolidated stream-derived NOI (Tier 1b) wins instead.
    if "noi" in out and _noi_reconciles(out["noi"].get("terminal"), canonical) is False:
        out.pop("noi")
    if "noi" not in out:                          # Tier 1b — derive NOI from the stream
        derived = _operating_from_stream(oracle, canonical, hold)
        if derived:
            out["noi"] = derived
    if "noi" not in out and "noi" in canonical:   # Tier 2 — summary point, reconciled
        info = canonical["noi"]
        prov = info.get("provenance") or {}
        traces = bool(prov.get("crosses_sheet"))
        out["noi"] = {
            "by_year": None, "going_in": None,
            "terminal": float(info["value"]), "stabilized": float(info["value"]),
            "source": info.get("source"),
            "provenance": "summary_traced" if traces else "summary_point",
        }
    return out


# ---------------------------------------------------------------------------
# 3. Per-concept reconciliation → canonical winner
# ---------------------------------------------------------------------------

# Source role preference per concept (concept-appropriate truth).
_SOURCE_PREF = {
    "purchase_price": ("inputs", "summary", "support", "model"),
    "total_cost":     ("inputs", "summary", "model", "support"),
    "debt":           ("inputs", "summary", "model"),
    "equity":         ("inputs", "summary", "model"),
    "ltc":            ("inputs", "summary"),
    "ltv":            ("inputs", "summary"),
    "exit_cap":       ("inputs", "summary"),
    "going_in_cap":   ("inputs", "summary"),
    "interest_rate":  ("inputs", "model"),
    "debt_yield":     ("inputs", "summary", "model"),
    "dscr":           ("inputs", "model", "summary"),
    "exit_value":     ("summary", "model", "support"),
    "yield_on_cost":  ("summary", "returns", "model"),
}


def _canonical_for_concept(concept: str, cands: list[dict],
                           referenced: set[str]) -> dict | None:
    """Pick the canonical candidate for a non-stream concept, and detect conflict.

    Collapses display→source duplicates via provenance (a summary cell and the
    input it links to are ONE fact), prefers the concept-appropriate role, and —
    the key for picking the value the model actually USES — prefers an input that
    is REFERENCED by a display elsewhere over an unreferenced lookalike. Flags a
    conflict when genuinely distinct magnitudes remain."""
    usable = [c for c in cands if _passes_domain(concept, c["value"]) and not c.get("series")]
    if not usable:
        usable = [c for c in cands if _passes_domain(concept, c["value"])]
    if not usable:
        return None

    # Collapse by formula source. Keep the REAL source cell as representative
    # (not the display), and mark it referenced if a display pointed to it.
    groups_by_src: dict[str, dict] = {}
    for c in usable:
        prov = c.get("provenance") or {}
        a1 = f"{c['sheet']}!{c.get('cell')}"
        key = prov.get("source") or a1
        g = groups_by_src.setdefault(key, {"rep": None, "members": [], "referenced": False})
        g["members"].append(c)
        if a1 == key or g["rep"] is None:        # prefer the source cell itself
            g["rep"] = c
        if a1 != key:                            # this candidate DISPLAYS the source
            g["referenced"] = True
    for key, g in groups_by_src.items():
        if key in referenced:
            g["referenced"] = True
    reps = [(g["rep"], g["referenced"]) for g in groups_by_src.values()]

    # Distinct magnitudes (beyond tolerance).
    mags: list[float] = []
    for rep, _ in reps:
        v = float(rep["value"])
        if not any(abs(v - mm) / max(abs(v), abs(mm), 1e-9) <= _REL_TOL for mm in mags):
            mags.append(v)

    pref = _SOURCE_PREF.get(concept, ("inputs", "summary", "returns", "model", "support"))
    rep_vals = [float(r["value"]) for r, _ in reps]

    def support(v: float) -> int:
        return sum(1 for vv in rep_vals if abs(v - vv) / max(abs(v), abs(vv), 1e-9) <= _REL_TOL)

    def rank(item: tuple) -> tuple:
        c, ref = item
        r = c.get("role")
        role_rank = pref.index(r) if r in pref else len(pref)
        # role-appropriate first; then the value the most sources agree on; then
        # the input a display actually references; then magnitude.
        return (role_rank, -support(float(c["value"])), 0 if ref else 1, -abs(float(c["value"])))

    winner, _ = min(reps, key=rank)

    # A conflict is real only if a competing value isn't the winner AND the
    # competitors don't DECOMPOSE it: total equity = LP + GP, total cost = a sum
    # of line items — a part-of-whole split, not a disagreement. Components are
    # same-signed and smaller than the whole; they must sum to it.
    w = float(winner["value"])
    others = [mm for mm in mags if abs(mm - w) / max(abs(mm), abs(w), 1e-9) > _REL_TOL]
    comps = [o for o in others if o * w > 0 and abs(o) < abs(w)]
    part_of_whole = bool(comps) and abs(sum(comps) - w) / max(abs(w), 1e-9) <= _REL_TOL
    conflict = bool(others) and not part_of_whole
    return {
        "concept": concept,
        "value": winner["value"],
        "source": (winner.get("provenance") or {}).get("source") or winner["display"],
        "display": winner["display"],
        "role": winner.get("role"),
        "provenance": winner.get("provenance"),
        "conflict": conflict,
        "alternatives": ([{"value": rep["value"], "source":
                           (rep.get("provenance") or {}).get("source") or rep["display"]}
                          for rep, _ in reps] if conflict else None),
    }


# ---------------------------------------------------------------------------
# Deal-type inference (from cash-flow shape, not names)
# ---------------------------------------------------------------------------

def _infer_deal_type(oracle: dict, ops: dict) -> str:
    """development if the cash flow shows a CONSTRUCTION RUN (a stretch of early
    outflows before income), acquisition if income starts ~immediately. The run
    of leading outflows in the unlevered stream is the robust signal (a single
    acquisition outflow vs many construction draws); NOI ramp is the backup."""
    stream = oracle.get("unlevered") or oracle.get("levered")
    if stream and stream.get("flows"):
        neg_lead = 0
        for _, v in stream["flows"]:
            if v < 0:
                neg_lead += 1
            elif v > 0:
                break
        # monthly: a multi-month draw period; annual: ≥2 outflow years.
        thresh = 3 if stream.get("periodicity") == "monthly" else 1
        if neg_lead > thresh:
            return "development"
    noi = ops.get("noi")
    if noi and noi.get("by_year"):
        ys = [v for _, v in noi["by_year"] if v > 0]
        if ys:
            mx = max(ys)
            ramp = sum(1 for v in ys if v < 0.5 * mx)
            return "development" if ramp >= max(1, 0.25 * len(ys)) else "acquisition"
    return "acquisition" if stream else "unknown"


# ---------------------------------------------------------------------------
# 4. Guardrails — generated dynamically from validation results
# ---------------------------------------------------------------------------

_NON_NEGOTIABLE = ("purchase_price", "total_cost", "exit_value", "exit_cap",
                   "noi", "levered_irr")


def _build_guardrails(canonical: dict, oracle: dict, identities: list[dict],
                      ops: dict, deal_type: str, m: dict) -> list[dict]:
    g: list[dict] = []

    def add(code, message, evidence=None):
        g.append({"code": code, "message": message, "evidence": evidence})

    # (1) missing non-negotiables
    for c in _NON_NEGOTIABLE:
        present = c in canonical or (c == "levered_irr" and oracle.get("levered"))
        if c == "noi":
            present = bool(ops.get("noi")) or "noi" in canonical
        if not present:
            add("missing_metric",
                f"{c} not found in this workbook — GPT may say 'not found' but must "
                f"NOT infer strategy or risk from its absence.", {"concept": c})

    # (2) source conflicts
    for c, info in canonical.items():
        if info.get("conflict"):
            add("source_conflict",
                f"Sources disagree on {c}; use ONLY the canonical winner "
                f"{info['value']} from {info['source']}, and disclose the conflict.",
                {"concept": c, "winner": info["value"], "alternatives": info["alternatives"]})

    # (3) display vs underlying source (formula-traced)
    for c, info in canonical.items():
        prov = info.get("provenance") or {}
        if prov.get("crosses_sheet"):
            add("display_vs_source",
                f"{c} is shown at {prov.get('display')} but computed at "
                f"{prov.get('source')}; treat the formula source as the underlying "
                f"value and the summary cell as display only.",
                {"concept": c, "display": prov.get("display"), "source": prov.get("source")})

    # (4) cash-flow total must not be read as a price/exit value
    for e in m["candidates"].get("levered_cf", []):
        if (e.get("provenance") or {}).get("op") == "SUM" and abs(float(e["value"])) > 1e6:
            add("cashflow_total_not_price",
                f"{e['display']} = {e['value']} is a SUM of cash flows (a total/"
                f"profit), NOT a price or exit value — do not present it as one.",
                {"cell": e["display"]})
            break

    # (5) exit value not tied to the NOI / exit-cap identity
    exit_id = next((i for i in identities if i["name"] == "exit_value≈NOI/exit_cap"), None)
    if exit_id and exit_id.get("checked") and not exit_id["passed"]:
        add("exit_unverified",
            "Exit value does not reconcile with terminal NOI / exit cap; GPT may "
            "not assert it as a clean exit value without flagging the mismatch.",
            exit_id)

    # (6) non-standard / model-stated multiple
    lev = oracle.get("levered")
    if lev and lev.get("recomputed_em") is not None:
        stated_em = next((float(e["value"]) for e in m["candidates"].get("equity_multiple", [])
                          if 0.1 <= float(e["value"]) <= 25), None)
        if stated_em is not None and abs(stated_em - lev["recomputed_em"]) > 0.1:
            add("nonstandard_multiple",
                f"Stated equity multiple ({stated_em}x) differs from the cash-flow "
                f"multiple ({lev['recomputed_em']}x); label stated multiples as "
                f"model-defined / non-standard, not institutional-standard.",
                {"stated": stated_em, "cashflow": lev["recomputed_em"]})

    # (7) development → yield-on-cost language
    if deal_type == "development":
        add("development_language",
            "Deal type is development (cash flow ramps from construction/lease-up); "
            "prefer yield-on-cost / ROC language over going-in cap unless an "
            "acquisition cap is explicitly supported.", {"deal_type": deal_type})

    # (8/9) debt service / DSCR coverage
    has_ds = any(b["kind"] == "debt_service" for b in m["timeseries_blocks"]) \
        or "debt_service" in canonical
    if not has_ds:
        add("no_debt_schedule",
            "No debt-service schedule detected; GPT may not compute DSCR or infer "
            "debt-coverage risk from the loan amount alone.", None)

    # (10) returns not validated against cash flow
    if not (lev and lev.get("validated")):
        add("returns_unvalidated",
            "Levered IRR/EM did not validate against a cash-flow stream; present "
            "returns as model-stated and unverified.",
            {"levered": lev})
    return g


# ---------------------------------------------------------------------------
# Identity checks
# ---------------------------------------------------------------------------

def _run_identities(canonical: dict, oracle: dict, ops: dict) -> list[dict]:
    out: list[dict] = []

    def val(c):
        return float(canonical[c]["value"]) if c in canonical else None

    def approx(a, b):
        denom = max(abs(a), abs(b), 1e-9)
        return abs(a - b) / denom <= 0.12

    debt, equity, cost = val("debt"), val("equity"), val("total_cost")
    if None not in (debt, equity, cost):
        out.append({"name": "debt+equity≈cost", "checked": True,
                    "passed": approx(debt + equity, cost),
                    "detail": f"{debt:,.0f} + {equity:,.0f} vs {cost:,.0f}"})

    # Cross-check the capital outlay against the cash flow ITSELF: total cost vs
    # the cumulative outflows in the unlevered stream; acquisition cost vs the
    # first-period outflow (land/purchase at close). Promotes a costed input to a
    # cash-flow-validated fact when they agree.
    unlev = oracle.get("unlevered")
    flows = unlev.get("flows") if unlev else None
    if flows and cost:
        draws = -sum(v for _, v in flows if v < 0)
        out.append({"name": "total_cost≈CF draws", "checked": True,
                    "passed": approx(cost, draws),
                    "detail": f"cost {cost:,.0f} vs unlevered outflows {draws:,.0f}"})
    acq = val("purchase_price")
    if flows and acq:
        init = -min(v for _, v in flows)        # largest single outflow ~ acquisition/land
        out.append({"name": "acq_cost≈initial outflow", "checked": True,
                    "passed": approx(acq, init),
                    "detail": f"acq {acq:,.0f} vs initial outflow {init:,.0f}"})

    # exit_value ≈ terminal NOI / exit_cap (fall back to the canonical NOI point
    # when no full trajectory was extractable — e.g. a development model where
    # the operating build-up doesn't align with the cash-flow date axis).
    ev, ec = val("exit_value"), val("exit_cap")
    term_noi = (ops.get("noi") or {}).get("terminal")
    if term_noi is None:
        term_noi = val("noi")
    ecf = (ec / 100.0) if (ec and ec > 1.5) else ec      # percent → fraction
    if ev and ecf and term_noi:
        implied = term_noi / ecf
        out.append({"name": "exit_value≈NOI/exit_cap", "checked": True,
                    "passed": approx(ev, implied),
                    "detail": f"exit {ev:,.0f} vs NOI/cap {implied:,.0f} "
                              f"({term_noi:,.0f}/{ecf:.4f})"})
    else:
        out.append({"name": "exit_value≈NOI/exit_cap", "checked": False, "passed": None,
                    "detail": "missing exit value, exit cap, or terminal NOI"})

    # recomputed vs stated IRR (already in oracle, surfaced as an identity)
    lev = oracle.get("levered")
    if lev:
        out.append({"name": "recomputed_irr≈stated", "checked": True,
                    "passed": bool(lev.get("validated")),
                    "detail": f"recomputed {lev['recomputed_irr']} vs stated "
                              f"{lev['stated_irr']} ({lev['match_err_bps']} bps)"})
    return out


# ---------------------------------------------------------------------------
# Deal brief — the non-negotiable fact set, all cash-flow-sourced
# ---------------------------------------------------------------------------

_BRIEF_RATE = {"exit_cap", "going_in_cap", "ltv", "levered_irr", "unlevered_irr"}
_BRIEF_MULT = {"equity_multiple", "unlevered_equity_multiple"}


def _bfmt(concept: str, v: float) -> str:
    if concept in _BRIEF_RATE:
        return f"{v * 100:.2f}%" if abs(v) <= 1.5 else f"{v:.2f}%"
    if concept in _BRIEF_MULT:
        return f"{v:.2f}x"
    return f"${v:,.0f}"


def deal_brief_facts(canonical: dict, oracle: dict, hold: dict | None,
                     ops: dict) -> list[dict]:
    """The non-negotiable brief set the user requires — acquisition cost / going-in
    cap, total cost, LTV·debt·equity, hold period (from the sale event), sales
    price, exit cap, and returns (levered+unlevered IRR & EM). Each item carries
    its value + source, or is flagged 'not found'. All cash-flow-sourced."""
    out: list[dict] = []

    def item(label: str, concept: str, *, value=None, source=None, note=None):
        info = canonical.get(concept) if value is None else None
        cf_val = bool(info and info.get("cf_validated"))
        if value is None and info is not None:
            value, source = info["value"], info.get("source")
        found = value is not None
        disp = _bfmt(concept, float(value)) if found else "— not found —"
        out.append({"label": label, "concept": concept, "found": found,
                    "display": disp, "source": source, "note": note,
                    "cf_validated": cf_val})

    item("Acquisition cost", "purchase_price")
    item("Going-in cap", "going_in_cap")
    item("Total cost", "total_cost")
    item("LTV", "ltv")
    item("Debt", "debt")
    item("Equity", "equity")
    if hold:
        note = (f"sells at month {hold['months']} of a {hold['model_months']}-month model"
                if hold.get("sells_before_model_end")
                else f"{hold['months']} months")
        out.append({"label": "Hold period", "concept": "hold_period", "found": True,
                    "display": f"{hold['months']} mo ({hold['years']:g} yr)",
                    "source": hold.get("source"), "note": note})
    else:
        out.append({"label": "Hold period", "concept": "hold_period", "found": False,
                    "display": "— not found —", "source": None, "note": None})
    item("Sales price (gross)", "sale_price",
         note="net proceeds: " + (_bfmt("exit_value", float(canonical["exit_value"]["value"]))
                                  if "exit_value" in canonical else "—"))
    item("Exit cap rate", "exit_cap")
    item("Levered IRR", "levered_irr")
    item("Unlevered IRR", "unlevered_irr")
    item("Levered equity multiple", "equity_multiple")
    item("Unlevered equity multiple", "unlevered_equity_multiple")
    return out


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def _oracle_from_spine(spine) -> dict:
    """Adapt the validated spine streams into the oracle shape the rest of the
    pipeline consumes (hold detection, operating derivation, returns). The spine
    is the authoritative source — these streams already reproduced the stated IRR."""
    matched = dict(spine.matched)
    if "levered" not in matched and "primary" in matched:
        matched["levered"] = matched.pop("primary")
    out: dict[str, Any] = {"levered": None, "unlevered": None, "candidates": []}
    for leg in ("levered", "unlevered"):
        m = matched.get(leg)
        if not m:
            continue
        entry = {
            "sheet": m["sheet"], "row": m.get("loc"), "loc": m.get("loc"), "label": m["label"],
            "recomputed_irr": m["recomputed_irr"], "recomputed_em": m["recomputed_em"],
            "stated_irr": m["stated_irr"], "validated": True,
            "match_err_bps": m["match_err_bps"], "initial_outflow": m["initial_outflow"],
            "flows": m["flows"], "periodicity": m.get("periodicity"),
        }
        out[leg] = entry
        # candidate carries the leg as its label so _operating_from_stream can find it
        out["candidates"].append({**entry, "label": leg, "xirr": m["recomputed_irr"],
                                  "em": m["recomputed_em"]})
    return out


def _nearest_pow10(x: float) -> float:
    import math
    if x <= 0:
        return 1.0
    return float(10 ** round(math.log10(x)))


def _derive_stack_from_spine(oracle: dict, canonical: dict) -> float:
    """Rule 8: the capital stack / cost / sale come FROM the validated streams,
    not vocab-matched summary cells. First detect the streams' DOLLAR SCALE from
    the debt anchor (loan amounts are stated in full dollars: scale ≈ debt /
    (cost − equity), since debt + equity = cost) and SCALE THE FLOWS in place —
    so every downstream dollar derivation (stack, NOI, sale-backout, identities)
    is in full dollars and internally consistent. Then derive:
    equity = Σ levered outflows; total cost = Σ unlevered outflows; sale =
    terminal inflow. Returns the scale applied."""
    lev, unlev = oracle.get("levered"), oracle.get("unlevered")
    legs = [s for s in (lev, unlev) if s]
    if not legs:
        return 1.0
    # Streams arrive in FULL DOLLARS — the spine already applied each sheet's
    # DECLARED units ("$ in 000s" etc.). So derive the stack directly.
    oracle["candidates"] = [{**s, "label": name}
                            for name, s in (("levered", lev), ("unlevered", unlev)) if s]

    def setc(concept, raw, leg, label):
        if raw is None:
            return
        canonical[concept] = {
            "concept": concept, "value": raw,
            "source": f"{leg['loc']} ({label})", "method": "recomputed",
            "validated": True, "conflict": False, "cf_validated": True}
    if lev:
        setc("equity", -sum(v for _, v in lev["flows"] if v < 0), lev, "Σ equity outflows")
    if unlev:
        setc("total_cost", -sum(v for _, v in unlev["flows"] if v < 0), unlev, "Σ unlevered outflows")
        setc("sale_price", max((v for _, v in unlev["flows"]), default=None), unlev, "terminal inflow")
        # Acquisition cost = the largest single outflow (the asset purchase at
        # close dwarfs monthly operating/construction flows).
        mn = min((v for _, v in unlev["flows"]), default=0.0)
        if mn < 0:
            setc("purchase_price", -mn, unlev, "acquisition outflow (largest single)")

    # Bring vocab-sourced cost/debt onto the full-$ scale set by the stream-
    # derived EQUITY (the reliable full-$ anchor), using the scale-invariant LTC
    # ratio. Needed when there's no unlevered stream, so total_cost/debt came from
    # vocab in their sheet's units. cost = equity / (1 − LTC); debt = cost − equity.
    eq = _val_of(canonical, "equity")
    ltc = _val_of(canonical, "ltc")
    ci = canonical.get("total_cost")
    if eq and ci and ci.get("method") != "recomputed" and ltc and 0 < ltc < 1:
        cv = abs(float(ci["value"]))
        if cv > 0:
            sc = _nearest_pow10((eq / (1 - ltc)) / cv)
            if sc in (1e-3, 1e3, 1e6):
                canonical["total_cost"] = {**ci, "value": float(ci["value"]) * sc}
    cost = _val_of(canonical, "total_cost")
    di = canonical.get("debt")
    if di and cost and eq and (cost - eq) > 0:
        dv = abs(float(di["value"]))
        if dv > 0:
            sc = _nearest_pow10((cost - eq) / dv)
            if sc in (1e-3, 1e3, 1e6):
                canonical["debt"] = {**di, "value": float(di["value"]) * sc}
    return 1.0


def _val_of(canonical: dict, c: str):
    try:
        return float(canonical[c]["value"]) if c in canonical else None
    except (TypeError, ValueError):
        return None


def _engine_not_found(file_path: Path, spine) -> dict[str, Any]:
    """Honest result when the cash-flow engine can't be identified (rule 9): no
    reconstructed numbers, an explicit reason, and the non-negotiables all 'not
    found'. Downstream must NOT present anything as validated."""
    reason = spine.diagnostics.get("reason") or (
        "No cash-flow stream reproduced the model's stated IRR — the engine "
        "could not be identified. Numbers were NOT reconstructed.")
    return {
        "version": DEAL_TRUTH_VERSION, "file": file_path.name,
        "ok": False, "engine_found": False, "reason": reason,
        "deal_type": "unknown", "cashflow_engine": None, "hold": None,
        "canonical": {}, "brief_facts": deal_brief_facts({}, {}, None, {}),
        "operating_series": {}, "oracle": {}, "identities": [],
        "guardrails": [{"code": "engine_not_found", "message": reason,
                        "evidence": spine.diagnostics}],
        "spine": {"ok": False, "stated_irrs": len(spine.stated_irrs),
                  "candidate_streams": spine.n_candidate_streams},
    }


def build_deal_truth(file_path: str | Path) -> dict[str, Any]:
    file_path = Path(file_path)
    m = build_workbook_map(file_path)
    if m.get("error"):
        return {"error": m["error"], "version": DEAL_TRUTH_VERSION}
    try:
        tables = parse_workbook_tables_cached(file_path)
    except Exception as e:
        tables = []
        log.error("table parse failed: %s", e)

    # The spine is the anchor: find the one true cash-flow stream by matching its
    # recomputed IRR to the model's stated IRR (rules 1-7). If none validates,
    # the engine wasn't found — do NOT reconstruct (rules 6, 9).
    spine = find_spine(file_path)
    if not spine.ok:
        return _engine_not_found(file_path, spine)
    oracle = _oracle_from_spine(spine)
    # The engine's anchor sheet (the build-up usually lives with the unlevered
    # stream) becomes the cash-flow engine the operating reader prefers.
    anchor = (oracle.get("unlevered") or oracle.get("levered") or {}).get("sheet")
    if anchor:
        m["cashflow_engine"] = anchor
    hold = detect_hold(oracle)

    # Cells that some display cell formula-traces to — i.e. inputs the model
    # actually USES. Lets reconciliation prefer the used input over a lookalike.
    referenced: set[str] = set()
    for lst in m["candidates"].values():
        for e in lst:
            src = (e.get("provenance") or {}).get("source")
            if src and (e.get("provenance") or {}).get("crosses_sheet"):
                referenced.add(src)

    # Canonical winners: returns/hold come from the oracle; the rest from
    # concept-appropriate reconciliation.
    canonical: dict[str, dict] = {}
    static_concepts = [c for c in m["candidates"]
                       if c not in ("levered_irr", "unlevered_irr", "equity_multiple",
                                    "levered_cf", "unlevered_cf", "hold_period")]
    for c in static_concepts:
        picked = _canonical_for_concept(c, m["candidates"][c], referenced)
        if picked:
            canonical[c] = picked

    # Returns from the validated streams (override any mis-binned candidates),
    # including BOTH equity multiples (levered + unlevered) from their streams.
    if oracle.get("levered"):
        canonical["levered_irr"] = {
            "concept": "levered_irr", "value": oracle["levered"]["recomputed_irr"],
            "source": f"{oracle['levered']['loc']} (XIRR)",
            "method": "recomputed", "validated": oracle["levered"]["validated"],
            "conflict": False}
        if oracle["levered"].get("recomputed_em") is not None:
            canonical["equity_multiple"] = {
                "concept": "equity_multiple", "value": oracle["levered"]["recomputed_em"],
                "source": f"{oracle['levered']['loc']} (ΣCF)", "method": "recomputed",
                "validated": oracle["levered"]["validated"], "conflict": False}
    if oracle.get("unlevered"):
        canonical["unlevered_irr"] = {
            "concept": "unlevered_irr", "value": oracle["unlevered"]["recomputed_irr"],
            "source": f"{oracle['unlevered']['loc']} (XIRR)",
            "method": "recomputed", "validated": oracle["unlevered"]["validated"],
            "conflict": False}
        if oracle["unlevered"].get("recomputed_em") is not None:
            canonical["unlevered_equity_multiple"] = {
                "concept": "unlevered_equity_multiple",
                "value": oracle["unlevered"]["recomputed_em"],
                "source": f"{oracle['unlevered']['loc']} (ΣCF)", "method": "recomputed",
                "validated": oracle["unlevered"]["validated"], "conflict": False}

    # Rule 8: derive the capital stack / cost / sale FROM the validated stream
    # (engine wins over vocab-matched summary cells), detecting the stream's
    # dollar scale so mixed-unit models report full dollars.
    stack_scale = _derive_stack_from_spine(oracle, canonical)

    # Sales price comes from the stream (above). Only fall back to a vocab exit
    # candidate if the stream didn't yield one.
    if "sale_price" not in canonical:
        ev_cands = [e for e in m["candidates"].get("exit_value", [])
                    if _passes_domain("exit_value", e["value"])]
        if ev_cands:
            gross = max(ev_cands, key=lambda e: float(e["value"]))
            canonical["sale_price"] = {
                "concept": "sale_price", "value": float(gross["value"]),
                "source": (gross.get("provenance") or {}).get("source") or gross["display"],
                "provenance": gross.get("provenance"), "conflict": False}

    # Tiered operating trajectory (needs sale_price to back the sale out in 1b).
    ops = operating_trajectory(m, tables, oracle, canonical, hold, file_path)

    # Going-in cap: explicit if present, else the derived entry yield on cost.
    if "going_in_cap" not in canonical:
        noi0 = (ops.get("noi") or {}).get("stabilized") or (ops.get("noi") or {}).get("terminal")
        if noi0 is None and "noi" in canonical:
            noi0 = float(canonical["noi"]["value"])
        cost = float(canonical["total_cost"]["value"]) if "total_cost" in canonical else None
        if noi0 and cost:
            canonical["going_in_cap"] = {
                "concept": "going_in_cap", "value": round(noi0 / cost, 4),
                "source": "derived: stabilized NOI / total cost (entry yield on cost)",
                "method": "derived", "conflict": False}

    deal_type = _infer_deal_type(oracle, ops)
    identities = _run_identities(canonical, oracle, ops)
    # Promote a costed input to cash-flow-validated when it ties to the stream.
    for ident_name, concept in (("total_cost≈CF draws", "total_cost"),
                                ("acq_cost≈initial outflow", "purchase_price")):
        idr = next((i for i in identities if i["name"] == ident_name), None)
        if idr and idr.get("passed") and concept in canonical:
            canonical[concept]["cf_validated"] = True
    guardrails = _build_guardrails(canonical, oracle, identities, ops, deal_type, m)
    brief_facts = deal_brief_facts(canonical, oracle, hold, ops)

    result = {
        "version": DEAL_TRUTH_VERSION,
        "file": file_path.name,
        "ok": True,
        "engine_found": True,
        "spine_anchor": spine.diagnostics.get("anchor_sheets"),
        "deal_type": deal_type,
        "cashflow_engine": m.get("cashflow_engine"),
        "hold": hold,
        "canonical": canonical,
        "brief_facts": brief_facts,
        "operating_series": {k: {kk: vv for kk, vv in v.items() if kk != "by_year"}
                             | {"by_year": v.get("by_year")} for k, v in ops.items()},
        "oracle": {leg: {k: v for k, v in (oracle.get(leg) or {}).items() if k != "flows"}
                   for leg in ("levered", "unlevered") if oracle.get(leg)},
        "identities": identities,
        "guardrails": guardrails,
    }
    log.info("TRUTH %s — type=%s, hold=%s, %d canonical, %d identities, %d guardrails",
             file_path.name, deal_type, (hold or {}).get("months"),
             len(canonical), len(identities), len(guardrails))
    return result


def render_truth_text(d: dict) -> str:
    if d.get("error"):
        return f"ERROR: {d['error']}"
    L = [f"DEAL TRUTH — {d['file']}   (type: {d['deal_type']}, engine: {d['cashflow_engine']})",
         "=" * 78, "\nDEAL BRIEF — non-negotiables (all cash-flow-sourced)"]
    for b in d.get("brief_facts", []):
        miss = "" if b["found"] else "   ⟵ MISSING"
        cf = " ✓CF" if b.get("cf_validated") else ""
        note = f"   ({b['note']})" if b.get("note") else ""
        L.append(f"  {b['label']:<26} {b['display']:<14}{cf:<5} {b.get('source') or ''}{note}{miss}")
    h = d.get("hold")
    if h and h.get("sells_before_model_end"):
        L.append(f"  ⚠ sells early: month {h['months']} of a {h['model_months']}-month model")

    L.append("\nCANONICAL FACTS")
    for c, info in sorted(d["canonical"].items()):
        flag = "  ⚠CONFLICT" if info.get("conflict") else ""
        meth = f"  [{info['method']}]" if info.get("method") else ""
        val = info["value"]
        L.append(f"  {c:<22} = {val!r:<20} ← {info['source']}{meth}{flag}")
    L.append("\nOPERATING TRAJECTORY")
    for c, s in d["operating_series"].items():
        prov = s.get("provenance", "?")
        gi = s.get("going_in")
        gi_s = f"{gi:,.0f}" if isinstance(gi, (int, float)) else "—"
        tm = s.get("terminal")
        tm_s = f"{tm:,.0f}" if isinstance(tm, (int, float)) else "—"
        L.append(f"  {c}: going-in {gi_s} → terminal {tm_s}   [{prov}]")
    L.append("\nORACLE (returns recomputed from cash flow)")
    for leg in ("levered", "unlevered"):
        o = d["oracle"].get(leg)
        if o:
            ok = "✓validated" if o["validated"] else f"✗ {o['match_err_bps']}bps off"
            L.append(f"  {leg:<10} IRR {o['recomputed_irr']*100:.2f}% (stated "
                     f"{o['stated_irr']*100:.2f}%) EM {o.get('recomputed_em')}  {ok}"
                     f"  ← {o['sheet']}!row{o['row']} '{o['label']}'")
    L.append("\nIDENTITY CHECKS")
    for i in d["identities"]:
        st = "—" if not i["checked"] else ("PASS" if i["passed"] else "FAIL")
        L.append(f"  [{st}] {i['name']}: {i['detail']}")
    L.append("\nGUARDRAILS (what a later GPT layer may NOT say)")
    for gr in d["guardrails"]:
        L.append(f"  • [{gr['code']}] {gr['message']}")
    return "\n".join(L)


# ---------------------------------------------------------------------------
# Bridges to the product layers (Layer 3 Initial View + chat grounding)
# ---------------------------------------------------------------------------

# canonical concept -> the field label Layer 3 (investment_intel) understands.
_INTEL_FIELD = {
    "exit_cap": "Exit Cap Rate", "going_in_cap": "Going-in Cap Rate",
    "noi": "Stabilized NOI",                       # the forward/stabilized NOI
    "exit_value": "Exit Value", "total_cost": "Total Project Cost",
    "purchase_price": "Purchase Price", "debt": "Total Debt",
    "equity": "Total Equity", "ltc": "Loan-to-Cost (LTC)",
    "ltv": "Loan-to-Value (LTV)", "dscr": "DSCR", "debt_yield": "Debt Yield",
    "interest_rate": "Interest Rate", "levered_irr": "Levered IRR",
    "unlevered_irr": "Unlevered IRR", "equity_multiple": "Equity Multiple",
    "yield_on_cost": "Yield on Cost", "hold_period": "Hold Period",
}
_INTEL_RATE = {"exit_cap", "going_in_cap", "ltc", "ltv", "debt_yield",
               "interest_rate", "levered_irr", "unlevered_irr", "yield_on_cost"}
_INTEL_MULT = {"equity_multiple", "dscr"}


def _intel_display(concept: str, v: float) -> str:
    if concept in _INTEL_RATE:
        return f"{v * 100:.2f}%" if abs(v) <= 1.5 else f"{v:.2f}%"
    if concept in _INTEL_MULT:
        return f"{v:.2f}x"
    if concept == "hold_period":
        return f"{v:g}"
    return f"${v:,.0f}"


def to_intel_facts(dt: dict) -> list[dict]:
    """Convert canonical deal-truth facts into the verified-fact shape Layer 3
    (investment_intel) consumes — so the Initial View reasons over VALIDATED,
    complete facts instead of the thin brief set. Returned facts carry
    verdict 'show' (already validated here) and their canonical source."""
    facts: list[dict] = []
    for concept, info in (dt.get("canonical") or {}).items():
        field = _INTEL_FIELD.get(concept)
        if not field:
            continue
        try:
            v = float(info["value"])
        except (TypeError, ValueError):
            continue
        src = info.get("source") or ""
        sheet, _, cell = src.partition("!")
        facts.append({
            "field": field, "value": v, "display": _intel_display(concept, v),
            "sheet": sheet or None, "cell": cell or None,
            "trust": {"verdict": "show", "confidence": "high"},
        })
    return facts


def guardrail_lines(dt: dict) -> list[str]:
    """The guardrail messages as plain lines, for injection into a GPT prompt."""
    return [g["message"] for g in (dt.get("guardrails") or [])]


if __name__ == "__main__":
    args = sys.argv[1:]
    targets: list[Path] = []
    for a in args:
        p = Path(a)
        if p.is_dir():
            targets += sorted(p.glob("*.xlsx")) + sorted(p.glob("*.xlsm"))
        elif p.exists():
            targets.append(p)
    if not targets:
        up = Path("uploads")
        targets = sorted(up.glob("*.xlsx")) if up.exists() else []
    if not targets:
        print("usage: python deal_truth.py <file.xlsx | dir>")
        raise SystemExit(1)
    for t in targets:
        print("\n" + render_truth_text(build_deal_truth(t)) + "\n")
