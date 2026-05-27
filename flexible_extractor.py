from pathlib import Path
import json
import pandas as pd
import openpyxl

from metric_catalog import load_metric_catalog


UPLOAD_DIR = Path("uploads")
REPOSITORY_DIR = Path("repository")


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def is_numeric(value):
    return isinstance(value, (int, float)) and not pd.isna(value)


def normalize_text(value):
    return clean_text(value).lower()


def cell_address(row, col):
    return openpyxl.utils.get_column_letter(col) + str(row)


def find_nearby_value(ws, row, col):
    """
    Search nearby cells for a value.
    Priority:
    1. Same row, cells to the right
    2. Same column, cells below
    3. Small surrounding area
    """

    # Look right
    for offset in range(1, 6):
        value = ws.cell(row=row, column=col + offset).value
        if is_numeric(value):
            return value, cell_address(row, col + offset), "right"

    # Look below
    for offset in range(1, 6):
        value = ws.cell(row=row + offset, column=col).value
        if is_numeric(value):
            return value, cell_address(row + offset, col), "below"

    # Look nearby grid
    for r_offset in range(-2, 4):
        for c_offset in range(-2, 6):
            r = row + r_offset
            c = col + c_offset

            if r < 1 or c < 1:
                continue

            value = ws.cell(row=r, column=c).value
            if is_numeric(value):
                return value, cell_address(r, c), "nearby"

    return None, None, None


# Which data_nature values are relevant per SSOT layer.
# "mixed" metrics are always included (meaningful in both projection and actual contexts).
# "underwriting" scans all three because acquisition/closing models routinely contain both
# projected values (IRR, NOI proforma, exit cap) and actual values (closing costs paid,
# loan amount drawn, actual purchase price confirmed at closing).
_LAYER_DATA_NATURE: dict[str, set] = {
    "underwriting":    {"projection", "actual", "mixed"},
    "business_plan":   {"projection", "actual", "mixed"},
    "actuals_2020":    {"actual", "mixed"},
    "actuals_2021":    {"actual", "mixed"},
    "actuals_2022":    {"actual", "mixed"},
    "actuals_2023":    {"actual", "mixed"},
    "actuals_2024":    {"actual", "mixed"},
    "actuals_2025":    {"actual", "mixed"},
    "actuals_recent":  {"actual", "mixed"},
    "rent_roll":       {"actual", "mixed"},
    "debt":            {"actual", "mixed"},
}


def filter_catalog_for_layer(catalog: list, layer: str) -> list:
    """
    Return only the metrics relevant to a given SSOT layer.

    Two filters applied:
    1. Skip calculated metrics (metric_source == "calculated") — these are
       derived after extraction, not extracted from cells.
    2. Keep only metrics whose data_nature matches the layer's expected type.
       e.g. an underwriting file should not be scanned for Current LTV or DSCR
       (those are actual/current-state metrics).
    """
    allowed_natures = _LAYER_DATA_NATURE.get(layer, {"projection", "actual", "mixed"})
    return [
        m for m in catalog
        if m.get("metric_source", "extracted") == "extracted"
        and m.get("data_nature", "mixed") in allowed_natures
    ]


def scan_workbook_for_all_metrics(file_path, catalog):
    """
    Load the workbook ONCE and scan all catalog metrics in a single pass.

    This is the fast path used by v2's tools.extract_from_file. It replaces
    the prior pattern of calling scan_workbook_for_metric in a loop, which
    re-loaded the same Excel file once per metric (≈97x per file).

    Returns {metric_id: best_match_dict_or_None} for every metric in the catalog.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return {m["metric_id"]: None for m in catalog}

    # Pre-normalize every alias once, paired with its parent metric.
    # Each entry: (normalized_alias_text, metric_dict, original_alias_string)
    alias_index = []
    for metric in catalog:
        for alias in metric.get("aliases", []):
            alias_text = normalize_text(alias)
            if alias_text:
                alias_index.append((alias_text, metric, alias))

    matches_by_metric: dict = {m["metric_id"]: [] for m in catalog}
    file_name = Path(file_path).name

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell_text = normalize_text(cell.value)
                if not cell_text:
                    continue

                for alias_text, metric, original_alias in alias_index:
                    if alias_text not in cell_text:
                        continue

                    # Label quality: penalise matches where the alias is a small
                    # fraction of the cell label (e.g. "noi" inside "noi to offset
                    # interest"). An exact or near-exact label match scores 1.0;
                    # a substring-in-long-label scores proportionally lower.
                    label_ratio = len(alias_text) / max(len(cell_text), 1)
                    # Also penalise if alias appears mid-word (e.g. "irr" in "irrespective")
                    idx = cell_text.find(alias_text)
                    char_before = cell_text[idx - 1] if idx > 0 else " "
                    char_after  = cell_text[idx + len(alias_text)] if idx + len(alias_text) < len(cell_text) else " "
                    mid_word = char_before.isalpha() or char_after.isalpha()
                    if mid_word:
                        continue  # alias embedded inside a longer word — skip

                    value, value_cell, direction = find_nearby_value(
                        ws, cell.row, cell.column
                    )
                    if value is None:
                        continue

                    # Confidence tiers:
                    #   "exact"  — alias covers ≥80% of the cell label, value right/below
                    #   "high"   — value right/below (alias may be partial label)
                    #   "medium" — value found nearby
                    #   "partial"— alias is a small fragment of a longer label (label_ratio < 0.4)
                    if direction in ("right", "below"):
                        confidence = "exact" if label_ratio >= 0.8 else "high"
                    else:
                        confidence = "partial" if label_ratio < 0.4 else "medium"

                    matches_by_metric[metric["metric_id"]].append({
                        "metric_id": metric["metric_id"],
                        "metric_name": metric["metric_name"],
                        "category": metric["category"],
                        "definition": metric["definition"],
                        "value": value,
                        "source_file": file_name,
                        "sheet": sheet_name,
                        "label_cell": cell.coordinate,
                        "value_cell": value_cell,
                        "matched_alias": original_alias,
                        "confidence": confidence,
                        "label_ratio": round(label_ratio, 2),
                        "match_method": direction,
                    })

    # Best match per metric — ranked by confidence tier then label quality.
    # Tier order: exact > high > medium > partial
    _TIER = {"exact": 0, "high": 1, "medium": 2, "partial": 3}
    best = {}
    for metric_id, matches in matches_by_metric.items():
        if not matches:
            best[metric_id] = None
        else:
            matches.sort(key=lambda x: (
                _TIER.get(x["confidence"], 9),
                -x.get("label_ratio", 0),  # higher label_ratio wins ties
            ))
            best[metric_id] = matches[0]
    return best


def scan_workbook_for_metric(file_path, metric):
    """
    Search one Excel workbook for one metric.
    Returns best match or None.

    NOTE: kept for backward compatibility with v1 modules. The fast path is
    scan_workbook_for_all_metrics, which avoids reloading the workbook per metric.
    """

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None

    aliases = metric.get("aliases", [])
    matches = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                cell_text = normalize_text(cell.value)

                if not cell_text:
                    continue

                for alias in aliases:
                    alias_text = normalize_text(alias)

                    if not alias_text:
                        continue

                    if alias_text in cell_text:
                        value, value_cell, direction = find_nearby_value(
                            ws,
                            cell.row,
                            cell.column
                        )

                        if value is not None:
                            confidence = "high" if direction in ["right", "below"] else "medium"

                            matches.append({
                                "metric_id": metric["metric_id"],
                                "metric_name": metric["metric_name"],
                                "category": metric["category"],
                                "definition": metric["definition"],
                                "value": value,
                                "source_file": Path(file_path).name,
                                "sheet": sheet_name,
                                "label_cell": cell.coordinate,
                                "value_cell": value_cell,
                                "matched_alias": alias,
                                "confidence": confidence,
                                "match_method": direction,
                            })

    if not matches:
        return None

    # Prefer high confidence matches first
    matches = sorted(
        matches,
        key=lambda x: 0 if x["confidence"] == "high" else 1
    )

    return matches[0]


def extract_raw_labeled_pairs(file_path, max_pairs: int = 600) -> list[dict]:
    """
    Extract ALL (sheet, label, value) pairs from a workbook without any
    catalog filtering. This is the input for Pass 2 (GPT insight pass).

    Returns a list of dicts:
        {"sheet": str, "label": str, "value": numeric, "cell": str,
         "direction": "right"|"below"|"nearby", "label_len": int}

    Quality fields (used by run_raw_insight_pass to filter noise):
      direction: "right"/"below" = label directly precedes value — high signal
                 "nearby" = value found in surrounding area — lower signal
      label_len: very short labels (< 5 chars) are often headers/indices, not metrics

    Capped at max_pairs. Priority sheets (summary, assumptions, waterfall) come first.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    # Prioritise summary/assumption sheets so if we hit the cap we keep the
    # most analytically relevant rows.
    priority_keywords = [
        "summary", "assumption", "return", "waterfall", "overview",
        "sources", "uses", "debt", "equity", "cashflow", "cash flow",
        "proforma", "pro forma", "irr", "exit",
    ]

    def sheet_priority(name: str) -> int:
        nl = name.lower()
        return 0 if any(kw in nl for kw in priority_keywords) else 1

    sorted_sheets = sorted(wb.sheetnames, key=sheet_priority)

    pairs = []
    seen_labels: set[str] = set()

    for sheet_name in sorted_sheets:
        if len(pairs) >= max_pairs:
            break
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            if len(pairs) >= max_pairs:
                break
            for cell in row:
                cell_text = clean_text(cell.value)
                if not cell_text or len(cell_text) < 3:
                    continue
                # Only process text cells (labels)
                if not isinstance(cell.value, str):
                    continue

                value, value_cell, direction = find_nearby_value(
                    ws, cell.row, cell.column
                )
                if value is None:
                    continue

                # Deduplicate by (sheet, normalised label) to avoid
                # repeated header rows skewing GPT's reading.
                key = f"{sheet_name}|{normalize_text(cell_text)}"
                if key in seen_labels:
                    continue
                seen_labels.add(key)

                pairs.append({
                    "sheet":     sheet_name,
                    "label":     cell_text,
                    "value":     value,
                    "cell":      value_cell,
                    "direction": direction,
                    "label_len": len(cell_text),
                })

    return pairs


def classify_file_layer(file_name):
    """
    Classify a file by its investment lifecycle layer based on its name.
    Returns one of: 'underwriting', 'business_plan', 'actuals_2021',
    'actuals_2022', 'actuals_recent', or 'unknown'.

    These names must match ssot.KNOWN_LAYERS exactly.

    Keyword groups reflect institutional RE naming conventions:
      - 'proforma' / 'pro forma' is the most common name for an UW model
      - 'BP' alone is risky (matches too much) so we anchor with word boundaries
      - financial statements: 'fs', 'financial', 'p&l', 'income statement',
        'operating statement', 't12'
    """
    name_lower = file_name.lower()

    # --- Financial Statements / actuals (check first; "2022 P&L" should NOT
    # match business plan via the year). ---
    # We pad with leading/trailing spaces so " fs " matches "FS 2022.xlsx"
    padded = f" {name_lower} "
    actuals_keywords = [
        "financial statement", "income statement", "operating statement",
        "p&l", "pl statement", "actual", "actuals",
        " fs ", "_fs_", "_fs.", " fs.", "t12", "trailing 12",
    ]
    if any(kw in padded for kw in actuals_keywords):
        for year in ("2020", "2021", "2022", "2023", "2024", "2025"):
            if year in name_lower:
                return f"actuals_{year}"
        return "actuals_recent"

    # --- Acquisition Underwriting (proforma / UW model / deal memo / closing docs) ---
    uw_keywords = [
        "acquisition", "underwriting",
        "proforma", "pro forma", "pro-forma",
        "uw model", "deal memo",
        "closing", "settlement",  # closing statement / settlement statement
        "psa", "purchase agreement",        # purchase & sale agreement
        "ic memo", "investment committee",  # IC package
    ]
    # Word-boundary check for the short token " uw" (avoid matching "answer"!)
    uw_token_match = (
        " uw" in name_lower or "_uw" in name_lower
        or name_lower.endswith(" uw") or name_lower.endswith("_uw")
    )
    if any(kw in name_lower for kw in uw_keywords) or uw_token_match:
        return "underwriting"

    # --- Business Plan (revised plan post-acquisition) ---
    bp_keywords = [
        "business plan", "budget", "forecast", "revised plan",
        "annual plan", "asset plan", "hold plan",
    ]
    if any(kw in name_lower for kw in bp_keywords):
        return "business_plan"
    # " bp " as a standalone token (so "abp_2022.xlsx" doesn't false-match)
    if " bp " in name_lower or "_bp_" in name_lower or "_bp." in name_lower or " bp." in name_lower:
        return "business_plan"

    return "unknown"


def scan_uploaded_files(upload_dir=UPLOAD_DIR):
    """
    Scan all uploaded Excel files against the metric catalog.
    Extracts each metric from EVERY file where found, tagged by source layer,
    so the analysis can compare underwriting vs business plan vs actuals.
    """

    upload_dir = Path(upload_dir)
    REPOSITORY_DIR.mkdir(exist_ok=True)

    catalog = load_metric_catalog()

    excel_files = list(upload_dir.glob("*.xlsx")) + list(upload_dir.glob("*.xlsm"))

    extracted = []
    missing = []

    for metric in catalog:
        all_matches = []

        for file_path in excel_files:
            match = scan_workbook_for_metric(file_path, metric)

            if match:
                match["source_layer"] = classify_file_layer(file_path.name)
                all_matches.append(match)

        if all_matches:
            extracted.extend(all_matches)
        else:
            missing.append({
                "metric_id": metric["metric_id"],
                "metric_name": metric["metric_name"],
                "category": metric["category"],
                "definition": metric["definition"],
                "source": metric.get("source", ""),
                "priority": metric.get("priority", "medium"),
                "aliases": metric.get("aliases", []),
                "status": "missing"
            })

    result = {
        "status": "success",
        "total_metrics": len(catalog),
        "extracted_count": len(extracted),
        "missing_count": len(missing),
        "extracted_metrics": extracted,
        "missing_metrics": missing,
    }

    with open(REPOSITORY_DIR / "flexible_extraction_result.json", "w") as f:
        json.dump(result, f, indent=2, default=str)

    pd.DataFrame(extracted).to_csv(
        REPOSITORY_DIR / "extracted_metrics_report.csv",
        index=False
    )

    pd.DataFrame(missing).to_csv(
        REPOSITORY_DIR / "missing_metrics_report.csv",
        index=False
    )

    return result


if __name__ == "__main__":
    result = scan_uploaded_files()
    print(f"Total metrics: {result['total_metrics']}")
    print(f"Extracted: {result['extracted_count']}")
    print(f"Missing: {result['missing_count']}")
    print("Saved reports to repository/")