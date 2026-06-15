"""
workbook_map.py — Slice 1 of the deal-reconstruction rebuild (2026-06-15).

Before Collie extracts a single number it must prove "I understand how this
workbook is organized." This module produces that proof: a deterministic,
GPT-free structural map of a workbook —

    sheet roles          (what each tab IS, by behavior)
    economic blocks      (what regions MEAN: revenue / NOI / debt / sale / ...)
    formula provenance   (where a displayed number actually COMES FROM)
    candidate facts      (every place a concept appears, with its provenance)
    diagnostics          (so a misread is inspectable)

It deliberately stops BEFORE judgment: it does NOT pick conflict winners, label
returns verified/unverified, or write narrative. Reconciliation, the IRR/EM
validation oracle, canonical winners, and the GPT guardrails live in the next
module (deal_truth.py). Keeping the map judgment-free means the same structural
read can be inspected regardless of how the deal later resolves.

Design rules (non-negotiable):
  * No GPT. Pure openpyxl + arithmetic + vocabulary.
  * No per-file logic. Nothing here keys off a specific deal, tab, row, or
    value. Sheets are understood by BEHAVIOR (roles, formulas, date axes,
    vocabulary), so it must survive acquisition / development / refi / value-add
    / mixed-use / weird-sheet-name models, not just one test file.

Built ON existing deterministic layers (no duplication):
  * workbook_orientation.orient_workbook  → sheet roles + ref counts + signals
  * financial_model_parser.parse_workbook_tables_cached → period-axis tables
  * formula_tracer._parse_refs / _split_sheet → cross-sheet formula precedents

Public:
  build_workbook_map(path) -> dict   (the artifact; JSON-serializable)
  render_map_text(map) -> str        (human-readable dump for the harness)
"""
from __future__ import annotations

import logging
import re
import sys
from pathlib import Path
from typing import Any, Callable

from metric_resolver import parse_numeric_value

log = logging.getLogger("fb.map")
if not log.handlers:
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("[fb.map] %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

WORKBOOK_MAP_VERSION = "2026-06-15.1"

# ---------------------------------------------------------------------------
# Concept vocabulary — maps a free-form label to the economic role it plays.
# Order matters: MORE SPECIFIC variants first so substrings don't mis-bind
# ("unlevered irr" contains "irr"; "exit cap" contains "cap"). First match wins.
# This is the ONLY place economic meaning is hard-coded, and it is generic
# vocabulary, never a per-deal value.
# ---------------------------------------------------------------------------
_CONCEPT_VOCAB: list[tuple[str, tuple[str, ...]]] = [
    # --- pricing / exit ---
    ("exit_cap",       ("exit cap", "terminal cap", "disposition cap", "reversion cap")),
    ("going_in_cap",   ("going-in cap", "going in cap", "entry cap", "in-place cap",
                        "acquisition cap")),
    ("exit_value",     ("terminal value", "reversion value", "gross sale", "sale price",
                        "sales price", "net sale", "net proceeds", "net exit proceeds",
                        "exit value", "disposition value", "residual value")),
    ("purchase_price", ("purchase price", "acquisition price", "acquisition cost", "land pp")),
    ("total_cost",     ("total project cost", "total development cost", "total budget",
                        "total dev cost", "all-in basis", "total basis", "total uses",
                        "total sources & uses", "total sources", "total cost")),
    # --- capital stack ---
    ("debt",           ("loan amount", "debt amount", "total debt", "senior loan",
                        "loan proceeds", "mortgage", "debt")),
    ("equity",         ("lp equity", "gp equity", "sponsor equity", "total equity",
                        "equity required", "equity invested", "peak equity", "equity")),
    ("ltc",            ("loan-to-cost", "loan to cost", "ltc")),
    ("ltv",            ("loan-to-value", "loan to value", "ltv")),
    ("dscr",           ("dscr", "debt service coverage", "debt coverage")),
    ("debt_yield",     ("debt yield",)),
    ("interest_rate",  ("interest rate", "all-in rate", "coupon", "fixed rate")),
    # --- operations ---
    ("noi",            ("net operating income", "operating income", "noi")),
    ("revenue",        ("effective gross", "gross potential", "total revenue",
                        "total income", "rental income", "net revenue", "egi", "gpr")),
    ("opex",           ("operating expense", "total expense", "total opex", "opex")),
    ("capex",          ("capital expenditure", "capex", "tenant improvement", "ti/lc",
                        "reserve")),
    ("debt_service",   ("debt service", "interest expense", "principal & interest",
                        "p&i", "loan d/s")),
    # --- returns ---
    ("unlevered_irr",  ("unlevered irr", "unleveraged irr", "un-levered irr",
                        "unlevered return", "project un-levered")),
    ("levered_irr",    ("levered irr", "leveraged irr", "project levered", "project irr",
                        "irr")),
    ("equity_multiple",("equity multiple", "equity multiplier", "moic", "eqmx",
                        "multiple on invested capital")),
    ("yield_on_cost",  ("yield on cost", "yield-on-cost", "return on cost", "roc",
                        "development yield", "stabilized yield", "untrended yield")),
    ("cash_on_cash",   ("cash-on-cash", "cash on cash", "cash yield")),
    ("hold_period",    ("hold period", "investment period", "exit year", "hold (years)")),
    # --- cash-flow stream rows (time series). Unlevered BEFORE levered: the
    #     substring "levered cash flow" lives inside "unlevered cash flows". ---
    ("unlevered_cf",   ("unlevered cash flow", "unleveraged cash flow", "un-levered cash flow",
                        "project cash flow")),
    ("levered_cf",     ("levered cash flow", "leveraged cash flow", "net levered",
                        "cash flow to equity", "equity cash flow", "net cash flow to equity")),
]

# Concepts whose appearance as a TABLE ROW (a time series) marks a sheet block.
_ROW_KINDS = {
    "revenue", "opex", "noi", "capex", "debt_service", "levered_cf", "unlevered_cf",
}

# Block kinds inferred from the set of concepts found in a region/table.
_BLOCK_FOR_CONCEPT = {
    "purchase_price": "acquisition", "total_cost": "sources_uses",
    "debt": "sources_uses", "equity": "sources_uses", "ltc": "sources_uses",
    "ltv": "sources_uses", "revenue": "revenue", "opex": "expenses",
    "noi": "noi", "capex": "capex", "debt_service": "debt_service",
    "interest_rate": "debt", "dscr": "debt", "debt_yield": "debt",
    "exit_value": "sale", "exit_cap": "sale",
    "levered_irr": "returns", "unlevered_irr": "returns", "equity_multiple": "returns",
    "yield_on_cost": "returns", "cash_on_cash": "returns", "hold_period": "returns",
    "levered_cf": "cashflow", "unlevered_cf": "cashflow",
}


def _concept_of(label: str) -> str | None:
    """Map a label to its economic concept (first/most-specific match)."""
    if not label:
        return None
    f = label.strip().lower()
    for concept, kws in _CONCEPT_VOCAB:
        if any(kw in f for kw in kws):
            return concept
    return None


# Each concept has an expected value DOMAIN. A label match whose value falls
# outside its concept's domain is almost always a vocabulary false-positive
# (e.g. a "debt" of 19, an "equity multiple" of $35M, a "hold period" of
# 286,144). Gating by domain removes that noise generically — no per-deal tuning.
_RATE = {"exit_cap", "going_in_cap", "ltc", "ltv", "debt_yield", "interest_rate",
         "levered_irr", "unlevered_irr", "yield_on_cost", "cash_on_cash"}
_MULTIPLE = {"equity_multiple", "dscr"}
_MONEY = {"purchase_price", "total_cost", "debt", "equity", "exit_value", "noi",
          "revenue", "opex", "capex", "debt_service", "levered_cf", "unlevered_cf"}
_COUNT = {"hold_period"}


def _passes_domain(concept: str, value: Any) -> bool:
    """Is `value` plausible for `concept`? Accepts both fraction (0.0525) and
    percent (5.25) forms for rates; demands real magnitude for money concepts."""
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return False
    v = float(value)
    if concept in _RATE:
        return -1.0 <= v <= 100.0          # fraction or percent form
    if concept in _MULTIPLE:
        return 0.1 <= v <= 25.0
    if concept in _MONEY:
        return abs(v) >= 1000.0
    if concept in _COUNT:
        return 0 < v <= 40 and float(v).is_integer()
    return True


# ---------------------------------------------------------------------------
# Formula provenance — follow a displayed cell to where its value is really
# computed (the "formula source"), distinguishing display from truth.
# ---------------------------------------------------------------------------

# A pure pass-through: '=Sheet!A1', '=+A1', '=  Other!$B$2'  (one ref, nothing else).
_PASSTHROUGH_RE = re.compile(
    r"^=\s*[+\-]?\s*((?:'[^']+'|[A-Za-z0-9_.]+)!)?\$?[A-Z]{1,3}\$?\d+\s*$"
)
_FUNC_RE = re.compile(r"^=\s*([A-Za-z][A-Za-z0-9.]*)\s*\(")


def _formula_op(formula: str) -> str | None:
    """The operation a formula performs: a function name (SUM/XIRR/NPV/...),
    'ARITH' for bare arithmetic over >1 reference, or None for a pass-through."""
    m = _FUNC_RE.match(formula)
    if m:
        return m.group(1).upper()
    if _PASSTHROUGH_RE.match(formula):
        return None
    return "ARITH"


def _passthrough_target(formula: str, default_sheet: str,
                        parse_refs: Callable) -> tuple[str, str] | None:
    """If `formula` just forwards one cell, return that (sheet, coord); else None."""
    if not _PASSTHROUGH_RE.match(formula):
        return None
    refs = parse_refs(formula, default_sheet)
    return refs[0] if len(refs) == 1 else None


def trace_provenance(form_get: Callable[[str, str], Any], sheet: str, coord: str,
                     parse_refs: Callable, max_hops: int = 8) -> dict:
    """
    Follow a cell's formula chain to its source. `form_get(sheet, coord)` returns
    the FORMULA string (or literal value) at a cell.

    Returns:
      display       : "Sheet!Coord" (where the number is shown)
      chain         : ["Sheet!Coord", ...] the full hop path
      source        : "Sheet!Coord" where the value is computed or hard-coded
      source_sheet  : the sheet of `source`
      op            : the operation at the source ('SUM'|'XIRR'|'ARITH'|None)
      is_hardcode   : the source is a literal value, not a reference
      ref_sheets    : sheets the source formula pulls from (for aggregations)
      crosses_sheet : the chain spans more than one sheet
    """
    start = f"{sheet}!{coord}"
    chain = [start]
    cur_s, cur_c = sheet, coord
    op = None
    is_hardcode = False
    ref_sheets: list[str] = []

    for _ in range(max_hops):
        val = form_get(cur_s, cur_c)
        if not (isinstance(val, str) and val.startswith("=")):
            is_hardcode = True            # leaf is a literal
            break
        op = _formula_op(val)
        try:
            refs = parse_refs(val, cur_s)
        except Exception:
            refs = []
        pt = _passthrough_target(val, cur_s, parse_refs)
        if pt is not None:
            ns, nc = pt
            a1 = f"{ns}!{nc}"
            if a1 in chain:               # cycle guard
                break
            chain.append(a1)
            cur_s, cur_c = ns, nc
            continue
        # Aggregation / arithmetic: this cell IS the computation site. Stop.
        ref_sheets = list(dict.fromkeys(s for s, _ in refs))
        break

    source = chain[-1]
    return {
        "display": start,
        "chain": chain,
        "source": source,
        "source_sheet": cur_s,
        "op": op,
        "is_hardcode": is_hardcode,
        "ref_sheets": ref_sheets,
        "crosses_sheet": len({c.split("!", 1)[0] for c in chain}) > 1,
    }


# ---------------------------------------------------------------------------
# Grid helpers (in-memory; random openpyxl access on read_only is pathological)
# ---------------------------------------------------------------------------

def _col_letter(idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def _nearest_value(grid: list[tuple], r: int, c: int, reach: int = 8):
    """First numeric value to the RIGHT of a label, then BELOW it. Returns
    (value, coord) or (None, None). 0-based grid; r,c are 0-based here."""
    row = grid[r] if r < len(grid) else ()
    for cc in range(c + 1, min(c + 1 + reach, len(row))):
        v = row[cc]
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return v, f"{_col_letter(cc + 1)}{r + 1}"
    for rr in range(r + 1, min(r + 1 + 3, len(grid))):
        v = grid[rr][c] if c < len(grid[rr]) else None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return v, f"{_col_letter(c + 1)}{rr + 1}"
    return None, None


# ---------------------------------------------------------------------------
# Map construction
# ---------------------------------------------------------------------------

def _load_grids(file_path: Path) -> dict[str, list[tuple]]:
    """Per-sheet value grids, pulled once with sequential iteration (fast)."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    grids: dict[str, list[tuple]] = {}
    for s in wb.sheetnames:
        if s.rstrip(">").strip() == "" or s.endswith(">"):
            continue
        ws = wb[s]
        if not ws.max_row:
            grids[s] = []
            continue
        grids[s] = [
            row for row in ws.iter_rows(
                min_row=1, max_row=min(ws.max_row, 250),
                max_col=min(ws.max_column or 1, 60), values_only=True)
        ]
    try:
        wb.close()
    except Exception:
        pass
    return grids


def _scan_static_facts(grid: list[tuple], sheet: str, role: str) -> list[dict]:
    """Vocabulary-anchored scan of a static sheet: every label that names a
    concept, paired with the nearest numeric value. Spatial (left-to-right,
    top-to-bottom), position-agnostic. Returns candidate facts (no provenance
    yet — that's attached in build_workbook_map where formulas are available)."""
    out: list[dict] = []
    for r, row in enumerate(grid):
        for c, v in enumerate(row):
            if not (isinstance(v, str) and v.strip()):
                continue
            concept = _concept_of(v)
            if not concept:
                continue
            val, coord = _nearest_value(grid, r, c)
            if val is None or not _passes_domain(concept, val):
                continue
            out.append({
                "concept": concept,
                "label": v.strip()[:60],
                "value": val,
                "sheet": sheet,
                "cell": coord,
                "role": role,
                "kind": "static",
            })
    return out


def _blocks_from_table(table: dict) -> dict | None:
    """Turn a parsed period-axis table into an economic block: classify its rows
    into concepts and infer the block kind from what's present."""
    concepts_seen: dict[str, dict] = {}
    for rrow in table.get("rows", []):
        concept = _concept_of(rrow.get("label", ""))
        if concept and concept not in concepts_seen:
            concepts_seen[concept] = {
                "label": (rrow.get("label") or "")[:60],
                "row": rrow.get("row"),
                "values_by_period": rrow.get("values_by_period"),
            }
    if not concepts_seen:
        return None
    # Block kind = most operationally-specific concept present.
    priority = ["cashflow", "noi", "sale", "debt_service", "revenue", "expenses", "capex"]
    kinds = {_BLOCK_FOR_CONCEPT.get(k) for k in concepts_seen}
    kind = next((k for k in priority if k in kinds), "operating")
    return {
        "kind": kind,
        "sheet": table["sheet"],
        "title": table.get("title"),
        "periodicity": table.get("periodicity"),
        "period_axis": "horizontal",
        "header_row": table.get("header_row"),
        "date_headers": table.get("date_headers"),
        "concepts": concepts_seen,
    }


def build_workbook_map(file_path: str | Path, use_cache: bool = True) -> dict[str, Any]:
    """
    Build the deterministic structural map of a workbook. No GPT, no judgment.

    Returns a JSON-serializable dict:
      {
        "version", "file",
        "engine": "<sheet most-referenced by others, or null>",
        "sheets": { name: {role, confidence, behavior:{...}, static_blocks:[kind,...]} },
        "timeseries_blocks": [ {kind, sheet, title, periodicity, concepts:{...}} ],
        "candidates": { concept: [ {value, display, sheet, cell, role, label,
                                    provenance:{...}, periodicity?} ] },
        "diagnostics": {...}
      }  or {"error": "..."}
    """
    file_path = Path(file_path)
    import openpyxl
    from workbook_orientation import orient_workbook
    from financial_model_parser import parse_workbook_tables_cached
    try:
        from formula_tracer import _parse_refs
    except Exception:                       # pragma: no cover - defensive
        _parse_refs = lambda f, s: []       # noqa: E731

    # --- Layer 1: sheet roles (reuse orientation) -------------------------
    orient = orient_workbook(file_path)
    if orient.get("error"):
        return {"error": orient["error"], "version": WORKBOOK_MAP_VERSION}
    osheets: dict[str, dict] = orient.get("sheets", {})

    # The sheet other sheets reference most is the INPUTS HUB (assumptions are
    # read by everyone) — NOT necessarily the computation engine. The cashflow
    # engine is identified later from the time-series blocks.
    inputs_hub, hub_refs = None, 0
    for name, info in osheets.items():
        n = (info.get("signals") or {}).get("referenced_by_other_sheets", 0)
        if n > hub_refs:
            inputs_hub, hub_refs = name, n

    # --- load value grids + a formula workbook for provenance -------------
    grids = _load_grids(file_path)
    try:
        wb_form = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        return {"error": f"Could not open workbook for formulas: {e}",
                "version": WORKBOOK_MAP_VERSION}

    def form_get(sheet: str, coord: str):
        try:
            return wb_form[sheet][coord].value
        except Exception:
            return None

    # --- Layer 2: economic blocks (time series) ---------------------------
    timeseries_blocks: list[dict] = []
    try:
        tables = parse_workbook_tables_cached(file_path)
    except Exception as e:
        tables = []
        log.error("table parse failed for %s: %s", file_path.name, e)
    for t in tables:
        blk = _blocks_from_table(t)
        if blk:
            timeseries_blocks.append(blk)

    # Cashflow engine = the cashflow time-series block richest in deal-level CF
    # concepts (bonus for carrying BOTH levered + unlevered streams, finer
    # periodicity preferred). This is the real computation hub — where returns,
    # NOI, and proceeds are produced — independent of which tab is referenced most.
    def _engine_score(b: dict) -> tuple:
        cs = set(b.get("concepts", {}))
        rich = len(cs & {"levered_cf", "unlevered_cf", "noi", "revenue", "opex", "total_cost"})
        both = 1 if {"levered_cf", "unlevered_cf"} <= cs else 0
        fine = 1 if b.get("periodicity") == "monthly" else 0
        return (both, rich, fine)
    cf_blocks = [b for b in timeseries_blocks if b["kind"] == "cashflow"]
    cashflow_engine = max(cf_blocks, key=_engine_score)["sheet"] if cf_blocks else None

    # --- Layer 2: economic blocks (static) + candidate facts --------------
    candidates: dict[str, list[dict]] = {}
    sheet_static_kinds: dict[str, set[str]] = {s: set() for s in grids}

    def _add_candidate(fact: dict, provenance: dict | None, periodicity: str | None = None):
        entry = {
            "value": fact["value"],
            "display": f"{fact['sheet']}!{fact['cell']}",
            "sheet": fact["sheet"],
            "cell": fact["cell"],
            "role": fact["role"],
            "label": fact["label"],
            "provenance": provenance,
        }
        if periodicity:
            entry["periodicity"] = periodicity
        candidates.setdefault(fact["concept"], []).append(entry)

    for sheet, grid in grids.items():
        role = (osheets.get(sheet, {}) or {}).get("role", "other")
        # Static facts from summary / inputs / returns / support tabs (the places
        # headline numbers are displayed). Model tabs are covered by time series.
        if role in ("summary", "inputs", "returns", "support", "other"):
            for fact in _scan_static_facts(grid, sheet, role):
                prov = trace_provenance(form_get, fact["sheet"], fact["cell"], _parse_refs)
                _add_candidate(fact, prov)
                sheet_static_kinds[sheet].add(_BLOCK_FOR_CONCEPT.get(fact["concept"], "other"))

    # Time-series concept rows also become candidates (NOI, CF rows, etc.), with
    # provenance traced from the row's first populated period cell.
    for blk in timeseries_blocks:
        for concept, info in blk["concepts"].items():
            vbp = info.get("values_by_period") or {}
            # representative value = last non-null period (terminal/stabilized)
            rep_val, rep_period = None, None
            for per, vv in vbp.items():
                if isinstance(vv, (int, float)) and not isinstance(vv, bool):
                    rep_val, rep_period = vv, per
            if rep_val is None or not _passes_domain(concept, rep_val):
                continue
            row_no = info.get("row")
            # find the period column for rep_period to cite a real cell
            cell = None
            prov = None
            if row_no:
                cell = f"R{row_no}"  # row-level reference; period cell resolved in deal_truth
            candidates.setdefault(concept, []).append({
                "value": rep_val,
                "display": f"{blk['sheet']}!row{row_no}",
                "sheet": blk["sheet"],
                "cell": cell,
                "role": "model",
                "label": info.get("label"),
                "provenance": None,
                "periodicity": blk.get("periodicity"),
                "series": True,
            })

    try:
        wb_form.close()
    except Exception:
        pass

    # Dedup candidates by (sheet, cell) within each concept — the same value can
    # be reached via two nearby labels; keep one entry per source cell.
    for concept, entries in candidates.items():
        seen: set[tuple] = set()
        deduped = []
        for e in entries:
            key = (e["sheet"], e.get("cell"))
            if key in seen:
                continue
            seen.add(key)
            deduped.append(e)
        candidates[concept] = deduped

    # --- assemble sheet summaries ----------------------------------------
    sheets_out: dict[str, dict] = {}
    for name, info in osheets.items():
        sig = info.get("signals", {}) or {}
        sheets_out[name] = {
            "role": info.get("role"),
            "confidence": info.get("confidence"),
            "behavior": {
                "hardcode_ratio": sig.get("hardcode_ratio"),
                "pulls_from_other_sheets": sig.get("pulls_from_other_sheets"),
                "referenced_by_other_sheets": sig.get("referenced_by_other_sheets"),
                "time_series_rows": sig.get("time_series_rows"),
                "date_header": sig.get("date_header"),
                "populated_cells": sig.get("populated_cells"),
            },
            "static_blocks": sorted(k for k in sheet_static_kinds.get(name, set())
                                    if k != "other"),
        }

    result = {
        "version": WORKBOOK_MAP_VERSION,
        "file": file_path.name,
        "cashflow_engine": cashflow_engine,
        "inputs_hub": inputs_hub,
        "sheets": sheets_out,
        "timeseries_blocks": timeseries_blocks,
        "candidates": candidates,
        "diagnostics": {
            "n_sheets": len(osheets),
            "n_timeseries_blocks": len(timeseries_blocks),
            "n_candidate_concepts": len(candidates),
            "inputs_hub_inbound_refs": hub_refs,
        },
    }
    log.info("MAP %s — cf_engine=%s, inputs_hub=%s, %d sheets, %d ts-blocks, %d concepts",
             file_path.name, cashflow_engine, inputs_hub,
             len(osheets), len(timeseries_blocks), len(candidates))
    return result


# ---------------------------------------------------------------------------
# Human-readable dump (the harness output)
# ---------------------------------------------------------------------------

def render_map_text(m: dict) -> str:
    if m.get("error"):
        return f"ERROR: {m['error']}"
    L: list[str] = []
    L.append(f"WORKBOOK MAP — {m['file']}")
    L.append(f"  cashflow engine: {m.get('cashflow_engine') or '—'}   "
             f"inputs hub: {m.get('inputs_hub') or '—'}")
    L.append("=" * 78)
    L.append("\nSHEET ROLES")
    for name, s in m["sheets"].items():
        b = s["behavior"]
        flags = []
        if b.get("date_header"):
            flags.append("date-axis")
        if b.get("referenced_by_other_sheets"):
            flags.append(f"in:{b['referenced_by_other_sheets']}")
        if b.get("pulls_from_other_sheets"):
            flags.append(f"out:{b['pulls_from_other_sheets']}")
        blk = ("  blocks=" + ",".join(s["static_blocks"])) if s["static_blocks"] else ""
        L.append(f"  {name:<22} {str(s['role']):<8} conf={s['confidence']}"
                 f"  [{' '.join(flags)}]{blk}")

    L.append("\nTIME-SERIES BLOCKS")
    for blk in m["timeseries_blocks"]:
        cs = ", ".join(sorted(blk["concepts"]))
        L.append(f"  {blk['sheet']:<22} {blk['kind']:<12} {blk['periodicity']:<8} :: {cs}")

    L.append("\nCANDIDATE FACTS (concept → where it appears)")
    for concept in sorted(m["candidates"]):
        L.append(f"  {concept}")
        for e in m["candidates"][concept]:
            prov = e.get("provenance") or {}
            tag = ""
            if prov.get("crosses_sheet"):
                tag = f"  ⇒ {prov['source']}"
                if prov.get("op"):
                    tag += f" ({prov['op']})"
            elif prov.get("op"):
                tag = f"  ({prov['op']})"
            elif prov.get("is_hardcode"):
                tag = "  (hardcode)"
            if e.get("series"):
                tag = f"  [series/{e.get('periodicity')}]"
            L.append(f"      {e['display']:<26} = {e['value']!r:<22}{tag}")
    return "\n".join(L)


if __name__ == "__main__":
    import json
    args = sys.argv[1:]
    targets: list[Path] = []
    if args:
        for a in args:
            p = Path(a)
            if p.is_dir():
                targets += sorted(p.glob("*.xlsx")) + sorted(p.glob("*.xlsm"))
            elif p.exists():
                targets.append(p)
    else:
        up = Path("uploads")
        targets = sorted(up.glob("*.xlsx")) + sorted(up.glob("*.xlsm")) if up.exists() else []
    if not targets:
        print("usage: python workbook_map.py <file.xlsx | dir> [...]   (or put files in uploads/)")
        raise SystemExit(1)
    for t in targets:
        m = build_workbook_map(t)
        print("\n" + render_map_text(m) + "\n")
        if "--json" in args:
            print(json.dumps(m, indent=2, default=str))
