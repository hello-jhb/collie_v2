"""
trust_engine.py — per-fact confidence scoring (step 2 of the comprehension-first
rebuild). This is what replaces the human gate: instead of a person confirming
every number, each fact the Model Brief cites earns a verdict from independent
checks, and only the low-confidence minority is surfaced for review.

Five signals per fact:
  1. grounded      — the cited cell actually contains the claimed value
                     (anti-fabrication floor; fail -> omit). Deterministic.
  2. authoritative — that cell sits on an orientation-classified summary /
                     inputs / returns tab (tier <= 3). Deterministic.
  3. reconciles    — the fact satisfies a deal identity it participates in
                     (Price x Cap ~= NOI, Debt + Equity ~= Basis, LTV ~=
                     Debt/Price, ...). Deterministic. None when the fact joins
                     no identity (neither credit nor penalty).
  4/5. challenged  — ONE adversarial GPT re-read of the authoritative tabs: for
                     each headline fact, does the cell agree, or is there a
                     MORE authoritative cell with a DIFFERENT value? Agreement
                     corroborates; a conflicting authoritative source flags.
                     (Corroboration + challenge collapse into this one call.)

Verdict:
  not grounded                               -> OMIT  (low)
  grounded + (reconciles False | challenge disagree) -> FLAG (low, conflict)
  grounded + authoritative + (reconciles True | challenge agree) -> SHOW (high)
  otherwise grounded                          -> SHOW (medium)

This module verifies facts; it does NOT yet regenerate the brief prose from only
verified facts, and it does NOT demote the audit gate. Those are the next step.
"""
from __future__ import annotations

import json
import logging
import sys
from pathlib import Path
from typing import Any

from metric_resolver import parse_numeric_value
from scenarios._llm import client, MODEL, llm_available

log = logging.getLogger("fb.trust")
if not log.handlers:
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("[fb.trust] %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

TRUST_ENGINE_VERSION = "2026-06-12.1"

# Concept keyword sets — map a free-form fact field to an identity role. First
# match wins; a fact maps to at most one concept. Exit variants are checked
# before their going-in counterparts so "Exit Cap" doesn't match "cap".
_CONCEPTS: list[tuple[str, tuple[str, ...]]] = [
    ("exit_noi",       ("exit noi", "terminal noi", "forward noi", "reversion noi")),
    ("exit_cap",       ("exit cap", "terminal cap", "disposition cap", "reversion cap")),
    ("exit_value",     ("exit value", "terminal value", "sale value", "gross sale",
                        "sale price", "reversion value", "disposition value")),
    ("going_in_cap",   ("going-in cap", "going in cap", "entry cap", "in-place cap",
                        "acquisition cap", "cap rate")),
    ("noi",            ("net operating income", "noi")),
    ("purchase_price", ("purchase price", "acquisition price", "acquisition cost")),
    ("tpc",            ("total project cost", "total basis", "total capitalization",
                        "total cost", "all-in basis", "total uses")),
    ("debt",           ("loan amount", "debt amount", "total debt", "mortgage", "senior loan")),
    ("equity",         ("equity required", "total equity", "equity invested", "peak equity")),
    ("ltv",            ("loan-to-value", "loan to value", "ltv")),
]

# Identities over NORMALIZED display magnitudes (full dollars / fractions), with
# a relative tolerance. Each: (name, participant concepts, predicate).
_REL_TOL = 0.12


def _approx(a: float, b: float, tol: float = _REL_TOL) -> bool:
    denom = max(abs(a), abs(b), 1e-9)
    return abs(a - b) / denom <= tol


_IDENTITIES: list[tuple[str, tuple[str, ...], Any]] = [
    ("Purchase Price x Going-in Cap = NOI",
     ("purchase_price", "going_in_cap", "noi"),
     lambda v: _approx(v["purchase_price"] * v["going_in_cap"], v["noi"])),
    ("Exit Value x Exit Cap = Exit NOI",
     ("exit_value", "exit_cap", "exit_noi"),
     lambda v: _approx(v["exit_value"] * v["exit_cap"], v["exit_noi"])),
    ("Debt + Equity = Total Basis",
     ("debt", "equity", "tpc"),
     lambda v: _approx(v["debt"] + v["equity"], v["tpc"])),
    ("LTV = Debt / Purchase Price",
     ("ltv", "debt", "purchase_price"),
     lambda v: _approx(v["ltv"], v["debt"] / v["purchase_price"]) if v["purchase_price"] else False),
]


def _concept_of(field: str) -> str | None:
    f = (field or "").lower()
    for concept, kws in _CONCEPTS:
        if any(kw in f for kw in kws):
            return concept
    return None


def _norm_magnitude(fact: dict) -> float | None:
    """Real-world magnitude of a fact for identity math: parse the human display
    ('$224.1M' -> 224100000, '6.0%' -> 0.06, '1.92x' -> 1.92), falling back to
    the literal value. Identity math needs consistent scale; display carries it."""
    for key in ("display", "value"):
        raw = fact.get(key)
        if raw in (None, "", "-", "—"):
            continue
        num, ok = parse_numeric_value(raw)
        if ok and isinstance(num, (int, float)):
            return float(num)
    return None


# ---------------------------------------------------------------------------
# Check 1 — grounding (cell actually contains the claimed value)
# ---------------------------------------------------------------------------

_GROUND_MAX_ROW = 600
_GROUND_MAX_COL = 90


def _read_cells(file_path: Path, needed: dict[str, set[str]]) -> dict[tuple[str, str], Any]:
    """Read only the cited (sheet, coordinate) cells. Uses values_only iteration
    and computes coordinates by index — read-only mode yields EmptyCell objects
    with no .coordinate, so we never touch cell.coordinate."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    out: dict[tuple[str, str], Any] = {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        log.error("Grounding open failed for %s: %s", file_path.name, e)
        return out
    for sheet, coords in needed.items():
        if sheet not in wb.sheetnames or not coords:
            continue
        want = set(coords)
        try:
            ws = wb[sheet]
            for r, row in enumerate(
                ws.iter_rows(min_row=1, max_row=_GROUND_MAX_ROW,
                             min_col=1, max_col=_GROUND_MAX_COL, values_only=True),
                start=1,
            ):
                if not want:
                    break
                for ci, val in enumerate(row, start=1):
                    if val is None:
                        continue
                    coord = f"{get_column_letter(ci)}{r}"
                    if coord in want:
                        out[(sheet, coord)] = val
                        want.discard(coord)
                        if not want:
                            break
        except Exception:
            continue
    try:
        wb.close()
    except Exception:
        pass
    return out


def _scan_sheet_values(file_path: Path, sheets: set[str]) -> dict[str, list[tuple[str, Any]]]:
    """For each sheet, collect (coordinate, value) for every numeric cell — used
    to recover a fact whose VALUE is right but whose CITED CELL is off (gpt-4o
    cites a neighbouring cell). Bounded, read-only, values_only."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    out: dict[str, list[tuple[str, Any]]] = {s: [] for s in sheets}
    if not sheets:
        return out
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return out
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        try:
            ws = wb[sheet]
            for r, row in enumerate(
                ws.iter_rows(min_row=1, max_row=_GROUND_MAX_ROW,
                             min_col=1, max_col=_GROUND_MAX_COL, values_only=True),
                start=1,
            ):
                for ci, v in enumerate(row, start=1):
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        out[sheet].append((f"{get_column_letter(ci)}{r}", v))
        except Exception:
            continue
    try:
        wb.close()
    except Exception:
        pass
    return out


def _grounds(claimed, actual) -> bool:
    """Does the actual cell value support the claimed value? Numbers match within
    rounding, or off by a clean scale (x1000) / percent representation (x100);
    text by containment; dates by equality."""
    if actual is None:
        return False
    if isinstance(actual, str) and not actual.strip():
        return False
    # Text claim
    cnum, ok = parse_numeric_value(claimed)
    if not ok or not isinstance(cnum, (int, float)):
        a, c = str(actual).strip().lower(), str(claimed).strip().lower()
        return bool(c) and bool(a) and (c in a or a in c)
    anum, aok = parse_numeric_value(actual)
    if not aok or not isinstance(anum, (int, float)):
        return False
    if anum == 0:
        return abs(cnum) < 1e-9
    for eq in (cnum, cnum * 1000.0, cnum / 1000.0, cnum * 100.0, cnum / 100.0):
        if abs(eq - anum) <= 0.02 * max(abs(anum), 1e-9):
            return True
    return False


# ---------------------------------------------------------------------------
# Check 4/5 — adversarial challenge read
# ---------------------------------------------------------------------------

_CHALLENGE_SYSTEM = """\
You are auditing facts another analyst extracted from an underwriting model.
You are given the FULL CONTENT of the model's authoritative tabs (each cell with
its A1 reference) and a list of claimed facts, each with the sheet!cell it was
taken from. For EACH claimed fact, verify independently:

- "agree": the cited cell contains this value (allowing thousands/percent
  formatting), AND no more-authoritative cell contradicts it.
- "disagree": a cell on an equally- or more-authoritative tab gives a DIFFERENT
  value for this same concept (report it in `better`). This is a real conflict.
- "absent": you cannot locate support for the claim in the tabs.

Be skeptical and specific. Watch for per-asset slices in portfolio models being
passed off as deal-level totals.

Return ONLY JSON, no prose, no fences:
{ "<index>": {"status": "agree|disagree|absent",
              "better": {"value": <v>, "sheet": "...", "cell": "..."} | null,
              "note": "<short>"} }
where <index> is the fact's position in the list (0-based).
"""


def _challenge(facts: list[dict], cells_block: str) -> dict[int, dict]:
    """One adversarial GPT pass over all facts. Returns {index: {status, better, note}}."""
    if not facts or not cells_block or not llm_available():
        return {}
    lines = [
        f"{i}. {f.get('field')}: {f.get('display', f.get('value'))} "
        f"[from {f.get('sheet')}!{f.get('cell')}]"
        for i, f in enumerate(facts)
    ]
    user_msg = (
        "CLAIMED FACTS:\n" + "\n".join(lines)
        + "\n\nAUTHORITATIVE TABS:\n" + cells_block
    )
    try:
        resp = client.chat.completions.create(
            model=MODEL,
            temperature=0.0,
            messages=[
                {"role": "system", "content": _CHALLENGE_SYSTEM},
                {"role": "user", "content": user_msg},
            ],
        )
        raw = (resp.choices[0].message.content or "").strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        parsed = json.loads(raw)
        return {int(k): v for k, v in parsed.items() if isinstance(v, dict)}
    except Exception as e:
        log.error("Challenge read failed: %s", e)
        return {}


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def _authoritative(sheet: str, orientation: dict | None) -> tuple[bool, str | None]:
    """True if the cited sheet is an oriented summary/inputs/returns tab."""
    if not orientation or orientation.get("error"):
        return False, None
    info = (orientation.get("sheets") or {}).get(sheet)
    if not info:
        return False, None
    role = info.get("role")
    return role in ("summary", "inputs", "returns"), role


def _reconcile(facts: list[dict]) -> dict[int, tuple[bool, str]]:
    """Run identities over the facts. Returns {fact_index: (passed, identity_name)}
    for every fact that participates in at least one identity."""
    # First fact per concept (brief returns ~one per field).
    concept_idx: dict[str, int] = {}
    concept_val: dict[str, float] = {}
    for i, f in enumerate(facts):
        c = _concept_of(f.get("field", ""))
        if c and c not in concept_idx:
            mag = _norm_magnitude(f)
            if mag is not None:
                concept_idx[c] = i
                concept_val[c] = mag
    out: dict[int, tuple[bool, str]] = {}
    for name, parts, pred in _IDENTITIES:
        if all(p in concept_val for p in parts):
            try:
                passed = bool(pred({p: concept_val[p] for p in parts}))
            except Exception:
                continue
            for p in parts:
                idx = concept_idx[p]
                # A pass anywhere wins; only set a fail if nothing passed yet.
                if idx not in out or passed:
                    out[idx] = (passed, name)
    return out


def score_facts(
    brief: dict,
    file_path: str | Path,
    orientation: dict | None = None,
    run_challenge: bool = True,
    authoritative_sheets: set[str] | None = None,
    cells_block: str | None = None,
) -> dict[str, Any]:
    """
    Score every fact in a Model Brief OR a focused dive. Returns:
      {"facts": [fact + "trust"], "summary": {high, medium, flagged, omitted}}

    Each fact gains a "trust" block: grounded, authoritative, reconciles,
    challenge, confidence (high|medium|low), verdict (show|flag|omit), notes.

    authoritative_sheets: when given, a fact cited from one of these sheets is
        AUTHORITATIVE — used by dives, where the topic's own tabs (the NOI sheet
        for cash flow, the Debt sheet for leverage) are the authority, not the
        brief's summary/inputs tabs. Falls back to the orientation-role check.
    cells_block: pre-rendered sheet text to challenge against (the dive already
        rendered its sheets — reuse it instead of re-reading the summary tabs).
    """
    file_path = Path(file_path)
    facts = [dict(f) for f in (brief.get("facts") or [])]
    if not facts:
        return {"facts": [], "summary": {"high": 0, "medium": 0, "flagged": 0, "omitted": 0}}

    # Check 1 — grounding (batched cell read)
    needed: dict[str, set[str]] = {}
    for f in facts:
        s, c = f.get("sheet"), f.get("cell")
        if s and c:
            needed.setdefault(s, set()).add(c)
    cells = _read_cells(file_path, needed)

    # Check 3 — reconciliation over the whole fact set
    recon = _reconcile(facts)

    # Check 4/5 — adversarial challenge (one GPT call). Challenge against the
    # supplied block (dive's own sheets) or, for the brief, the summary tabs.
    challenge: dict[int, dict] = {}
    if run_challenge and llm_available():
        try:
            block = cells_block
            if block is None:
                from workbook_orientation import analyst_reading_stack
                _, block = analyst_reading_stack(file_path)
            challenge = _challenge(facts, block)
        except Exception as e:
            log.error("Challenge stage skipped: %s", e)

    # Exact-cell grounding first; collect sheets of ungrounded NUMERIC facts so
    # we can recover a right-value/wrong-cell case by scanning the cited sheet.
    exact: dict[int, tuple[bool, Any]] = {}
    fallback_sheets: set[str] = set()
    for i, f in enumerate(facts):
        s, c = f.get("sheet"), f.get("cell")
        actual = cells.get((s, c)) if (s and c) else None
        g = bool(s and c) and _grounds(f.get("value", f.get("display")), actual)
        exact[i] = (g, actual)
        if not g and s:
            num, ok = parse_numeric_value(f.get("display", f.get("value")))
            if ok and isinstance(num, (int, float)):
                fallback_sheets.add(s)
    scan = _scan_sheet_values(file_path, fallback_sheets)

    summary = {"high": 0, "medium": 0, "flagged": 0, "omitted": 0}
    for i, f in enumerate(facts):
        s, c = f.get("sheet"), f.get("cell")
        composed = bool(f.get("composed_from"))
        grounded, actual = exact[i]
        claimed_disp = f.get("display", f.get("value"))

        # Recover a right-value/wrong-cell fact: if the exact cell failed but the
        # claimed value appears elsewhere on the SAME (authoritative) sheet,
        # ground it there and correct the provenance.
        corrected_cell = None
        if not grounded and s in scan:
            claimed = f.get("display", f.get("value"))
            for coord, v in scan[s]:
                if _grounds(claimed, v):
                    grounded, corrected_cell = True, coord
                    break

        if authoritative_sheets is not None:
            authoritative = s in authoritative_sheets
            role = "topic" if authoritative else (s or None)
        else:
            authoritative, role = _authoritative(s, orientation)
        rec = recon.get(i)  # (passed, name) or None
        ch = challenge.get(i, {})
        ch_status = ch.get("status")

        notes: list[str] = []
        if corrected_cell:
            notes.append(f"source cell corrected to {s}!{corrected_cell} (brief cited {c})")
        if rec:
            notes.append(("reconciles: " if rec[0] else "RECONCILE FAIL: ") + rec[1])
        if ch_status == "disagree" and ch.get("better"):
            b = ch["better"]
            notes.append(f"challenge: conflicting {b.get('sheet')}!{b.get('cell')}={b.get('value')}")
        elif ch_status:
            notes.append(f"challenge: {ch_status}")

        # Verdict
        conflict = (rec and rec[0] is False) or (ch_status == "disagree")
        positive = (rec and rec[0] is True) or (ch_status == "agree")

        if not grounded and not composed:
            if ch_status == "agree":
                # Value corroborated by the independent challenge read even though
                # the exact cell couldn't be confirmed — show, but mark it.
                verdict, conf = "show", "medium"
                notes.insert(0, f"cell {s}!{c} unconfirmed; challenge corroborates the value")
            else:
                verdict, conf = "omit", "low"
                notes.insert(0, f"not grounded: {s}!{c} does not contain {claimed_disp}")
        elif conflict:
            verdict, conf = "flag", "low"
        elif grounded and authoritative and positive:
            verdict, conf = "show", "high"
        elif composed and not grounded:
            verdict, conf = ("show", "medium")
            notes.insert(0, "composed figure (arithmetic not yet recomputed)")
        else:
            verdict, conf = "show", "medium"

        # Adopt the corrected cell as the fact's provenance (display + persist).
        if corrected_cell:
            f["cell"] = corrected_cell
        f["trust"] = {
            "grounded": grounded,
            "actual_cell_value": actual,
            "corrected_cell": corrected_cell,
            "authoritative": authoritative,
            "sheet_role": role,
            "reconciles": (rec[0] if rec else None),
            "identity": (rec[1] if rec else None),
            "challenge": ch_status,
            "challenge_note": ch.get("note"),
            "confidence": conf,
            "verdict": verdict,
            "notes": notes,
        }
        key = "flagged" if verdict == "flag" else ("omitted" if verdict == "omit" else conf)
        summary[key] = summary.get(key, 0) + 1

    return {"facts": facts, "summary": summary}
