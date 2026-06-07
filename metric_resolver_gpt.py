"""
metric_resolver_gpt.py — Phase 2 GPT resolver for candidate-pool disambiguation.

Runs only when a bounded metric has multiple validation-passing candidates and
the deterministic ranker can't tell which is right. GPT sees:
  - the metric schema (unit, period, range, definition)
  - each candidate's value, sheet, cell, and SURROUNDING CELL CONTEXT
    (row label in column A, above/below labels, column headers)
GPT picks one with reasoning and confidence.

Cost: ~$0.005-0.01 per call with gpt-4o-mini. Fires only on truly ambiguous
candidate_pool records (not when candidates all agree on value).

Also exposes run_identity_checks() — deterministic cross-checks like
Equity + Debt ≈ Total Project Cost. Flags inconsistencies as suspicious.
"""
from __future__ import annotations
import json
import logging
import sys
from pathlib import Path
from typing import Any

from scenarios._llm import client, MODEL_FAST, llm_available
from re_knowledge import knowledge_block, IDENTITY_RELATIONSHIPS

log = logging.getLogger("fb.resolver_gpt")
if not log.handlers:
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("[fb.resolver_gpt] %(asctime)s %(levelname)s %(message)s"))
    log.addHandler(h)
    log.setLevel(logging.INFO)


RESOLVER_GPT_VERSION = "phase2.v2"  # improved prompt: period glossary + industry conventions


# =============================================================================
# Identity arithmetic — deterministic cross-checks
# =============================================================================
#
# Each check is a function (bounded_metrics) -> list of flags.
# A flag is a dict {metric_name, reason, expected, found, severity}.
# severity: "suspicious" (>5% off) or "info" (small mismatch).
#
# Currently implemented checks:
#   - Going-in Cap Rate = NOI / Purchase Price (within 10% tolerance)
#   - LTV = Debt Amount / Purchase Price (or Total Project Cost)
#   - Equity + Debt ≈ Total Project Cost
#   - Equity Multiple consistent with IRR + Hold Period (rough order check)

def _get_numeric(bm: dict, name: str):
    rec = bm.get(name)
    if not rec:
        return None
    v = rec.get("normalized_value")
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def _check_going_in_cap_rate(bm: dict) -> list[dict]:
    noi = _get_numeric(bm, "Net Operating Income (NOI)")
    price = _get_numeric(bm, "Purchase Price")
    cap = _get_numeric(bm, "Going-in Cap Rate")
    if not all(v is not None and v > 0 for v in (noi, price, cap)):
        return []
    implied_cap = noi / price
    if abs(implied_cap - cap) / max(cap, implied_cap) > 0.10:
        return [{
            "metric_name": "Going-in Cap Rate",
            "reason": (
                f"Stated going-in cap {cap*100:.2f}% does not reconcile with "
                f"NOI / Purchase Price = {implied_cap*100:.2f}%. "
                f"One of NOI ({noi:,.0f}), Purchase Price ({price:,.0f}), "
                f"or Cap Rate is likely wrong."
            ),
            "severity": "suspicious",
        }]
    return []


def _check_ltv_consistency(bm: dict) -> list[dict]:
    debt = _get_numeric(bm, "Debt Amount") or _get_numeric(bm, "Loan Amount")
    price = _get_numeric(bm, "Purchase Price")
    cost = _get_numeric(bm, "Total Project Cost")
    ltv = _get_numeric(bm, "Original LTV")
    if ltv is None or (debt is None) or not (price or cost):
        return []
    # LTV may be relative to purchase price OR total project cost
    candidates = [c for c in (price, cost) if c and c > 0]
    flags = []
    if not any(abs(debt/denom - ltv) / max(ltv, debt/denom) < 0.05 for denom in candidates):
        flags.append({
            "metric_name": "Original LTV",
            "reason": (
                f"Stated LTV {ltv*100:.1f}% does not reconcile with "
                f"Debt / Price ({debt/price*100:.1f}% if price={price:,.0f}) "
                f"or Debt / Total Cost. Debt or LTV likely wrong."
            ),
            "severity": "suspicious",
        })
    return flags


def _check_sources_uses(bm: dict) -> list[dict]:
    """
    Sources = Uses check, made phased-funding aware.

    For value-add / development deals, the INITIAL debt + equity legitimately
    funds less than Total Project Cost — future CapEx is funded by later draws.
    So Total Project Cost >= initial (Debt + Equity) is NORMAL and must NOT be
    flagged. We only flag two genuine errors:
      (a) sources EXCEED uses materially (Debt + Equity > TPC by >5%) — impossible
          in a balanced model, signals a wrong extraction.
      (b) TPC is implausibly larger than initial cap (TPC > 3x (Debt+Equity)) —
          likely TPC landed on the wrong cell (e.g. a cumulative cash-flow total).
    """
    debt = _get_numeric(bm, "Debt Amount") or _get_numeric(bm, "Loan Amount")
    equity = _get_numeric(bm, "Equity Invested") or _get_numeric(bm, "Total Equity")
    cost = _get_numeric(bm, "Total Project Cost")
    if not all(v is not None and v > 0 for v in (debt, equity, cost)):
        return []
    sum_se = debt + equity

    # (a) sources exceed uses — real error
    if sum_se > cost * 1.05:
        return [{
            "metric_name": "Total Project Cost",
            "reason": (
                f"Sources exceed uses: Debt ({debt:,.0f}) + Equity ({equity:,.0f}) "
                f"= {sum_se:,.0f} > Total Project Cost {cost:,.0f}. One of these is "
                f"likely wrong (or Debt/Equity are totals while TPC is partial)."
            ),
            "severity": "suspicious",
        }]
    # (b) TPC implausibly large vs initial capitalization
    if cost > sum_se * 3.0:
        return [{
            "metric_name": "Total Project Cost",
            "reason": (
                f"Total Project Cost {cost:,.0f} is >3x initial capitalization "
                f"(Debt+Equity = {sum_se:,.0f}). TPC may have landed on a wrong "
                f"cell (e.g. a cumulative cash-flow total)."
            ),
            "severity": "suspicious",
        }]
    # TPC between (Debt+Equity) and 3x = normal phased funding → no flag.
    return []


_IDENTITY_CHECKS = [
    _check_going_in_cap_rate,
    _check_ltv_consistency,
    _check_sources_uses,
]


def run_identity_checks(bounded_metrics: dict[str, Any]) -> dict[str, list[str]]:
    """
    Run all identity arithmetic checks. Returns {metric_name: [list of flag reasons]}.
    Callers can append these to validation_notes and/or downgrade status to suspicious.
    """
    flags_by_metric: dict[str, list[str]] = {}
    for check in _IDENTITY_CHECKS:
        try:
            for flag in check(bounded_metrics):
                metric = flag["metric_name"]
                flags_by_metric.setdefault(metric, []).append(flag["reason"])
        except Exception as e:
            log.warning("Identity check %s crashed: %s", check.__name__, e)
    return flags_by_metric


# =============================================================================
# GPT candidate-pool resolver
# =============================================================================

SYSTEM_PROMPT = """\
You are an experienced real estate underwriting analyst choosing the deal-level
value for a specific metric from a list of candidate cells in an Excel model.

""" + knowledge_block(include=["period", "deal_level", "debt"]) + """

ANTI-PATTERNS (reject these unless explicitly requested):
  - Per-unit / per-key / per-SF / per-property cells when total wanted
  - Historical year columns (2014, 2015, 2016, 2017, 2018-1, 2018-2) when
    going-in or stabilized wanted
  - Sub-property / single-asset cells when deal-level / consolidated wanted
  - Sensitivity table cells (rows of alternative cap rates / IRRs)
  - Scenario alternatives (downside, upside) when base case wanted
  - Single-line items in a longer build-up (e.g., "Construction Contingency"
    when "Total Hard Costs" or "Total Project Cost" wanted)

If NONE of the candidates is semantically a clear match for the requested
metric, set "chosen_index" to null and explain why. Do not pick just because
confidence is highest — only pick when the row label + context semantically
match.

Return ONLY JSON:
{
  "chosen_index": <int 0-based> | null,
  "reasoning": "short sentence citing the row label / column header / context clue",
  "confidence": "high" | "medium" | "low"
}
No prose, no markdown fences.
"""


def _get_cell_text(ws, row: int, col: int) -> str:
    """Safe cell-text getter for context gathering."""
    try:
        v = ws.cell(row=row, column=col).value
        if v is None:
            return ""
        s = str(v).strip()
        return s[:60] if len(s) > 60 else s
    except Exception:
        return ""


def _gather_candidate_context(wb, candidate: dict) -> dict:
    """
    Read the cells around a candidate's label/value cell to give GPT semantic context.

    Returns:
      {
        row_label, label_above, label_below,
        col_a_at_row, col_b_at_row,  # often the actual row label
        column_header,
        col_left_value, col_right_value
      }
    """
    sheet = candidate.get("sheet")
    label_cell = candidate.get("label_cell") or candidate.get("value_cell")
    value_cell = candidate.get("value_cell")
    if not sheet or sheet not in wb.sheetnames or not label_cell:
        return {}

    import openpyxl.utils as _u
    try:
        ws = wb[sheet]
        # Parse the label cell ref to row + col
        col_letters = "".join(ch for ch in label_cell if ch.isalpha())
        row_digits  = "".join(ch for ch in label_cell if ch.isdigit())
        if not col_letters or not row_digits:
            return {}
        lcol = _u.column_index_from_string(col_letters)
        lrow = int(row_digits)
    except Exception:
        return {}

    ctx = {
        "row_label":      _get_cell_text(ws, lrow, lcol),
        "label_above":    _get_cell_text(ws, lrow - 1, lcol),
        "label_below":    _get_cell_text(ws, lrow + 1, lcol),
        "col_a_at_row":   _get_cell_text(ws, lrow, 1),
        "col_b_at_row":   _get_cell_text(ws, lrow, 2),
    }

    # Get the column header for the value cell.
    # This is the SINGLE most useful piece of context for Phase 2 — a column
    # labeled "Going-In NOI" vs "Exit NOI" tells GPT which period the value is.
    column_header_path: list[str] = []
    if value_cell and value_cell != label_cell:
        try:
            vcol_letters = "".join(ch for ch in value_cell if ch.isalpha())
            vrow_digits  = "".join(ch for ch in value_cell if ch.isdigit())
            vcol = _u.column_index_from_string(vcol_letters)
            vrow = int(vrow_digits)
            # Collect ALL text cells in this column above the value, up to 8 rows.
            # Multi-row headers are common; capture the stack so GPT sees the full label.
            for r in range(max(1, vrow - 8), vrow):
                txt = _get_cell_text(ws, r, vcol)
                if txt and not txt.replace(",", "").replace(".", "").replace("-", "").replace("%", "").isdigit():
                    column_header_path.append(txt)
            if column_header_path:
                ctx["column_header"] = " / ".join(column_header_path[-3:])  # most recent 3
        except Exception:
            pass

    # ALSO show the row's neighboring values so GPT can see if this is part of
    # a wider table (per-key, per-unit, total columns side by side).
    row_neighbors: list[str] = []
    try:
        for c_offset in range(-3, 4):
            c = lcol + c_offset
            if c < 1 or c == lcol:
                continue
            v = ws.cell(row=lrow, column=c).value
            if v is None or str(v).strip() == "":
                continue
            cell_ref = _u.get_column_letter(c) + str(lrow)
            row_neighbors.append(f"{cell_ref}={str(v)[:30]}")
    except Exception:
        pass
    if row_neighbors:
        ctx["row_neighbors"] = ", ".join(row_neighbors)

    return ctx


def _format_candidate_for_prompt(idx: int, candidate: dict, ctx: dict) -> str:
    """Render a candidate + context as readable text block for the GPT prompt."""
    lines = [
        f"Candidate {idx}:",
        f"  value:        {candidate.get('value')}",
        f"  sheet:        {candidate.get('sheet')}",
        f"  cell:         {candidate.get('value_cell')}",
        f"  sheet_tier:   {candidate.get('sheet_tier')} (lower = more authoritative)",
        f"  matched alias: {candidate.get('matched_alias')!r}",
    ]
    if ctx:
        if ctx.get("row_label"):
            lines.append(f"  row label:    {ctx['row_label']!r}")
        if ctx.get("col_a_at_row"):
            lines.append(f"  col A label:  {ctx['col_a_at_row']!r}")
        if ctx.get("col_b_at_row") and ctx["col_b_at_row"] != ctx.get("row_label"):
            lines.append(f"  col B label:  {ctx['col_b_at_row']!r}")
        if ctx.get("label_above"):
            lines.append(f"  cell above:   {ctx['label_above']!r}")
        if ctx.get("label_below"):
            lines.append(f"  cell below:   {ctx['label_below']!r}")
        if ctx.get("column_header"):
            lines.append(f"  column hdr:   {ctx['column_header']!r}")
        if ctx.get("row_neighbors"):
            lines.append(f"  row neighbors: {ctx['row_neighbors']}")
    return "\n".join(lines)


def _candidates_substantially_agree(candidates: list[dict]) -> bool:
    """
    Return True if all validation-passing candidates have the same value
    (within 1% tolerance for numerics). Saves GPT calls when there's no real
    ambiguity — different sheets reporting the same number is fine.
    """
    passing = [c for c in candidates if c.get("passes_validation")]
    if len(passing) < 2:
        return True
    values = [c.get("value") for c in passing if c.get("value") is not None]
    if len(values) < 2:
        return True
    # Numeric agreement test
    try:
        nums = [float(v) for v in values]
        if max(nums) == min(nums):
            return True
        tol = 0.01 * max(abs(max(nums)), abs(min(nums)), 1)
        return (max(nums) - min(nums)) <= tol
    except (TypeError, ValueError):
        return all(v == values[0] for v in values)


def _gpt_pick(metric: dict, candidates: list[dict], context_by_idx: dict[int, dict]) -> dict:
    """Send candidates + context to GPT, get back chosen_index + reasoning."""
    blocks = [
        _format_candidate_for_prompt(i, c, context_by_idx.get(i, {}))
        for i, c in enumerate(candidates)
    ]
    user_msg = (
        f"METRIC: {metric['metric_name']}\n"
        f"DEFINITION: {metric.get('definition', '')}\n"
        f"EXPECTED UNIT: {metric.get('unit')}\n"
        f"EXPECTED PERIOD: {metric.get('period')}\n"
        f"PREFERRED SHEETS: {', '.join(metric.get('preferred_sheets', []) or [])}\n"
        f"\nCANDIDATES:\n\n" + "\n\n".join(blocks)
    )

    try:
        response = client.chat.completions.create(
            model=MODEL_FAST,
            temperature=0.0,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        return json.loads(raw)
    except json.JSONDecodeError as e:
        log.error("GPT picker JSON parse failed for %s: %s", metric["metric_name"], e)
        return {"chosen_index": None, "reasoning": "JSON parse error", "confidence": "low"}
    except Exception as e:
        log.error("GPT picker API failed for %s: %s", metric["metric_name"], e)
        return {"chosen_index": None, "reasoning": f"API error: {e}", "confidence": "low"}


def resolve_pool_with_gpt(record: dict, metric: dict, file_path: Path) -> dict:
    """
    Phase 2 — if `record` is candidate_pool, send candidates + context to GPT
    and pick the right one. Returns an updated record (or the original if
    GPT call wasn't possible).

    Skips the GPT call if:
      - LLM unavailable
      - candidates substantially agree on value (already verified-equivalent)
      - record status != "candidate_pool"
    """
    if record.get("status") != "candidate_pool":
        return record
    if not llm_available():
        return record

    candidates = record.get("candidates", [])
    passing = [c for c in candidates if c.get("passes_validation")]
    if len(passing) < 2:
        # Only 1 (or 0) passing — nothing to disambiguate
        return record

    if _candidates_substantially_agree(passing):
        # All passing candidates have ~same value; just promote to verified
        record["status"] = "verified"
        record["validation_notes"] = (record.get("validation_notes") or []) + [
            f"Promoted to verified — all {len(passing)} passing candidates agree on value."
        ]
        return record

    # Gather context per passing candidate
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        log.error("Could not open workbook for context gathering: %s", e)
        return record

    context_by_idx: dict[int, dict] = {}
    for i, c in enumerate(passing):
        context_by_idx[i] = _gather_candidate_context(wb, c)

    try:
        wb.close()
    except Exception:
        pass

    log.info(
        "GPT-resolve START for %s (%d passing candidates)",
        metric["metric_name"], len(passing),
    )
    result = _gpt_pick(metric, passing, context_by_idx)

    chosen_idx = result.get("chosen_index")
    reasoning  = result.get("reasoning", "")
    confidence = result.get("confidence", "low")

    if chosen_idx is None or not (0 <= chosen_idx < len(passing)):
        # GPT couldn't decide
        record["validation_notes"] = (record.get("validation_notes") or []) + [
            f"GPT resolver could not confidently pick a candidate. Reasoning: {reasoning[:160]}"
        ]
        log.info("GPT-resolve INCONCLUSIVE for %s — %s",
                 metric["metric_name"], reasoning[:80])
        return record

    chosen = passing[chosen_idx]
    # Apply the chosen candidate
    record["raw_value"]     = chosen.get("value")
    record["source_sheet"]  = chosen.get("sheet")
    record["source_cell"]   = chosen.get("value_cell")
    record["sheet_tier"]    = chosen.get("sheet_tier")
    record["extractor_confidence"] = chosen.get("confidence")

    # Apply scale correction if the chosen candidate needed it
    raw = chosen.get("value")
    scale_correction = chosen.get("scale_correction")
    if scale_correction == "1000":
        record["normalized_value"] = float(raw) * 1000
    elif scale_correction == "1000000":
        record["normalized_value"] = float(raw) * 1_000_000
    else:
        record["normalized_value"] = raw

    # Re-format display value
    from metric_resolver import _format_display
    record["display_value"] = _format_display(
        record["normalized_value"], metric.get("unit"), metric.get("scale"),
    )

    record["status"] = "verified"
    record["validation_notes"] = (record.get("validation_notes") or []) + [
        f"Phase 2 GPT resolver picked candidate {chosen_idx} ({confidence} confidence). "
        f"Reasoning: {reasoning[:200]}"
    ]
    log.info(
        "GPT-resolve PICKED %s for %s — %s",
        chosen.get("value_cell"), metric["metric_name"], reasoning[:80],
    )
    return record


# =============================================================================
# Comprehension verification — Phase 2.5 / C
# =============================================================================
#
# After extraction + reconciliation, send the full set of extracted values to
# GPT for a holistic "does this deal cohere?" review. This catches errors the
# deterministic identity checks miss — e.g. an exit value that's implausible
# given the NOI and cap rate, a unit count inconsistent with the price/unit, a
# property type that contradicts the metrics.
#
# Unlike the per-metric picker, this reasons over the WHOLE deal at once. It
# does not silently change values — it FLAGS them, appending to validation_notes
# and (for clear contradictions) downgrading verified→suspicious. The human and
# the memo then see the flag rather than a confident-wrong number.

COMPREHENSION_SYSTEM = f"""\
You are a senior real estate analyst doing a sanity review of an automated
extraction. You are given the metrics pulled from one underwriting model.
Your job: decide whether the numbers COHERE as a single real deal.

{IDENTITY_RELATIONSHIPS}

Flag a metric ONLY when it is clearly inconsistent with the others or
implausible for the asset (wrong by an order of magnitude, fails an identity
by a wide margin, contradicts the property type, etc.). Do NOT flag values
that are merely missing, nor nitpick small rounding differences.

Return ONLY JSON:
{{
  "flags": [
    {{"metric": "<exact metric name>",
      "issue": "one sentence — what's inconsistent and with what",
      "severity": "high" | "medium"}}
  ],
  "overall": "one sentence — does the deal cohere as extracted?"
}}
No prose, no code fences. If everything coheres, return "flags": [].
"""


def run_comprehension_review(bounded_metrics: dict) -> dict:
    """
    Holistic GPT coherence review over all extracted bounded metrics.

    Returns {"flags": [...], "overall": str}. Empty flags = coheres / skipped.
    Mutates nothing — the caller decides how to apply flags.
    """
    if not llm_available() or not bounded_metrics:
        return {"flags": [], "overall": ""}

    # Build a compact value table for GPT (only metrics that have a value)
    lines = []
    for name, rec in bounded_metrics.items():
        status = rec.get("status")
        if status in ("missing",):
            continue
        val = rec.get("display_value", "—")
        lines.append(f"  {name}: {val}  [{status}]")
    if not lines:
        return {"flags": [], "overall": ""}

    user_msg = "Extracted metrics for one deal:\n\n" + "\n".join(lines)

    try:
        response = client.chat.completions.create(
            model=MODEL_FAST,
            temperature=0.0,
            messages=[
                {"role": "system", "content": COMPREHENSION_SYSTEM},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        parsed = json.loads(raw)
        flags = parsed.get("flags", []) or []
        log.info(
            "Comprehension review — %d flag(s). Overall: %s",
            len(flags), str(parsed.get("overall", ""))[:100],
        )
        return parsed
    except json.JSONDecodeError as e:
        log.error("Comprehension review JSON parse failed: %s", e)
        return {"flags": [], "overall": ""}
    except Exception as e:
        log.error("Comprehension review API failed: %s", e)
        return {"flags": [], "overall": ""}
