"""
metric_fallback.py — Phase 1.5b GPT-as-reader fallback.

Runs only when Pass 1 catalog returned ZERO candidates for a bounded metric.
For each such metric, sends a tight per-metric GPT call against the top-priority
sheet(s) and either:
  - returns a new candidate record (which the resolver then validates), OR
  - reports definitively "not found in file" (so the memo doesn't lie).

Designed to be cheap: one GPT call per missing metric, each scoped to a
single sheet rendered as a compact markdown table (~1-3K tokens). With ~5
missing metrics per file, total fallback cost is ~$0.01-0.02.

The fallback uses gpt-4o-mini and is silently skipped when no OpenAI key is set.
"""
from __future__ import annotations
import json
import logging
import sys
from pathlib import Path
from typing import Any

from scenarios._llm import client, MODEL_FAST, llm_available
from flexible_extractor import sorted_sheets_by_priority

log = logging.getLogger("fb.fallback")
if not log.handlers:
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("[fb.fallback] %(asctime)s %(levelname)s %(message)s"))
    log.addHandler(h)
    log.setLevel(logging.INFO)


FALLBACK_VERSION = "phase1_5b.v1"

# Cap on raw sheet content sent per call. Most metrics will be found in the
# first 60 rows of a summary sheet anyway.
_MAX_ROWS_PER_SHEET = 80
_MAX_COLS_PER_SHEET = 20


SYSTEM_PROMPT = """\
You are a precise data extractor for a real estate underwriting model.
You will be shown ONE sheet from an Excel workbook and asked to find ONE specific metric.

CRITICAL RULES:
- Return only what is literally in the sheet. If the value is not present, say not found.
- DO NOT SUBSTITUTE. If the row whose label matches the requested metric has an
  EMPTY value cell, the metric is NOT FOUND — never borrow a value from an
  adjacent or nearby row that is a DIFFERENT metric. Example: if "Occupancy %"
  is blank but the next row "Discount to Replacement Cost %" has 17.5%, the
  occupancy is NOT FOUND — do not return 17.5%.
- The matched row label must be SEMANTICALLY the requested metric, not merely
  near it or numerically plausible.
- The value must come from a single cell. Cite the cell (e.g. "C11").
- If the metric is ambiguous (multiple plausible cells), pick the one whose ROW LABEL
  is most semantically specific to the requested metric. Cite your reasoning.
- For ranges (rents, units), pick the TOTAL / aggregate figure unless the metric
  schema explicitly requests something else.
- Numbers must be returned as numbers, dates as ISO 8601 strings, text as text.

Return ONLY JSON of this shape:
{
  "found": true,
  "value": <number | string>,
  "cell": "C11",
  "label_in_sheet": "100% Purchase Price (Local Currency)",
  "reasoning": "Row 11 label specifies '100%' meaning the full deal-level purchase price, vs row 22 which is per-property."
}
OR if not found:
{
  "found": false,
  "reasoning": "No row in this sheet labels Purchase Price. Searched rows 1-80."
}
No prose outside the JSON. No code fences.
"""


def _sheet_to_text_block(file_path: Path, sheet_name: str) -> str:
    """Render a single sheet as compact text grid for GPT."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return f"(could not load workbook: {e})"

    if sheet_name not in wb.sheetnames:
        return f"(sheet {sheet_name!r} not found in workbook)"

    ws = wb[sheet_name]
    lines: list[str] = []
    for r in range(1, min(ws.max_row, _MAX_ROWS_PER_SHEET) + 1):
        for c in range(1, min(ws.max_column, _MAX_COLS_PER_SHEET) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or str(v).strip() == "":
                continue
            cell_ref = openpyxl.utils.get_column_letter(c) + str(r)
            # Truncate very long string cells
            s = str(v)
            if len(s) > 60:
                s = s[:57] + "..."
            lines.append(f"  {cell_ref}: {s}")
    wb.close()
    return "\n".join(lines)


def _resolve_target_sheets(metric: dict, available_sheets: list[str]) -> list[str]:
    """
    Translate metric.preferred_sheets keywords into actual sheet names that
    exist in this workbook. Returns up to 2 highest-priority matches.
    """
    preferred = metric.get("preferred_sheets") or []
    if not preferred:
        return []

    # Use sorted_sheets_by_priority to drop SKIP-tier and rank by tier
    ranked = sorted_sheets_by_priority(available_sheets, exclude_skipped=True)

    matches: list[str] = []
    for pref_kw in preferred:
        kw_lower = pref_kw.lower()
        for sheet in ranked:
            if kw_lower in sheet.lower() and sheet not in matches:
                matches.append(sheet)
                break  # first match per keyword
        if len(matches) >= 2:
            break

    # Catch-all: if the metric's preferred sheet keywords matched NOTHING in
    # this file (the common "no target sheets matched preferred list" SKIP on
    # complex models), fall back to the top-priority sheets so the metric still
    # gets a read attempt rather than going silently missing.
    if not matches:
        matches = ranked[:2]

    return matches[:2]


def _system_with_rules() -> str:
    """SYSTEM_PROMPT + any active metric_resolution knowledge patterns."""
    from knowledge_store import with_active_rules
    return with_active_rules(SYSTEM_PROMPT, ["metric_resolution"])


def _one_metric_gpt_call(metric: dict, sheet_name: str, sheet_text: str) -> dict:
    """Run a single GPT call for one metric on one sheet."""
    user_msg = (
        f"METRIC TO FIND: {metric['metric_name']}\n"
        f"DEFINITION:     {metric.get('definition', '(none)')}\n"
        f"EXPECTED UNIT:  {metric.get('unit', 'unknown')}\n"
        f"EXPECTED PERIOD:{metric.get('period', 'any')}\n"
        f"VALID RANGE:    [{metric.get('range_min')}, {metric.get('range_max')}]\n"
        f"\n"
        f"SHEET: {sheet_name}\n"
        f"{sheet_text}\n"
    )

    try:
        response = client.chat.completions.create(
            model=MODEL_FAST,
            temperature=0.0,
            messages=[
                {"role": "system", "content": _system_with_rules()},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        return json.loads(raw)
    except json.JSONDecodeError as e:
        log.error("Fallback JSON parse failed for %s: %s", metric["metric_name"], e)
        return {"found": False, "reasoning": f"GPT response unparseable: {e}"}
    except Exception as e:
        log.error("Fallback API call failed for %s: %s", metric["metric_name"], e)
        return {"found": False, "reasoning": f"API error: {e}"}


def fallback_find_metric(
    metric: dict,
    file_path: Path,
    available_sheets: list[str],
) -> dict | None:
    """
    Try to find this metric via targeted GPT read of top-priority sheets.

    Returns a CANDIDATE-shaped dict (compatible with resolve_metric input):
        {
          metric_id, metric_name, value, source_file, sheet, sheet_tier,
          label_cell, value_cell, matched_alias, confidence,
          label_ratio, match_method
        }
    or None if GPT could not find the metric in any target sheet.
    """
    if not llm_available():
        return None

    target_sheets = _resolve_target_sheets(metric, available_sheets)
    if not target_sheets:
        log.info("Fallback SKIP for %s — no target sheets matched preferred list",
                 metric["metric_name"])
        return None

    for sheet in target_sheets:
        sheet_text = _sheet_to_text_block(file_path, sheet)
        if not sheet_text:
            continue

        log.info("Fallback ATTEMPT for %s on sheet %r", metric["metric_name"], sheet)
        result = _one_metric_gpt_call(metric, sheet, sheet_text)

        if result.get("found"):
            log.info(
                "Fallback FOUND for %s on %s!%s — %r",
                metric["metric_name"], sheet, result.get("cell"),
                str(result.get("value"))[:50],
            )
            # Build a candidate-shaped record so resolve_metric can validate
            from flexible_extractor import sheet_priority_tier
            return {
                "metric_id":     metric["metric_id"],
                "metric_name":   metric["metric_name"],
                "value":         result.get("value"),
                "source_file":   file_path.name,
                "sheet":         sheet,
                "sheet_tier":    sheet_priority_tier(sheet),
                "label_cell":    result.get("cell"),
                "value_cell":    result.get("cell"),
                "matched_alias": result.get("label_in_sheet", "(GPT fallback)"),
                "confidence":    "gpt_fallback",
                "label_ratio":   1.0,
                "match_method":  "fallback",
                "fallback_reasoning": result.get("reasoning"),
            }

    log.info("Fallback NOT FOUND for %s after %d sheet(s)",
             metric["metric_name"], len(target_sheets))
    return None
