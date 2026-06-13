"""
workbook_orientation.py — lightweight Workbook Orientation stage (pre-AAM).

Before searching for any metric, a human analyst first orients: which sheet is
the summary, which holds the assumptions, which holds returns, which is the
actual financial model, and which are supporting schedules. This module does
that orientation deterministically — no GPT, no API key — by reading each
sheet's CONTENT and STRUCTURE:

    keyword concentration   what the sheet talks about (acquisition metrics,
                            returns vocabulary, schedule vocabulary, ...)
    hardcode ratio          inputs sheets are typed-in numbers, not formulas
    internal formula load   model sheets compute; summaries mostly reference
    cross-sheet ref flow    a summary PULLS values from other sheets; an
                            inputs sheet is REFERENCED BY other sheets
    time-series shape       model sheets carry wide dated roll-up tables

Sheet names are used only as one WEAK signal (a small score bonus) — a sheet
named "Tab 7" can be the real summary, and a sheet named "Summary" can be junk.

The output is a workbook map:

    {
      "version": ORIENTATION_VERSION,
      "sheets": { name: {"role", "confidence", "scores", "signals"} },
      "map":    { "summary": [...], "inputs": [...], "returns": [...],
                  "model": [...], "support": [...], "other": [...] },
    }

and `orientation_tier_map()` converts it into the {sheet: tier} dict the
existing extraction machinery already consumes (scan_workbook_for_candidates /
_stage1_context), with the search order the analyst uses:

    summary → 1, inputs → 2, returns → 3   (Stage-1 whitelist, searched first)
    model → 4, support → 5, other → 6      (Stage-2 territory)

Name-based SKIP sheets (sensitivity / comps / backup, tier 99) stay skipped
regardless of orientation — those tabs poison cell-matching by design, and a
content heuristic must not resurrect them.

This stage is non-destructive: it produces a map and a tier dict, nothing else.
If orientation fails, callers pass None and extraction behaves exactly as today.
"""
from __future__ import annotations

import logging
import re
import sys
from pathlib import Path
from typing import Any

from flexible_extractor import sheet_priority_tier

log = logging.getLogger("fb.orientation")
if not log.handlers:
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("[fb.orient] %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

ORIENTATION_VERSION = "2026-06-12.1"

# Scan bounds — match the extractor's window so orientation sees the same cells
# extraction will see.
_MAX_ROWS = 250
_MAX_COLS = 60

ROLES = ("summary", "inputs", "returns", "model", "support", "other")

# Role → extraction tier. Summary / inputs / returns land inside the Stage-1
# whitelist (tier ≤ 3) in exactly the order an analyst searches; model and
# support sheets are Stage-2 territory (formula trace reaches into them).
ROLE_TIER = {
    "summary": 1,
    "inputs":  2,
    "returns": 3,
    "model":   4,
    "support": 5,
    "other":   6,
}

# ---------------------------------------------------------------------------
# Content vocabulary (lowercase substring match against text cells ≤ 60 chars).
# Multi-word phrases preferred — short tokens ("sf", "cf") false-match too much.
# ---------------------------------------------------------------------------

_KW_ACQ = (
    "purchase price", "acquisition price", "acquisition cost",
    "total project cost", "total basis", "total cost", "basis build",
    "going in cap", "cap rate", "closing date", "closing cost",
    "transaction cost", "sources", "uses of funds", "hold period", "exit cap",
)
_KW_RET = (
    "irr", "xirr", "equity multiple", "moic", "npv",
    "waterfall", "promote", "preferred return", "distribution",
    "cash-on-cash", "cash on cash", "yield on cost",
)
_KW_DEBT = (
    "loan amount", "ltv", "dscr", "debt yield", "interest rate",
    "spread", "amortization", "maturity", "debt service",
)
_KW_IDENT = (
    "property name", "asset type", "property type", "location",
    "year built", "units", "square feet", "square footage", "keys",
)
_KW_INPUT = (
    # Deliberately EXCLUDES one-pager display items (hold period, exit cap):
    # a summary sheet shows those too, and they must not drag it into 'inputs'.
    "assumption", "rent growth", "expense growth", "growth rate",
    "vacancy", "inflation", "escalation", "renovation", "reserve",
)
_KW_MODEL = (
    "revenue", "rental income", "total income", "effective gross",
    "operating expense", "total expense", "net operating income", "noi",
    "capex", "capital expenditure", "cash flow", "debt service",
    "occupancy",
)
_KW_SUPPORT = (
    "rent roll", "unit mix", "tenant", "lease", "suite",
    "expiration", "stacking", "debt schedule", "amort schedule",
    "draw schedule", "loan schedule",
)

# Cross-sheet reference patterns inside formula strings: 'My Sheet'!A1 / Inputs!B2
_QUOTED_REF = re.compile(r"'([^']+)'!")
_BARE_REF = re.compile(r"\b([A-Za-z_][A-Za-z0-9_.]*)!")

# Sheets with fewer populated cells than this are classified "other" outright.
_MIN_POPULATED = 15
# A top score below this is too weak to assert a role.
_MIN_SCORE = 4.0


def _count_hits(texts: list[str], keywords: tuple[str, ...], cap: int = 12) -> int:
    """How many label cells mention any of `keywords` (each cell counts once)."""
    n = 0
    for t in texts:
        if any(kw in t for kw in keywords):
            n += 1
            if n >= cap:
                break
    return n


def _scan_values(ws) -> dict[str, Any]:
    """Values pass: label texts, numeric coordinates, time-series shape."""
    import datetime as _dt

    texts: list[str] = []
    numeric_coords: set[str] = set()
    ts_rows = 0
    date_header = False

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=_MAX_ROWS, min_col=1, max_col=_MAX_COLS),
        start=1,
    ):
        run = 0          # current consecutive-numeric run in this row
        best_run = 0
        n_dates = 0
        year_run = 0
        best_year_run = 0
        for cell in row:
            v = cell.value
            if v is None:
                run = 0
                year_run = 0
                continue
            if isinstance(v, str):
                # Normalize separators so "Yield-on-Cost" / "Going_In_Cap"
                # match the space-separated keyword vocabulary.
                s = v.strip().lower().replace("-", " ").replace("_", " ").replace("/", " ")
                if s and len(s) <= 60:
                    texts.append(s)
                run = 0
                year_run = 0
            elif isinstance(v, bool):
                run = 0
                year_run = 0
            elif isinstance(v, (_dt.datetime, _dt.date)):
                n_dates += 1
                run = 0
                year_run = 0
            elif isinstance(v, (int, float)):
                numeric_coords.add(cell.coordinate)
                run += 1
                best_run = max(best_run, run)
                # Header rows of annual roll-ups are often plain year integers.
                if isinstance(v, int) and 1990 <= v <= 2100:
                    year_run += 1
                    best_year_run = max(best_year_run, year_run)
                else:
                    year_run = 0
        if best_run >= 6:
            ts_rows += 1
        if r_idx <= 40 and (n_dates >= 4 or best_year_run >= 4):
            date_header = True

    return {
        "texts": texts,
        "numeric_coords": numeric_coords,
        "ts_rows": ts_rows,
        "date_header": date_header,
    }


def _scan_formulas(ws, sheet_name: str, all_sheets_lower: dict[str, str]):
    """Formula pass: formula coordinates + cross-sheet pulls (this sheet → others)."""
    formula_coords: set[str] = set()
    pulls = 0                       # formulas referencing ANOTHER sheet
    pulled_from: dict[str, int] = {}  # target sheet → ref count (for in_refs)

    for row in ws.iter_rows(min_row=1, max_row=_MAX_ROWS, min_col=1, max_col=_MAX_COLS):
        for cell in row:
            v = cell.value
            is_formula = (isinstance(v, str) and v.startswith("=")) or \
                v.__class__.__name__ == "ArrayFormula"
            if not is_formula:
                continue
            formula_coords.add(cell.coordinate)
            f = v if isinstance(v, str) else str(getattr(v, "text", "") or "")
            targets = set(_QUOTED_REF.findall(f)) | set(_BARE_REF.findall(f))
            hit_other = False
            for t in targets:
                real = all_sheets_lower.get(t.lower())
                if real and real != sheet_name:
                    pulled_from[real] = pulled_from.get(real, 0) + 1
                    hit_other = True
            if hit_other:
                pulls += 1

    return formula_coords, pulls, pulled_from


def _score_sheet(sig: dict[str, Any], name_tier: int) -> dict[str, float]:
    """Combine signals into per-role scores. Weights are heuristic v1 — kept as
    plain arithmetic so the debug panel can show WHY a sheet got its role."""
    texts = sig["texts"]
    acq = _count_hits(texts, _KW_ACQ)
    ret = _count_hits(texts, _KW_RET)
    debt = _count_hits(texts, _KW_DEBT)
    ident = _count_hits(texts, _KW_IDENT)
    inp = _count_hits(texts, _KW_INPUT)
    model = _count_hits(texts, _KW_MODEL)
    support = _count_hits(texts, _KW_SUPPORT)

    n_numeric = len(sig["numeric_coords"])
    n_formula = len(sig["formula_coords"])
    hardcoded = len(sig["numeric_coords"] - sig["formula_coords"])
    hardcode_ratio = hardcoded / n_numeric if n_numeric else 0.0
    internal_formulas = n_formula - sig["pulls"]
    populated = n_numeric + len(texts)
    internal_density = internal_formulas / populated if populated else 0.0
    text_share = len(texts) / populated if populated else 0.0

    # Summary: breadth across metric categories, pulls values from elsewhere
    # ("many inbound references, relatively few calculations"), label-dense.
    # NOTE: real one-pagers DO carry an annual cash-flow roll-up table (it is a
    # standard summary component), so time-series rows are only penalized
    # lightly and only beyond what a roll-up block explains. A one-pager
    # without return metrics isn't a one-pager → missing returns halves it.
    breadth = sum(1 for h in (acq, ret, debt, ident) if h > 0)
    pulls_norm = min(sig["pulls"], 40) / 40.0
    pulls_pen = min(sig["pulls"], 200) / 200.0   # wider range, for penalties
    in_norm = min(sig["in_refs"], 60) / 60.0
    summary_score = (
        1.2 * (acq + ret + debt + ident) + 3.0 * breadth
        + 7.0 * pulls_norm + 2.5 * text_share
        - 0.1 * max(0, sig["ts_rows"] - 30) - populated / 1500.0
    )
    if ret == 0:
        summary_score *= 0.5

    scores = {
        "summary": summary_score,
        # Inputs: assumption vocabulary + typed-in numbers + referenced by
        # others. PENALIZED by pulls: an inputs sheet feeds the model, it does
        # not pull hundreds of values FROM it — a one-pager (pull-heavy, partly
        # hardcoded) must not win 'inputs' because its display cells are typed.
        "inputs": 2.0 * inp + 0.8 * (acq + debt)
                  + (8.0 * hardcode_ratio if n_numeric >= 15 else 0.0)
                  + 5.0 * in_norm - 8.0 * pulls_pen,
        # Returns: returns vocabulary, with a floor so one stray "IRR" label
        # (e.g. in a sensitivity grid) can't claim the role by itself.
        "returns": 2.5 * ret if ret >= 2 else 0.0,
        # Model: flow vocabulary + wide dated time series + internal computation.
        "model": 0.8 * model + 0.6 * min(sig["ts_rows"], 20)
                 + (3.0 if sig["date_header"] else 0.0) + 5.0 * internal_density,
        "support": 2.5 * support,
    }

    # Sheet NAME as one supporting signal (the brief: names may be used as one
    # signal; content carries more total weight). Explicit tier-1 names
    # ("One Pager", "Executive Summary") get a slightly stronger nudge — they
    # are rarely lies, and content noise (an index matrix on the exec summary)
    # must not flip them into 'model'.
    name_hint, hint_w = {
        1: ("summary", 4.0), 2: ("inputs", 3.0), 3: ("summary", 2.0),
        4: ("model", 2.0), 5: ("support", 2.0), 6: ("support", 2.0),
        7: ("returns", 3.0),
    }.get(name_tier, (None, 0.0))
    if name_hint:
        scores[name_hint] += hint_w

    sig.update({
        "kw": {"acq": acq, "returns": ret, "debt": debt, "identity": ident,
               "inputs": inp, "model": model, "support": support},
        "n_numeric": n_numeric, "n_formula": n_formula,
        "hardcode_ratio": round(hardcode_ratio, 2),
        "pulls": sig["pulls"], "in_refs": sig["in_refs"],
        "populated": populated,
    })
    return scores


def orient_workbook(file_path: str | Path) -> dict[str, Any]:
    """
    Run Workbook Orientation. Deterministic, read-only, no GPT.

    Returns the workbook map described in the module docstring, or
    {"error": "..."} when the file can't be read (callers fall back to the
    existing name-based behavior).
    """
    import openpyxl

    file_path = Path(file_path)
    try:
        wb_vals = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        wb_form = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    except Exception as e:
        log.error("Orientation failed to open %s: %s", file_path.name, e)
        return {"error": f"Could not open workbook: {e}", "version": ORIENTATION_VERSION}

    all_sheets = list(wb_vals.sheetnames)
    sheets_lower = {s.lower(): s for s in all_sheets}

    # Pass 1+2 per sheet: values signals + formula signals.
    signals: dict[str, dict] = {}
    inbound: dict[str, int] = {s: 0 for s in all_sheets}
    for name in all_sheets:
        try:
            sig = _scan_values(wb_vals[name])
        except Exception:  # chartsheets etc.
            sig = {"texts": [], "numeric_coords": set(), "ts_rows": 0,
                   "date_header": False}
        try:
            fcoords, pulls, pulled_from = _scan_formulas(
                wb_form[name], name, sheets_lower)
        except Exception:
            fcoords, pulls, pulled_from = set(), 0, {}
        sig["formula_coords"] = fcoords
        sig["pulls"] = pulls
        for target, n in pulled_from.items():
            inbound[target] = inbound.get(target, 0) + n
        signals[name] = sig

    for p in (wb_vals, wb_form):
        try:
            p.close()
        except Exception:
            pass

    # Score and classify.
    result_sheets: dict[str, dict] = {}
    role_map: dict[str, list[str]] = {r: [] for r in ROLES}
    for name in all_sheets:
        sig = signals[name]
        sig["in_refs"] = inbound.get(name, 0)
        name_tier = sheet_priority_tier(name)
        scores = _score_sheet(sig, name_tier)

        ranked = sorted(scores.items(), key=lambda kv: -kv[1])
        top_role, top = ranked[0]
        second = ranked[1][1] if len(ranked) > 1 else 0.0
        if name_tier == 99:
            # Name-flagged junk (sensitivity / comps / backup) is excluded from
            # extraction no matter what its content resembles — sensitivity
            # grids LOOK like returns sheets, which is exactly the trap.
            role, confidence = "other", 0.0
        elif sig["populated"] < _MIN_POPULATED or top < _MIN_SCORE:
            role, confidence = "other", 0.2
        else:
            role = top_role
            confidence = round(min(1.0, max(0.0, (top - second) / top) + 0.15), 2)

        result_sheets[name] = {
            "role": role,
            "confidence": confidence,
            "scores": {k: round(v, 1) for k, v in scores.items()},
            "signals": {
                "kw": sig["kw"],
                "hardcode_ratio": sig["hardcode_ratio"],
                "pulls_from_other_sheets": sig["pulls"],
                "referenced_by_other_sheets": sig["in_refs"],
                "time_series_rows": sig["ts_rows"],
                "date_header": sig["date_header"],
                "populated_cells": sig["populated"],
            },
        }
        role_map[role].append(name)

    # Order each role group best-first (confidence desc) for display + nomination.
    for r in role_map:
        role_map[r].sort(key=lambda s: -result_sheets[s]["confidence"])

    log.info(
        "Orientation for %s — summary: %s | inputs: %s | returns: %s | model: %s | support: %s",
        file_path.name,
        ", ".join(role_map["summary"]) or "—",
        ", ".join(role_map["inputs"]) or "—",
        ", ".join(role_map["returns"]) or "—",
        ", ".join(role_map["model"]) or "—",
        ", ".join(role_map["support"]) or "—",
    )
    return {"version": ORIENTATION_VERSION, "sheets": result_sheets, "map": role_map}


def render_sheets_text(
    file_path: str | Path,
    sheet_names: list[str],
    formula_cells: dict[str, set[str]] | None = None,
    max_rows_per_sheet: int = 200,
    max_chars_per_sheet: int = 14_000,
    max_total_chars: int = 48_000,
) -> str:
    """
    Render whole sheets as compact text for an LLM read — the `extract-text`
    equivalent. Unlike labeled-pair extraction, this PRESERVES table structure
    (stacked line items, multi-column layouts, roll-up blocks), so the model
    reads the sheet the way an analyst does.

    Format (one line per non-empty row, every cell carries its A1 ref so the
    model can cite the exact cell — claims are later grounded against the
    workbook):

        === SHEET: One Pager ===
        R5: B5='Purchase Price' | C5=224100* | D5=892828.69
        ...

    A trailing `*` marks a hard-coded numeric cell (a typed-in modeler input,
    not a formula) when `formula_cells` is provided.

    Returns "" when nothing could be rendered (caller falls back to pairs).
    """
    import datetime as _dt
    import openpyxl

    file_path = Path(file_path)
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        log.error("Sheet render failed to open %s: %s", file_path.name, e)
        return ""

    blocks: list[str] = []
    total = 0
    for name in sheet_names:
        if name not in wb.sheetnames or total >= max_total_chars:
            continue
        try:
            ws = wb[name]
        except Exception:
            continue
        fset = (formula_cells or {}).get(name)
        lines = [f"=== SHEET: {name} ==="]
        n_chars = 0
        rows_out = 0
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=_MAX_ROWS, min_col=1, max_col=_MAX_COLS),
            start=1,
        ):
            if rows_out >= max_rows_per_sheet or n_chars >= max_chars_per_sheet:
                break
            cells: list[str] = []
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str):
                    s = v.strip()
                    if not s:
                        continue
                    cells.append(f"{cell.coordinate}='{s[:80]}'")
                elif isinstance(v, bool):
                    cells.append(f"{cell.coordinate}={v}")
                elif isinstance(v, (_dt.datetime, _dt.date)):
                    d = v.date() if isinstance(v, _dt.datetime) else v
                    cells.append(f"{cell.coordinate}={d.isoformat()}")
                elif isinstance(v, (int, float)):
                    mark = "*" if (fset is not None and cell.coordinate not in fset) else ""
                    num = f"{v:.6g}" if isinstance(v, float) else str(v)
                    cells.append(f"{cell.coordinate}={num}{mark}")
            if not cells:
                continue
            line = f"R{r_idx}: " + " | ".join(cells)
            lines.append(line)
            n_chars += len(line)
            rows_out += 1
        if rows_out:
            block = "\n".join(lines)
            blocks.append(block)
            total += len(block)

    try:
        wb.close()
    except Exception:
        pass
    return "\n\n".join(blocks)[:max_total_chars]


def orientation_tier_map(orientation: dict[str, Any] | None) -> dict[str, int] | None:
    """
    Convert an orientation result into the {sheet: tier} dict the extraction
    machinery consumes. Returns None when orientation is missing/failed so
    callers fall back to today's behavior unchanged.

    Name-flagged skip sheets (sensitivity / comps / backup → tier 99) STAY
    skipped: orientation prioritizes real sheets, it never resurrects junk.
    """
    if not orientation or orientation.get("error") or not orientation.get("sheets"):
        return None
    tiers: dict[str, int] = {}
    for name, info in orientation["sheets"].items():
        if sheet_priority_tier(name) == 99:
            tiers[name] = 99
        else:
            tiers[name] = ROLE_TIER.get(info.get("role"), 6)
    return tiers


def select_read_sheets(
    tier_map: dict[str, int],
    max_sheets: int = 8,
    quotas: tuple[tuple[int, int], ...] = ((1, 4), (2, 2), (3, 2)),
) -> list[str]:
    """
    Pick the sheets to read WHOLE — the analyst's short stack: a few summary
    sheets, the inputs sheet(s), a secondary summary or two. Within each tier
    the NAME tier breaks ties (an explicit "One Pager" / "Executive Summary"
    name beats template tabs that content-classified into the same tier),
    mirroring the resolver's own tiebreak. Backfills to max_sheets from the
    remaining whitelisted sheets when a tier is thin.
    """
    by_name_rank = lambda n: (sheet_priority_tier(n), n)
    picked: list[str] = []
    for tier, quota in quotas:
        group = sorted((n for n, t in tier_map.items() if t == tier), key=by_name_rank)
        picked.extend(group[:quota])
    if len(picked) < max_sheets:
        rest = sorted(
            (n for n, t in tier_map.items() if t != 99 and n not in picked),
            key=lambda n: (tier_map[n], *by_name_rank(n)),
        )
        picked.extend(rest[: max_sheets - len(picked)])
    return picked[:max_sheets]


def analyst_reading_stack(
    file_path: str | Path,
    max_sheets: int = 5,
    max_total_chars: int = 22_000,
) -> tuple[list[str], str]:
    """
    One-call helper: orient the workbook, pick the analyst's reading stack,
    and render those sheets whole. Used by narrative generation (Snapshot) to
    see the actual pages an analyst reads — basis build-ups, threshold notes,
    deal story — not just extracted metrics.

    Falls back to name-based tiers when orientation fails. Returns
    ([sheet names], rendered_text); ("", []) means nothing could be read.
    No hardcode marks (no formula pass) — this is reading context, not
    input-vs-derived analysis — which keeps it fast.
    """
    file_path = Path(file_path)
    tier_map = orientation_tier_map(orient_workbook(file_path))
    if tier_map is None:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            tier_map = {n: sheet_priority_tier(n) for n in wb.sheetnames}
            wb.close()
        except Exception:
            return [], ""
    sheets = select_read_sheets(tier_map, max_sheets=max_sheets)
    if not sheets:
        return [], ""
    text = render_sheets_text(
        file_path, sheets, formula_cells=None,
        max_chars_per_sheet=max(4_000, max_total_chars // max(len(sheets), 1)),
        max_total_chars=max_total_chars,
    )
    return sheets, text
