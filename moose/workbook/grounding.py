"""Code grounding validation for workbook claims."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .workbook_result import ClaimGroundingResult, WorkbookEvidencePackResult


class WorkbookClaimGroundingValidator:
    """Validate that claim source references are grounded in workbook context."""

    def validate(
        self,
        file_path: str | Path,
        claims: list[dict[str, Any]],
        evidence_pack: WorkbookEvidencePackResult,
    ) -> ClaimGroundingResult:
        """Return claims whose sheet/cell/value/evidence references are grounded."""
        path = Path(file_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            grounded: list[dict[str, Any]] = []
            rejected: list[dict[str, Any]] = []
            errors: list[str] = []

            for claim in claims:
                claim_errors = self._validate_claim(workbook, evidence_pack, claim)
                if claim_errors:
                    rejected.append({"claim": claim, "errors": claim_errors})
                    errors.extend(claim_errors)
                else:
                    grounded.append(claim)

            return ClaimGroundingResult(
                grounded_claims=grounded,
                rejected_claims=rejected,
                errors=errors,
            )
        finally:
            workbook.close()

    def _validate_claim(
        self,
        workbook: Any,
        evidence_pack: WorkbookEvidencePackResult,
        claim: dict[str, Any],
    ) -> list[str]:
        errors: list[str] = []
        source_location = claim.get("source_location")
        if not isinstance(source_location, dict):
            return ["source_location must include sheet and cell."]

        sheet_name = source_location.get("sheet")
        cell_ref = source_location.get("cell")
        if not sheet_name or sheet_name not in workbook.sheetnames:
            return [f"referenced sheet does not exist: {sheet_name}"]

        row_ref = source_location.get("row")
        if not cell_ref and row_ref:
            worksheet = workbook[sheet_name]
            if not isinstance(row_ref, int) or row_ref < 1 or row_ref > (worksheet.max_row or 0):
                return [f"referenced row does not exist: {sheet_name}!row{row_ref}"]
            if not self._row_in_evidence_pack(evidence_pack, sheet_name, row_ref) and not claim.get("fallback_only"):
                return [f"referenced row is outside the bounded evidence pack: {sheet_name}!row{row_ref}"]
            return []

        if not cell_ref:
            return ["source_location.cell is required."]
        if (
            not self._cell_in_evidence_pack(evidence_pack, sheet_name, cell_ref)
            and not claim.get("fallback_only")
        ):
            return [f"referenced cell is outside the bounded evidence pack: {sheet_name}!{cell_ref}"]

        worksheet = workbook[sheet_name]
        try:
            source_cell = worksheet[cell_ref]
        except ValueError:
            return [f"referenced cell is invalid: {cell_ref}"]

        if not self._values_match(source_cell.value, claim.get("value")):
            errors.append(f"claim value does not match workbook cell {sheet_name}!{cell_ref}.")

        nearby_label = source_location.get("nearby_label")
        if nearby_label and not self._label_near_cell(worksheet, source_cell.row, source_cell.column, nearby_label):
            errors.append(f"nearby label is not grounded near {sheet_name}!{cell_ref}.")

        return errors

    def _cell_in_evidence_pack(
        self,
        evidence_pack: WorkbookEvidencePackResult,
        sheet_name: str,
        cell_ref: str,
    ) -> bool:
        sampled_cells = evidence_pack.sampled_cells.get(sheet_name, [])
        if any(item.get("cell") == cell_ref for item in sampled_cells):
            return True

        for neighborhood in evidence_pack.candidate_neighborhoods:
            if neighborhood.get("sheet") != sheet_name:
                continue
            nearby_values = neighborhood.get("nearby_values") or []
            if any(item.get("cell") == cell_ref for item in nearby_values):
                return True
        return False

    def _row_in_evidence_pack(
        self,
        evidence_pack: WorkbookEvidencePackResult,
        sheet_name: str,
        row_ref: int,
    ) -> bool:
        sampled_cells = evidence_pack.sampled_cells.get(sheet_name, [])
        if any(item.get("row") == row_ref for item in sampled_cells):
            return True

        header_blocks = evidence_pack.section_header_blocks.get(sheet_name, [])
        if any(item.get("row") == row_ref for item in header_blocks):
            return True

        for neighborhood in evidence_pack.candidate_neighborhoods:
            if neighborhood.get("sheet") != sheet_name:
                continue
            row_context = neighborhood.get("row_context") or []
            if any(self._cell_row(item.get("cell")) == row_ref for item in row_context):
                return True
        return False

    def _cell_row(self, cell_ref: Any) -> int | None:
        if not isinstance(cell_ref, str):
            return None
        digits = "".join(ch for ch in cell_ref if ch.isdigit())
        return int(digits) if digits else None

    def _values_match(self, workbook_value: Any, claim_value: Any) -> bool:
        if isinstance(workbook_value, (int, float)) and isinstance(claim_value, (int, float)):
            return abs(float(workbook_value) - float(claim_value)) <= 0.000001
        return workbook_value == claim_value

    def _label_near_cell(self, worksheet: Any, row: int, col: int, label: str) -> bool:
        label_normalized = label.strip().lower()
        for candidate_row in range(max(1, row - 1), row + 2):
            for candidate_col in range(max(1, col - 2), col + 1):
                value = worksheet.cell(candidate_row, candidate_col).value
                if isinstance(value, str) and value.strip().lower() == label_normalized:
                    return True
        return False
