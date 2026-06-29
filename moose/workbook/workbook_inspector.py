"""Lightweight workbook structure inspection for Moose."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .workbook_result import WorkbookInspectionResult


MAX_SAMPLE_ROWS = 40
MAX_SAMPLE_COLS = 20
MAX_SAMPLES_PER_SHEET = 12


SECTION_HINTS = {
    "assumptions": ("assumption", "input", "growth", "vacancy", "rent growth"),
    "capital_structure": ("sources", "uses", "equity", "loan", "debt"),
    "operating_forecast": ("revenue", "expense", "noi", "cash flow", "occupancy"),
    "debt": ("debt", "loan", "interest", "amortization", "dscr"),
    "returns": ("irr", "return", "multiple", "waterfall", "promote"),
    "exit": ("exit", "sale", "terminal", "cap rate", "disposition"),
    "rent_roll": ("rent roll", "tenant", "suite", "lease", "expiration"),
}

IMPORTANT_SHEET_HINTS = (
    "summary",
    "assumption",
    "input",
    "cash flow",
    "cf",
    "debt",
    "returns",
    "sources",
    "uses",
    "rent roll",
)

IGNORED_SHEET_HINTS = (
    "cover",
    "disclaimer",
    "readme",
    "backup",
    "check",
    "audit",
)


class WorkbookInspector:
    """Read workbook shape and text headers without extracting metrics."""

    def inspect(self, file_path: str | Path) -> WorkbookInspectionResult:
        """Return workbook structure, visible state, dimensions, and text samples."""
        path = Path(file_path)
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet_names = list(workbook.sheetnames)
            visible_sheets: list[str] = []
            hidden_sheets: list[str] = []
            dimensions: dict[str, dict[str, int]] = {}
            samples: dict[str, list[dict[str, Any]]] = {}
            all_sample_text: dict[str, list[str]] = {}
            projection_candidates: set[str] = set()

            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                if worksheet.sheet_state == "visible":
                    visible_sheets.append(sheet_name)
                else:
                    hidden_sheets.append(sheet_name)

                dimensions[sheet_name] = {
                    "max_row": int(worksheet.max_row or 0),
                    "max_column": int(worksheet.max_column or 0),
                }
                sheet_samples, sheet_projection_candidates = self._sample_sheet(worksheet)
                samples[sheet_name] = sheet_samples
                all_sample_text[sheet_name] = [
                    str(item["text"]).lower()
                    for item in sheet_samples
                    if item.get("text")
                ]
                projection_candidates.update(sheet_projection_candidates)

            likely_important_sheets = self._likely_important_sheets(sheet_names, all_sample_text)
            possible_sections = self._possible_sections(sheet_names, all_sample_text)

            return WorkbookInspectionResult(
                file_name=path.name,
                sheet_names=sheet_names,
                visible_sheets=visible_sheets,
                hidden_sheets=hidden_sheets,
                dimensions=dimensions,
                sample_non_empty_cells=samples,
                likely_important_sheets=likely_important_sheets,
                possible_model_sections=possible_sections,
                projection_period_candidates=sorted(projection_candidates),
            )
        finally:
            workbook.close()

    def _sample_sheet(self, worksheet: Any) -> tuple[list[dict[str, Any]], set[str]]:
        samples: list[dict[str, Any]] = []
        projection_candidates: set[str] = set()

        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row or 1, MAX_SAMPLE_ROWS),
            min_col=1,
            max_col=min(worksheet.max_column or 1, MAX_SAMPLE_COLS),
        ):
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if isinstance(value, int) and 1990 <= value <= 2100:
                    projection_candidates.add(str(value))
                if not isinstance(value, str):
                    continue

                text = value.strip()
                if not text or text.startswith("="):
                    continue
                if len(text) > 80:
                    text = text[:77] + "..."
                samples.append({"cell": cell.coordinate, "text": text})
                if len(samples) >= MAX_SAMPLES_PER_SHEET:
                    return samples, projection_candidates

        return samples, projection_candidates

    def _likely_important_sheets(
        self,
        sheet_names: list[str],
        all_sample_text: dict[str, list[str]],
    ) -> list[str]:
        scored: list[tuple[int, str]] = []
        for sheet_name in sheet_names:
            name = sheet_name.lower()
            if any(hint in name for hint in IGNORED_SHEET_HINTS):
                continue
            score = sum(2 for hint in IMPORTANT_SHEET_HINTS if hint in name)
            sample_blob = " ".join(all_sample_text.get(sheet_name, []))
            score += sum(1 for hints in SECTION_HINTS.values() for hint in hints if hint in sample_blob)
            if score > 0:
                scored.append((score, sheet_name))
        return [sheet for _, sheet in sorted(scored, key=lambda item: (-item[0], sheet_names.index(item[1])))[:8]]

    def _possible_sections(
        self,
        sheet_names: list[str],
        all_sample_text: dict[str, list[str]],
    ) -> list[str]:
        sections: list[str] = []
        combined = " ".join(sheet.lower() for sheet in sheet_names)
        combined += " " + " ".join(" ".join(texts) for texts in all_sample_text.values())

        for section, hints in SECTION_HINTS.items():
            if any(hint in combined for hint in hints):
                sections.append(section)
        return sections
