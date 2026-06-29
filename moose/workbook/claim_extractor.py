"""Temporary deterministic workbook claim extraction fallback.

This is not Moose's primary claim extraction architecture. Primary extraction should use
`Workbook Evidence Pack -> WorkbookClaimDiscoveryAgent -> Grounding Validator`.

This fallback exists only while the LLM-backed discovery agent is stubbed or unavailable.
It uses a small metric label map as a smoke-test scaffold and must not grow into a broad
deterministic extractor.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from moose.intake import IntakeResult

from .workbook_result import (
    ModelBriefResult,
    WorkbookInspectionResult,
    WorkbookMentalModelResult,
    WorkbookOrientationResult,
)


MAX_CLAIMS = 20
SCAN_ROWS = 160
SCAN_COLS = 80

COLLIE_BASELINE_METRICS = {
    "purchase_price",
    "total_cost",
    "debt",
    "equity",
    "ltv",
    "interest_rate",
    "noi",
    "levered_irr",
    "unlevered_irr",
    "equity_multiple",
    "unlevered_equity_multiple",
    "exit_cap",
    "sale_price",
    "hold_period",
}

COLLIE_TO_MOOSE_METRIC = {
    "purchase_price": ("purchase_price", "currency", "investment_basis"),
    "total_cost": ("total_project_cost", "currency", "investment_basis"),
    "debt": ("debt_amount", "currency", "capital_structure"),
    "equity": ("equity_required", "currency", "capital_structure"),
    "ltv": ("loan_to_value", "percent", "debt"),
    "interest_rate": ("interest_rate", "percent", "debt"),
    "noi": ("stabilized_noi", "currency", "operating_performance"),
    "levered_irr": ("levered_irr", "percent", "returns"),
    "unlevered_irr": ("unlevered_irr", "percent", "returns"),
    "equity_multiple": ("equity_multiple", "multiple", "returns"),
    "unlevered_equity_multiple": ("unlevered_equity_multiple", "multiple", "returns"),
    "exit_cap": ("exit_cap_rate", "percent", "exit"),
    "sale_price": ("sale_value", "currency", "exit"),
    "hold_period": ("hold_period", "years", "exit"),
}

METRIC_DEFINITIONS = {
    "purchase_price": {
        "family": "investment_basis",
        "claim_type": "financial_metric",
        "unit": "currency",
        "labels": ("purchase price", "acquisition price"),
    },
    "total_project_cost": {
        "family": "investment_basis",
        "claim_type": "financial_metric",
        "unit": "currency",
        "labels": ("total project cost", "total cost", "total basis"),
    },
    "debt_amount": {
        "family": "capital_structure",
        "claim_type": "financial_metric",
        "unit": "currency",
        "labels": ("debt amount", "loan amount", "senior loan"),
    },
    "equity_required": {
        "family": "capital_structure",
        "claim_type": "financial_metric",
        "unit": "currency",
        "labels": ("equity required", "total equity", "equity invested"),
    },
    "loan_to_value": {
        "family": "debt",
        "claim_type": "ratio",
        "unit": "percent",
        "labels": ("ltv", "loan to value", "loan-to-value"),
    },
    "interest_rate": {
        "family": "debt",
        "claim_type": "ratio",
        "unit": "percent",
        "labels": ("interest rate", "coupon", "all-in rate"),
    },
    "levered_irr": {
        "family": "returns",
        "claim_type": "return_metric",
        "unit": "percent",
        "labels": ("levered irr", "leveraged irr", "project irr"),
    },
    "unlevered_irr": {
        "family": "returns",
        "claim_type": "return_metric",
        "unit": "percent",
        "labels": ("unlevered irr", "unleveraged irr"),
    },
    "equity_multiple": {
        "family": "returns",
        "claim_type": "return_metric",
        "unit": "multiple",
        "labels": ("equity multiple", "moic"),
    },
    "stabilized_noi": {
        "family": "operating_performance",
        "claim_type": "financial_metric",
        "unit": "currency",
        "labels": ("stabilized noi", "noi", "net operating income"),
    },
    "exit_cap_rate": {
        "family": "exit",
        "claim_type": "ratio",
        "unit": "percent",
        "labels": ("exit cap", "terminal cap"),
    },
    "sale_value": {
        "family": "exit",
        "claim_type": "financial_metric",
        "unit": "currency",
        "labels": ("sale value", "exit value", "terminal value"),
    },
    "hold_period": {
        "family": "exit",
        "claim_type": "duration",
        "unit": "years",
        "labels": ("hold period", "investment period"),
    },
}


class FallbackWorkbookClaimExtractor:
    """Temporary deterministic scaffold used only when GPT discovery is unavailable."""

    def extract(
        self,
        file_path: str | Path,
        intake_result: IntakeResult,
        inspection: WorkbookInspectionResult,
        orientation: WorkbookOrientationResult,
        model_brief: ModelBriefResult,
        mental_model: WorkbookMentalModelResult,
    ) -> list[dict[str, Any]]:
        """Return unverified fallback claims guided by the workbook mental model."""
        del intake_result, inspection, orientation, model_brief
        path = Path(file_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            claims = self._claims_from_collie_v2_baseline(path, mental_model)
            if claims:
                return claims[:MAX_CLAIMS]

            claims: list[dict[str, Any]] = []
            seen_metrics: set[str] = set()
            sheet_order = self._sheet_order(workbook.sheetnames, mental_model)
            wanted_families = set(mental_model.expected_metric_families)

            for sheet_name in sheet_order:
                worksheet = workbook[sheet_name]
                for claim in self._claims_from_sheet(path.name, worksheet, wanted_families, mental_model):
                    metric = claim["metric_or_subject"]
                    if metric in seen_metrics:
                        continue
                    claims.append(claim)
                    seen_metrics.add(metric)
                    if len(claims) >= MAX_CLAIMS:
                        return claims
            return claims
        finally:
            workbook.close()

    def _sheet_order(
        self,
        workbook_sheet_names: list[str],
        mental_model: WorkbookMentalModelResult,
    ) -> list[str]:
        preferred = mental_model.important_sheets + [
            sheet
            for sources in mental_model.likely_authoritative_sources.values()
            for sheet in sources
        ]
        ordered: list[str] = []
        for preferred_name in preferred:
            match = self._match_sheet(preferred_name, workbook_sheet_names)
            if match and match not in ordered:
                ordered.append(match)
        ordered.extend(sheet for sheet in workbook_sheet_names if sheet not in ordered)
        return ordered

    def _claims_from_sheet(
        self,
        source_document: str,
        worksheet: Any,
        wanted_families: set[str],
        mental_model: WorkbookMentalModelResult,
    ) -> list[dict[str, Any]]:
        claims: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row or 1, SCAN_ROWS),
            min_col=1,
            max_col=min(worksheet.max_column or 1, SCAN_COLS),
        ):
            for cell in row:
                label = cell.value
                if not isinstance(label, str) or not label.strip():
                    continue
                metric = self._metric_for_label(label, wanted_families)
                if not metric:
                    continue
                value_cell = self._nearby_value_cell(worksheet, cell.row, cell.column)
                if not value_cell:
                    continue
                definition = METRIC_DEFINITIONS[metric]
                claims.append(
                    {
                        "claim_id": f"claim:{metric}:{uuid4().hex[:8]}",
                        "claim_type": definition["claim_type"],
                        "metric_or_subject": metric,
                        "value": value_cell.value,
                        "unit": definition["unit"],
                        "period": None,
                        "source_document": source_document,
                        "source_location": {
                            "sheet": worksheet.title,
                            "cell": value_cell.coordinate,
                            "nearby_label": label.strip(),
                            "table_or_section": self._section_for_metric(metric, mental_model),
                        },
                        "evidence": [
                            {
                                "quote": label.strip(),
                                "cell": value_cell.coordinate,
                                "sheet": worksheet.title,
                                "nearby_label_cell": cell.coordinate,
                            }
                        ],
                        "confidence": self._confidence(worksheet.title, metric, mental_model),
                        "reasoning": (
                            f"Found a value adjacent to label '{label.strip()}' on an expected "
                            f"source sheet for {metric}."
                        ),
                        "extraction_method": "mental_model_guided_workbook_claim_extraction_v0",
                    }
                )
        return claims

    def _metric_for_label(self, label: str, wanted_families: set[str]) -> str | None:
        normalized = label.strip().lower().replace("_", " ").replace("-", " ")
        if "hedge maturity" in normalized or "extension option" in normalized:
            return None
        for metric, definition in METRIC_DEFINITIONS.items():
            if definition["family"] not in wanted_families:
                continue
            if metric == "levered_irr" and "unlevered irr" in normalized:
                continue
            if metric == "interest_rate" and any(blocked in normalized for blocked in ("hedge", "maturity")):
                continue
            if any(candidate in normalized for candidate in definition["labels"]):
                return metric
        return None

    def _claims_from_collie_v2_baseline(
        self,
        path: Path,
        mental_model: WorkbookMentalModelResult,
    ) -> list[dict[str, Any]]:
        """Bridge legacy Collie v2 truth into Moose as a temporary fallback baseline.

        TODO(Day 7+): Remove this bridge once GPT claim discovery can recover the
        same baseline facts from a bounded evidence pack. This code is deliberately
        fallback-only and should not become Moose's primary extraction architecture.
        """
        try:
            from deal_truth import build_deal_truth
        except Exception:
            return []

        try:
            truth = build_deal_truth(path)
        except Exception:
            return []

        canonical = truth.get("canonical") or {}
        if not isinstance(canonical, dict):
            return []

        claims: list[dict[str, Any]] = []
        wanted_families = set(mental_model.expected_metric_families)
        for collie_metric in COLLIE_BASELINE_METRICS:
            fact = canonical.get(collie_metric)
            if not isinstance(fact, dict) or fact.get("value") is None:
                continue
            mapped = COLLIE_TO_MOOSE_METRIC.get(collie_metric)
            if not mapped:
                continue
            metric, unit, family = mapped
            if wanted_families and family not in wanted_families:
                continue
            source_text = str(fact.get("source") or "")
            source_location = self._source_location_from_collie_source(source_text)
            if not source_location:
                source_location = {"sheet": None, "cell": None, "legacy_source": source_text}
            source_location["legacy_source"] = source_text
            if source_location.get("row") and not source_location.get("cell"):
                source_location["derivation"] = source_text.split("(", 1)[-1].rstrip(")") if "(" in source_text else "legacy row-derived metric"

            claims.append({
                "claim_id": f"claim:{metric}:{uuid4().hex[:8]}",
                "claim_type": METRIC_DEFINITIONS.get(metric, {}).get("claim_type", "financial_metric"),
                "metric_or_subject": metric,
                "value": fact.get("value"),
                "unit": unit,
                "period": fact.get("period"),
                "source_document": path.name,
                "source_location": source_location,
                "evidence": [{
                    "quote": source_text or f"Collie v2 canonical {collie_metric}",
                    "sheet": source_location.get("sheet"),
                    "cell": source_location.get("cell"),
                    "row": source_location.get("row"),
                }],
                "confidence": self._confidence(str(source_location.get("sheet") or ""), metric, mental_model),
                "reasoning": (
                    "Recovered from Collie v2 canonical truth as a temporary Moose "
                    "baseline fallback while GPT claim discovery is stubbed."
                ),
                "extraction_method": "collie_v2_baseline_fallback",
                "fallback_only": True,
            })
        return claims

    def _source_location_from_collie_source(self, source_text: str) -> dict[str, Any] | None:
        if "!" not in source_text:
            return None
        sheet_name, ref_and_note = source_text.split("!", 1)
        ref = ref_and_note.split(" ", 1)[0]
        ref = ref.strip()
        if not sheet_name or not ref:
            return None
        if ref.lower().startswith("row"):
            row_text = "".join(ch for ch in ref if ch.isdigit())
            if not row_text:
                return None
            return {"sheet": sheet_name, "row": int(row_text)}
        return {"sheet": sheet_name, "cell": ref}

    def _nearby_value_cell(self, worksheet: Any, row: int, col: int) -> Any | None:
        candidates = [
            (row, col + 1),
            (row, col + 2),
            (row + 1, col),
            (row + 1, col + 1),
        ]
        for candidate_row, candidate_col in candidates:
            value = worksheet.cell(candidate_row, candidate_col).value
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return worksheet.cell(candidate_row, candidate_col)
        return None

    def _section_for_metric(self, metric: str, mental_model: WorkbookMentalModelResult) -> str | None:
        family = METRIC_DEFINITIONS[metric]["family"]
        sections_by_family = {
            "investment_basis": {"assumptions", "capital_structure"},
            "capital_structure": {"capital_structure", "debt"},
            "debt": {"debt"},
            "returns": {"returns"},
            "operating_performance": {"operating_forecast", "rent_roll"},
            "exit": {"exit", "assumptions"},
        }
        for section in mental_model.expected_sections:
            if section in sections_by_family.get(family, set()):
                return section
        return None

    def _confidence(
        self,
        sheet_name: str,
        metric: str,
        mental_model: WorkbookMentalModelResult,
    ) -> float:
        confidence = 0.55
        authoritative_sources = mental_model.likely_authoritative_sources.get(metric, [])
        if any(self._sheet_name_matches(sheet_name, source) for source in authoritative_sources):
            confidence += 0.18
        if sheet_name in mental_model.important_sheets:
            confidence += 0.12
        confidence += min(mental_model.confidence * 0.1, 0.1)
        return round(min(confidence, 0.9), 2)

    def _match_sheet(self, preferred_name: str, sheet_names: list[str]) -> str | None:
        for sheet_name in sheet_names:
            if self._sheet_name_matches(sheet_name, preferred_name):
                return sheet_name
        return None

    def _sheet_name_matches(self, sheet_name: str, preferred_name: str) -> bool:
        return preferred_name.lower() in sheet_name.lower() or sheet_name.lower() in preferred_name.lower()


# Backward-compatible alias for early Day 4 code. New code should import
# FallbackWorkbookClaimExtractor or WorkbookClaimDiscoveryAgent explicitly.
WorkbookClaimExtractor = FallbackWorkbookClaimExtractor
