"""Build bounded workbook evidence for GPT claim discovery."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .workbook_result import WorkbookEvidencePackResult, WorkbookMentalModelResult


MAX_ROWS = 160
MAX_COLS = 80
MAX_SAMPLED_CELLS_PER_SHEET = 700
MAX_NEIGHBORHOODS = 300

BASELINE_SHEET_PRIORITY = (
    "General Information",
    "One Pager",
    "Key UW Metrics",
    "Inputs",
    "Debt Information",
    "Cash Flows",
    "Model",
    "NewKeys",
    "ClosingCosts",
)


class WorkbookEvidencePackBuilder:
    """Collect bounded workbook context without extracting claims."""

    def build(
        self,
        file_path: str | Path,
        mental_model: WorkbookMentalModelResult,
    ) -> WorkbookEvidencePackResult:
        """Return sheet samples, section/header blocks, and candidate neighborhoods."""
        path = Path(file_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_names = self._ordered_sheets(workbook.sheetnames, mental_model)
            sampled_cells: dict[str, list[dict[str, Any]]] = {}
            section_header_blocks: dict[str, list[dict[str, Any]]] = {}
            candidate_neighborhoods: list[dict[str, Any]] = []

            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                sampled_cells[sheet_name] = self._sample_cells(worksheet)
                section_header_blocks[sheet_name] = self._section_blocks(worksheet)
                if len(candidate_neighborhoods) < MAX_NEIGHBORHOODS:
                    remaining = MAX_NEIGHBORHOODS - len(candidate_neighborhoods)
                    candidate_neighborhoods.extend(
                        self._candidate_neighborhoods(worksheet, limit=remaining)
                    )

            return WorkbookEvidencePackResult(
                important_sheet_names=sheet_names,
                sampled_cells=sampled_cells,
                section_header_blocks=section_header_blocks,
                candidate_neighborhoods=candidate_neighborhoods,
                caveats=[
                    "Evidence pack is bounded; it does not include the full workbook.",
                    "Evidence pack is context for claim discovery, not verified fact.",
                ],
            )
        finally:
            workbook.close()

    def _ordered_sheets(
        self,
        workbook_sheet_names: list[str],
        mental_model: WorkbookMentalModelResult,
    ) -> list[str]:
        ordered = [
            sheet for sheet in mental_model.important_sheets
            if sheet in workbook_sheet_names
        ]
        for priority in BASELINE_SHEET_PRIORITY:
            match = self._match_sheet(priority, workbook_sheet_names)
            if match and match not in ordered:
                ordered.append(match)
        ordered.extend(sheet for sheet in workbook_sheet_names if sheet not in ordered)
        return ordered[:18]

    def _match_sheet(self, preferred_name: str, sheet_names: list[str]) -> str | None:
        for sheet_name in sheet_names:
            if preferred_name.lower() == sheet_name.lower():
                return sheet_name
        for sheet_name in sheet_names:
            if preferred_name.lower() in sheet_name.lower() or sheet_name.lower() in preferred_name.lower():
                return sheet_name
        return None

    def _sample_cells(self, worksheet: Any) -> list[dict[str, Any]]:
        samples: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row or 1, MAX_ROWS),
            min_col=1,
            max_col=min(worksheet.max_column or 1, MAX_COLS),
        ):
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                samples.append({
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "column": cell.column,
                    "value": value,
                })
                if len(samples) >= MAX_SAMPLED_CELLS_PER_SHEET:
                    return samples
        return samples

    def _section_blocks(self, worksheet: Any) -> list[dict[str, Any]]:
        blocks: list[dict[str, Any]] = []
        for row_idx, row in enumerate(worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row or 1, MAX_ROWS),
            min_col=1,
            max_col=min(worksheet.max_column or 1, min(worksheet.max_column or 1, 8)),
        ), start=1):
            texts = []
            for cell in row:
                value = getattr(cell, "value", None)
                if isinstance(value, str) and value.strip():
                    texts.append(value.strip())
            if not texts:
                continue
            blocks.append({
                "sheet": worksheet.title,
                "row": row_idx,
                "text": " | ".join(texts),
            })
        return blocks[:30]

    def _candidate_neighborhoods(self, worksheet: Any, limit: int = MAX_NEIGHBORHOODS) -> list[dict[str, Any]]:
        neighborhoods: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row or 1, MAX_ROWS),
            min_col=1,
            max_col=min(worksheet.max_column or 1, MAX_COLS),
        ):
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.strip():
                    continue
                numeric_neighbors = self._numeric_neighbors(worksheet, cell.row, cell.column)
                if not numeric_neighbors:
                    continue
                neighborhoods.append({
                    "sheet": worksheet.title,
                    "label_cell": cell.coordinate,
                    "label": cell.value.strip(),
                    "nearby_values": numeric_neighbors,
                    "row_context": self._row_context(worksheet, cell.row),
                })
                if len(neighborhoods) >= limit:
                    return neighborhoods
        return neighborhoods

    def _numeric_neighbors(self, worksheet: Any, row: int, col: int) -> list[dict[str, Any]]:
        neighbors: list[dict[str, Any]] = []
        for candidate_row, candidate_col in (
            (row, col + 1),
            (row, col + 2),
            (row + 1, col),
            (row + 1, col + 1),
        ):
            value = worksheet.cell(candidate_row, candidate_col).value
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                neighbors.append({
                    "cell": worksheet.cell(candidate_row, candidate_col).coordinate,
                    "value": value,
                })
        return neighbors

    def _row_context(self, worksheet: Any, row: int) -> list[dict[str, Any]]:
        context: list[dict[str, Any]] = []
        for cell in worksheet[row][:MAX_COLS]:
            if cell.value is not None:
                context.append({"cell": cell.coordinate, "value": cell.value})
        return context
