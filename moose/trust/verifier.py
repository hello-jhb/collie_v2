"""Code-based verification for Moose claims."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .authority import AuthorityResolver
from .verification_result import VerifiedFact, VerificationRunResult


class TrustVerifier:
    """Verify claims with code before they become facts."""

    check_names = (
        "source_location_present",
        "sheet_exists",
        "cell_exists",
        "value_matches_cell",
        "evidence_matches_neighborhood",
        "unit_scale_plausible",
        "authority",
    )

    def __init__(self, authority_resolver: AuthorityResolver | None = None) -> None:
        self.authority_resolver = authority_resolver or AuthorityResolver()

    def verify_claims(
        self,
        file_path: str | Path,
        claims: list[dict[str, Any]],
        mental_model: dict[str, Any] | None = None,
    ) -> VerificationRunResult:
        """Verify a batch of claims against workbook cells and the mental model."""
        path = Path(file_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            verified_facts = [
                self.verify_claim(workbook, claim, mental_model)
                for claim in claims
            ]
        finally:
            workbook.close()

        fact_dicts = [fact.as_dict() for fact in verified_facts]
        summary = {status: 0 for status in [
            "verified",
            "verified_with_caveat",
            "needs_review",
            "contradicted",
            "rejected",
        ]}
        for fact in fact_dicts:
            summary[fact["verification_status"]] += 1
        summary["claims_total"] = len(claims)

        caveats = [
            caveat
            for fact in fact_dicts
            for caveat in fact.get("caveats", [])
        ]
        rejected_claims = [
            {"claim_id": fact["claim_id"], "status": fact["verification_status"], "notes": fact.get("notes")}
            for fact in fact_dicts
            if fact["verification_status"] in {"contradicted", "rejected"}
        ]

        return VerificationRunResult(
            verified_facts=fact_dicts,
            summary=summary,
            caveats=caveats,
            rejected_claims=rejected_claims,
            reconciliation_notes=[],
        )

    def verify_claim(
        self,
        workbook: Any,
        claim: dict[str, Any],
        mental_model: dict[str, Any] | None = None,
    ) -> VerifiedFact:
        """Verify one claim against workbook source evidence."""
        checks: list[dict[str, Any]] = []
        caveats: list[str] = []
        notes: list[str] = []
        source_location = claim.get("source_location")

        if not isinstance(source_location, dict):
            checks.extend([
                self._check("source_location_present", "failed", "Claim source_location is missing or not structured."),
                self._not_run("sheet_exists"),
                self._not_run("cell_exists"),
                self._not_run("value_matches_cell"),
                self._not_run("evidence_matches_neighborhood"),
                self._check_unit_scale(claim),
                self._not_run("authority"),
            ])
            return self._fact(claim, "needs_review", checks, caveats, "Source location is required.")

        checks.append(self._check("source_location_present", "passed", "Claim includes structured source location."))
        sheet_name = source_location.get("sheet")
        cell_ref = source_location.get("cell")

        if not sheet_name or sheet_name not in workbook.sheetnames:
            checks.extend([
                self._check("sheet_exists", "failed", f"Sheet does not exist: {sheet_name}."),
                self._not_run("cell_exists"),
                self._not_run("value_matches_cell"),
                self._not_run("evidence_matches_neighborhood"),
                self._check_unit_scale(claim),
                self._not_run("authority"),
            ])
            return self._fact(claim, "rejected", checks, caveats, "Referenced sheet is missing.")

        checks.append(self._check("sheet_exists", "passed", f"Sheet exists: {sheet_name}."))
        worksheet = workbook[sheet_name]
        row_ref = source_location.get("row")
        if not cell_ref and row_ref:
            return self._verify_row_derived_claim(worksheet, claim, mental_model, checks, caveats, row_ref)

        try:
            source_cell = worksheet[cell_ref]
        except Exception:
            checks.extend([
                self._check("cell_exists", "failed", f"Cell does not exist or is invalid: {cell_ref}."),
                self._not_run("value_matches_cell"),
                self._not_run("evidence_matches_neighborhood"),
                self._check_unit_scale(claim),
                self._run_authority_check(claim, mental_model, caveats),
            ])
            return self._fact(claim, "rejected", checks, caveats, "Referenced cell is missing.")

        checks.append(self._check("cell_exists", "passed", f"Cell exists: {sheet_name}!{cell_ref}."))

        if self._values_match(source_cell.value, claim.get("value")):
            checks.append(self._check("value_matches_cell", "passed", "Claim value matches cited workbook cell."))
        else:
            checks.extend([
                self._check(
                    "value_matches_cell",
                    "failed",
                    f"Claim value {claim.get('value')} does not match cell value {source_cell.value}.",
                ),
                self._check_evidence_neighborhood(worksheet, source_cell.row, source_cell.column, claim),
                self._check_unit_scale(claim),
                self._run_authority_check(claim, mental_model, caveats),
            ])
            return self._fact(claim, "contradicted", checks, caveats, "Claim value contradicts cited cell.")

        checks.append(self._check_evidence_neighborhood(worksheet, source_cell.row, source_cell.column, claim))
        checks.append(self._check_unit_scale(claim))
        checks.append(self._run_authority_check(claim, mental_model, caveats))

        failed = [check for check in checks if check["status"] == "failed"]
        needs_review = [check for check in checks if check["status"] == "needs_review"]
        if failed:
            status = "needs_review"
            notes.append("One or more non-core checks failed.")
        elif needs_review:
            status = "verified_with_caveat"
            caveats.extend(check["details"] for check in needs_review)
        else:
            status = "verified"

        return self._fact(claim, status, checks, list(dict.fromkeys(caveats)), " ".join(notes) or None)

    def _fact(
        self,
        claim: dict[str, Any],
        status: str,
        checks: list[dict[str, Any]],
        caveats: list[str],
        notes: str | None,
    ) -> VerifiedFact:
        source = self._source_string(claim.get("source_location"))
        return VerifiedFact(
            fact_id=f"fact:{claim.get('claim_id', 'unknown')}",
            claim_id=str(claim.get("claim_id", "unknown")),
            metric_or_subject=str(claim.get("metric_or_subject", "unknown")),
            verified_value=claim.get("value") if status not in {"rejected", "contradicted"} else None,
            unit=claim.get("unit"),
            source=source,
            extraction_method=claim.get("extraction_method"),
            fact_origin="fallback" if claim.get("fallback_only") else "gpt_native",
            verification_status=status,
            checks=checks,
            caveats=caveats,
            notes=notes,
        )

    def _check(self, name: str, status: str, details: str) -> dict[str, str]:
        return {"name": name, "status": status, "details": details}

    def _not_run(self, name: str) -> dict[str, str]:
        return self._check(name, "not_run", "Skipped because a prerequisite check failed.")

    def _values_match(self, workbook_value: Any, claim_value: Any) -> bool:
        if isinstance(workbook_value, (int, float)) and isinstance(claim_value, (int, float)):
            return abs(float(workbook_value) - float(claim_value)) <= 0.000001
        return workbook_value == claim_value

    def _verify_row_derived_claim(
        self,
        worksheet: Any,
        claim: dict[str, Any],
        mental_model: dict[str, Any] | None,
        checks: list[dict[str, Any]],
        caveats: list[str],
        row_ref: Any,
    ) -> VerifiedFact:
        if not isinstance(row_ref, int) or row_ref < 1 or row_ref > (worksheet.max_row or 0):
            checks.extend([
                self._check("cell_exists", "failed", f"Referenced row is invalid: row{row_ref}."),
                self._not_run("value_matches_cell"),
                self._not_run("evidence_matches_neighborhood"),
                self._check_unit_scale(claim),
                self._run_authority_check(claim, mental_model, caveats),
            ])
            return self._fact(claim, "rejected", checks, caveats, "Referenced row is missing.")

        checks.extend([
            self._check("cell_exists", "passed", f"Row-derived source exists: {worksheet.title}!row{row_ref}."),
            self._check(
                "value_matches_cell",
                "needs_review",
                "Claim is derived from a row/cash-flow calculation rather than a single cited cell.",
            ),
            self._check("evidence_matches_neighborhood", "passed", "Legacy source row is present in the workbook."),
            self._check_unit_scale(claim),
            self._run_authority_check(claim, mental_model, caveats),
        ])
        caveats.append("Verified with caveat because this fallback claim is row-derived from Collie v2 baseline logic.")
        return self._fact(
            claim,
            "verified_with_caveat",
            checks,
            list(dict.fromkeys(caveats)),
            "Row-derived fallback fact; replace with GPT-discovered, code-verified source evidence in the next Moose iteration.",
        )

    def _check_evidence_neighborhood(
        self,
        worksheet: Any,
        row: int,
        col: int,
        claim: dict[str, Any],
    ) -> dict[str, str]:
        source_location = claim.get("source_location") or {}
        nearby_label = source_location.get("nearby_label")
        evidence_quotes = [
            item.get("quote")
            for item in claim.get("evidence", [])
            if isinstance(item, dict) and item.get("quote")
        ]
        labels = [nearby_label] if nearby_label else []
        labels.extend(evidence_quotes)
        if not labels:
            return self._check("evidence_matches_neighborhood", "needs_review", "No label or evidence quote provided.")

        for label in labels:
            if self._label_near_cell(worksheet, row, col, str(label)):
                return self._check("evidence_matches_neighborhood", "passed", "Evidence label appears near cited cell.")
        return self._check("evidence_matches_neighborhood", "needs_review", "Evidence label was not found near cited cell.")

    def _label_near_cell(self, worksheet: Any, row: int, col: int, label: str) -> bool:
        normalized = label.strip().lower()
        for candidate_row in range(max(1, row - 2), row + 3):
            for candidate_col in range(max(1, col - 3), col + 2):
                value = worksheet.cell(candidate_row, candidate_col).value
                if isinstance(value, str) and value.strip().lower() == normalized:
                    return True
        return False

    def _check_unit_scale(self, claim: dict[str, Any]) -> dict[str, str]:
        metric = claim.get("metric_or_subject")
        value = claim.get("value")
        unit = claim.get("unit")
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            return self._check("unit_scale_plausible", "needs_review", "Claim value is not numeric.")
        numeric_value = float(value)

        if unit == "percent" and -1.0 <= numeric_value <= 100.0:
            return self._check("unit_scale_plausible", "passed", "Percent-like value is within plausible range.")
        if unit == "multiple" and 0.0 <= numeric_value <= 50.0:
            return self._check("unit_scale_plausible", "passed", "Multiple-like value is within plausible range.")
        if unit == "years" and 0.0 <= numeric_value <= 100.0:
            return self._check("unit_scale_plausible", "passed", "Duration value is within plausible range.")
        if unit == "currency" and abs(numeric_value) >= 1.0:
            return self._check("unit_scale_plausible", "passed", "Currency-like value is non-trivial.")
        if metric in {"loan_to_value", "interest_rate", "levered_irr", "unlevered_irr", "exit_cap_rate"}:
            return self._check("unit_scale_plausible", "needs_review", "Ratio metric has unexpected unit or scale.")
        return self._check("unit_scale_plausible", "needs_review", f"Unit or scale needs review: {unit}.")

    def _run_authority_check(
        self,
        claim: dict[str, Any],
        mental_model: dict[str, Any] | None,
        caveats: list[str],
    ) -> dict[str, str]:
        result = self.authority_resolver.resolve(claim, mental_model)
        caveats.extend(result.get("caveats", []))
        return self._check("authority", result["status"], result["details"])

    def _source_string(self, source_location: Any) -> str:
        if isinstance(source_location, dict):
            sheet = source_location.get("sheet", "unknown")
            cell = source_location.get("cell", "unknown")
            if source_location.get("row") and not source_location.get("cell"):
                return f"{sheet}!row{source_location.get('row')}"
            return f"{sheet}!{cell}"
        return "unknown"
