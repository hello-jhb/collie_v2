"""Lightweight file identification for Moose intake."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from .context_resolver import KNOWLEDGE_DIR, load_simple_yaml
from .intake_result import DocumentIdentity


WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TEXT_EXTENSIONS = {".csv", ".txt", ".md"}
LOW_CONFIDENCE_THRESHOLD = 0.45


class FileIdentifier:
    """Classify files using cheap signals and knowledge-defined document types."""

    def __init__(self, knowledge_dir: str | Path = KNOWLEDGE_DIR) -> None:
        self.knowledge_dir = Path(knowledge_dir)
        self.document_types = load_simple_yaml(self.knowledge_dir / "document_types.yaml").get(
            "document_types", {}
        )
        self.supported_document_types = set(self.document_types) | {"unknown"}

    def identify(self, file_path: str | Path) -> DocumentIdentity:
        """Return a reviewable document identity without extracting investment metrics."""
        path = Path(file_path)
        suffix = path.suffix.lower()
        filename = path.stem.lower().replace("-", " ").replace("_", " ")

        evidence = [f"File extension is {suffix or 'missing'}."]
        sheet_names = self._read_sheet_names(path)
        if sheet_names:
            evidence.append(f"Workbook contains sheets named {', '.join(sheet_names[:8])}.")

        snippet = self._read_text_snippet(path)
        if snippet:
            evidence.append(f"Readable text/header snippet includes: {snippet[:180]}")

        candidates = self._score_candidates(filename, suffix, sheet_names, snippet)
        document_type, confidence, reasoning = self._select_candidate(candidates)

        if confidence < LOW_CONFIDENCE_THRESHOLD:
            document_type = "unknown"
            reasoning = "Signals were too weak or ambiguous for a confident v0 classification."

        return DocumentIdentity(
            document_type=document_type,
            confidence=round(confidence, 2),
            evidence=evidence,
            reasoning=reasoning,
            human_review_required=document_type == "unknown" or confidence < LOW_CONFIDENCE_THRESHOLD,
        )

    def _score_candidates(
        self,
        filename: str,
        suffix: str,
        sheet_names: list[str],
        snippet: str,
    ) -> dict[str, float]:
        text = " ".join([filename, " ".join(sheet_names).lower(), snippet.lower()])
        scores = {document_type: 0.0 for document_type in self.document_types}

        for document_type, context in self.document_types.items():
            formats = {f".{item}" for item in context.get("likely_formats", [])}
            if suffix in formats:
                scores[document_type] += 0.18

        # TODO(Day 2): Keep hints broad. Add or remove hints based on real intake examples.
        hints = {
            "financial_model": ["model", "underwriting", "cash flow", "debt", "returns", "assumptions"],
            "operating_statement": ["operating statement", "income statement", "actuals", "noi"],
            "rent_roll": ["rent roll", "tenant", "suite", "expiration", "lease expiry"],
            "lease": ["lease agreement", "landlord", "tenant", "premises"],
            "loan_agreement": ["loan agreement", "borrower", "lender", "maturity", "covenant"],
            "appraisal": ["appraisal", "appraised value", "valuation", "cap rate"],
            "offering_memorandum": ["offering memorandum", " om ", "offering", "investment opportunity"],
            "investment_memo": ["investment memo", "investment committee", "recommendation"],
            "business_plan": ["business plan", "strategy", "asset plan"],
            "budget_workbook": ["budget", "budget workbook", "budgeted"],
            "variance_report": ["variance", "budget vs actual", "actual to budget"],
            "property_management_report": ["property management", "management report", "work order"],
            "leasing_report": ["leasing", "prospect", "pipeline", "lease up"],
            "capital_project_tracker": ["capital project", "capex", "project tracker", "change order"],
            "portfolio_report": ["portfolio", "asset summary", "portfolio report"],
            "investor_report": ["investor report", "quarterly report", "investor"],
            "fund_model": ["fund model", "waterfall", "capital account", "promote"],
            "nav_package": ["nav", "net asset value"],
            "rfp": ["rfp", "request for proposal"],
            "vendor_proposal": ["proposal", "vendor", "fee proposal"],
            "service_agreement": ["service agreement", "scope of work", "engagement letter"],
        }

        for document_type, phrases in hints.items():
            if document_type not in scores:
                continue
            for phrase in phrases:
                if phrase in text:
                    scores[document_type] += 0.28

        if suffix in WORKBOOK_EXTENSIONS and sheet_names:
            workbook_sheet_text = " ".join(sheet_names).lower()
            if any(term in workbook_sheet_text for term in ["summary", "assumptions", "cash flow", "debt"]):
                scores["financial_model"] += 0.24
            if any(term in workbook_sheet_text for term in ["budget", "budget input", "budget detail"]):
                scores["budget_workbook"] += 0.24
            if any(term in workbook_sheet_text for term in ["rent roll", "tenant"]):
                scores["rent_roll"] += 0.24

        return scores

    def _select_candidate(self, candidates: dict[str, float]) -> tuple[str, float, str]:
        if not candidates:
            return "unknown", 0.0, "No supported document types were available."

        ranked = sorted(candidates.items(), key=lambda item: item[1], reverse=True)
        best_type, best_score = ranked[0]
        second_score = ranked[1][1] if len(ranked) > 1 else 0.0
        confidence = min(best_score, 0.88)

        if best_score <= 0:
            return "unknown", 0.0, "No lightweight file signals matched known document types."
        if best_score - second_score < 0.08:
            return "unknown", max(0.2, confidence), "Top document type signals were ambiguous."

        return (
            best_type,
            max(0.1, confidence),
            f"Lightweight file signals most closely match {best_type}.",
        )

    def _read_sheet_names(self, path: Path) -> list[str]:
        if path.suffix.lower() not in WORKBOOK_EXTENSIONS or not path.exists():
            return []
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                return list(workbook.sheetnames)
            finally:
                workbook.close()
        except Exception:
            return []

    def _read_text_snippet(self, path: Path) -> str:
        if path.suffix.lower() not in TEXT_EXTENSIONS or not path.exists():
            return ""
        try:
            return path.read_text(encoding="utf-8", errors="ignore").replace("\n", " ")[:500]
        except OSError:
            return ""
