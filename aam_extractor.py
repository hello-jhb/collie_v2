"""
aam_extractor.py — Stage-1 focused extraction of the Audit Appendix Metrics.

This is the redesigned first pass (2026-06-10). It replaces the broad, automatic
GPT sweep (5 section reads + insight pass + comprehension review over all 109
catalog metrics) with a tight, AAM-scoped pass:

    Workbook Mapper (classify_sheets, GPT call #1)
        |  role map
    Deterministic resolve (proximity scan + resolve_metric, scoped to ~20 AAM ids)
        |  fills what it can confidently, with schema + source-hierarchy validation
    Focused GPT gap-fill (ONE batched call, GPT call #2 -- only the gaps)
        |
    AAM records  ->  Audit Appendix  ->  human verification gate

The deterministic half runs with NO API key. Both GPT calls are skipped silently
when llm_available() is False, so the appendix still renders (with blanks the
user can later fill on demand).

This is NOT a new extractor. It is a scope + orchestration layer that reuses the
existing engine wholesale: flexible_extractor primitives, metric_resolver
ranking, and sheet_classifier -- restricted to the AAM.
"""
from __future__ import annotations

import json
import logging
import sys
from pathlib import Path
from typing import Any

from aam import aam_metrics, AAM_HARDCODED_INPUTS
from metric_catalog import load_metric_catalog
from flexible_extractor import (
    scan_workbook_for_candidates,
    extract_raw_labeled_pairs,
    sheet_priority_tier,
)
from metric_resolver import resolve_metric
from scenarios._llm import client, MODEL_FAST, llm_available

log = logging.getLogger("fb.aam_extractor")
if not log.handlers:
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("[fb.aam] %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

AAM_EXTRACTOR_VERSION = "2026-06-12.2"

# Statuses the GPT gap-fill may touch. STRICTLY blanks and flagged-wrong rows.
# candidate_pool is deliberately NOT here: it means the deterministic pass
# already ranked a value that passed validation — letting GPT "re-fill" those
# rows overwrote a correct $192M purchase price with a per-asset $88M slice on
# BAC. GPT fills holes; it does not second-guess the deterministic ranking.
_GAP_STATUSES = {"missing", "suspicious"}

# Max labeled pairs to hand the focused GPT read (legacy fallback payload).
_MAX_PAIRS = 450

# Max Stage-1 sheets rendered whole for the focused GPT read. Tier-ordered, so
# the one-pager / inputs sheets always make the cut; the per-sheet and total
# char caps in render_sheets_text bound the token cost.
_MAX_READ_SHEETS = 8

# Stage 1 reads ONLY the sheets an analyst opens first: the one-pager (name
# tier 1), inputs / assumptions / sources&uses (tier 2), and secondary
# summaries like general info / key UW metrics (tier 3). Cash flows, debt
# schedules, returns tabs, and portfolio-reference sheets are Stage-2 territory
# (formula trace). Anything above this tier is excluded from the first pass.
_STAGE1_MAX_TIER = 3


def extract_aam(
    file_path: str | Path,
    layer: str = "underwriting",
    sheet_classification: dict[str, dict] | None = None,
    use_gpt_gap_fill: bool = True,
    sheet_tier_map: dict[str, int] | None = None,
) -> dict[str, Any]:
    """
    Run the focused Stage-1 AAM extraction.

    Args:
      file_path:            workbook to read.
      layer:                SSOT layer (kept for parity / future per-layer AAM).
      sheet_classification: pre-computed role map; if None and an API key is
                            present, classify_sheets is called once (GPT #1).
      use_gpt_gap_fill:     run the single focused GPT gap-fill (GPT #2) for AAM
                            fields the deterministic pass left missing/ambiguous.
      sheet_tier_map:       pre-computed {sheet: tier} from Workbook Orientation.
                            When supplied it drives sheet priority directly and
                            the GPT sheet classification (call #1) is skipped.

    Returns:
      { metric_name: resolver_record }  for every AAM field, in AAM order.
    """
    file_path = Path(file_path)
    catalog = load_metric_catalog()
    # prefer_hardcoded: for modeler-input fields (cap rates, debt terms, dates,
    # identity), a hard-coded cell outranks a formula cell on the same sheet
    # rank — read the model the way it was built.
    aam = [
        dict(m, prefer_hardcoded=(m["metric_id"] in AAM_HARDCODED_INPUTS))
        for m in aam_metrics(catalog)
    ]

    # --- Step 1: Workbook Mapper ----------------------------------------------
    # Preferred source: the Workbook Orientation tier map (deterministic,
    # content-based, computed before extraction). Falls back to the GPT
    # classifier (call #1) / name-based tiers when no map is supplied.
    if sheet_tier_map is not None:
        classification = sheet_classification or {}
        log.info(
            "Using Workbook Orientation tier map for %s (%d sheets) — "
            "GPT sheet classification skipped.",
            file_path.name, len(sheet_tier_map),
        )
    else:
        classification, sheet_tier_map = _build_tier_map(file_path, sheet_classification)

    # --- Step 1.5: Stage-1 scope — summary / one-pager / inputs sheets ONLY ---
    stage1_tiers, formula_cells = _stage1_context(file_path, sheet_tier_map)
    log.info(
        "Stage-1 scope for %s: %s",
        file_path.name,
        ", ".join(sorted(n for n, t in stage1_tiers.items() if t != 99)) or "(fallback: all)",
    )

    # --- Step 2: Deterministic resolve, scoped to the AAM + Stage-1 sheets ----
    candidates_by_metric = scan_workbook_for_candidates(file_path, aam, stage1_tiers)
    for cands in candidates_by_metric.values():
        for c in cands:
            fset = formula_cells.get(c.get("sheet"))
            if fset is not None:
                c["is_hardcoded"] = c.get("value_cell") not in fset
    records: dict[str, Any] = {}
    for m in aam:
        cands = candidates_by_metric.get(m["metric_id"], [])
        records[m["metric_name"]] = resolve_metric(m, cands)

    det_counts = _status_counts(records)
    log.info(
        "AAM deterministic pass for %s -- %s",
        file_path.name, ", ".join(f"{k}={v}" for k, v in sorted(det_counts.items())),
    )

    # --- Step 3: Focused GPT gap-fill (GPT call #2, single batched call) ------
    if use_gpt_gap_fill and llm_available():
        gaps = [m for m in aam if records[m["metric_name"]]["status"] in _GAP_STATUSES]
        if gaps:
            filled = _focused_gap_fill(
                file_path, gaps, records, stage1_tiers, formula_cells
            )
            log.info(
                "AAM focused GPT gap-fill for %s -- attempted %d, filled %d",
                file_path.name, len(gaps), filled,
            )

    # --- Step 4: AAM-scoped normalization (representation fixes) --------------
    # The audit appendix must show the human the RIGHT number to verify. These
    # transforms fix how an extracted value is represented; they do not invent
    # facts. Cross-metric derivation (e.g. Exit Date = Purchase + Hold) is left
    # to post-verification reconciliation so it never derives off an unverified
    # input the user is about to correct.
    _normalize_aam(records)

    # Drive NOI from the pricing identity (NOI = Price × Cap) rather than trust a
    # cell-matched NOI, which grabs the wrong period/column. Runs after blank-fill
    # too (via fill_aam_blanks) so newly-filled price/cap inputs feed the derive.
    _derive_noi_from_pricing(records)

    return records


def fill_aam_blanks(
    file_path: str | Path,
    records: dict[str, Any],
    sheet_tier_map: dict[str, int] | None = None,
) -> int:
    """
    Run the focused GPT gap-fill (GPT #2) over the blank/ambiguous rows of an
    EXISTING records dict (produced by a prior deterministic extract_aam).

    This is the on-demand "Fill blanks with GPT" action: one batched call over
    only the gaps, re-validated through resolve_metric, then re-normalized.
    Mutates `records` in place; returns the number of fields filled. Silently
    no-ops (returns 0) when no API key is available.

    `sheet_tier_map` is the Workbook Orientation map; when supplied the fill is
    scoped by the same orientation-driven whitelist the extraction used.
    """
    if not llm_available():
        return 0
    catalog = load_metric_catalog()
    aam = [
        dict(m, prefer_hardcoded=(m["metric_id"] in AAM_HARDCODED_INPUTS))
        for m in aam_metrics(catalog)
    ]
    gaps = [
        m for m in aam
        if records.get(m["metric_name"], {}).get("status") in _GAP_STATUSES
    ]
    if not gaps:
        return 0
    # Rebuild the Stage-1 scope from the orientation map when available, else
    # name-based; same whitelist discipline as the extract-time fill.
    stage1_tiers, formula_cells = _stage1_context(Path(file_path), sheet_tier_map)
    filled = _focused_gap_fill(
        Path(file_path), gaps, records, stage1_tiers, formula_cells
    )
    _normalize_aam(records)
    # Re-derive NOI now that blank-fill may have supplied missing price/cap inputs.
    _derive_noi_from_pricing(records)
    return filled


def _stage1_context(
    file_path: Path,
    sheet_tier_map: dict[str, int] | None,
) -> tuple[dict[str, int], dict[str, set[str]]]:
    """
    Build the Stage-1 sheet scope: a strict tier map where every sheet above
    _STAGE1_MAX_TIER is forced to 99 (skip), plus a per-sheet set of FORMULA
    cell coordinates for the whitelisted sheets (bounded scan window).

    The formula map is how we read the model the way it was built: a cell whose
    value is typed in (no formula) is a modeler INPUT; a formula cell is derived.
    Falls back to the unrestricted tiers if the whitelist would be empty.
    """
    import openpyxl
    strict: dict[str, int] = {}
    formulas: dict[str, set[str]] = {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    except Exception as e:
        log.error("Stage-1 context build failed for %s: %s", file_path.name, e)
        return sheet_tier_map or {}, {}

    def _tier(name: str) -> int:
        if sheet_tier_map and name in sheet_tier_map:
            return sheet_tier_map[name]
        return sheet_priority_tier(name)

    for name in wb.sheetnames:
        t = _tier(name)
        strict[name] = t if t <= _STAGE1_MAX_TIER else 99

    if all(t == 99 for t in strict.values()):
        log.warning(
            "Stage-1 whitelist is EMPTY for %s (no summary/one-pager/inputs "
            "sheets recognized) — falling back to the unrestricted scan.",
            file_path.name,
        )
        strict = {name: _tier(name) for name in wb.sheetnames}

    for name, t in strict.items():
        if t == 99:
            continue
        ws = wb[name]
        fset: set[str] = set()
        try:
            for row in ws.iter_rows(min_row=1, max_row=250, min_col=1, max_col=60):
                for cell in row:
                    v = cell.value
                    if (isinstance(v, str) and v.startswith("=")) or \
                            v.__class__.__name__ == "ArrayFormula":
                        fset.add(cell.coordinate)
        except Exception:
            pass
        formulas[name] = fset

    try:
        wb.close()
    except Exception:
        pass
    return strict, formulas


def _build_tier_map(
    file_path: Path,
    classification: dict[str, dict] | None,
) -> tuple[dict[str, dict], dict[str, int] | None]:
    """
    Resolve the content-role classification + per-sheet tier map.

    Reuses a caller-supplied classification when given; otherwise runs
    classify_sheets once (GPT #1). Returns ({}, None) gracefully when no API key
    is available, so the deterministic scan falls back to name-based tiers.
    """
    classification = classification or {}
    if not classification and llm_available():
        try:
            from sheet_classifier import classify_sheets
            classification = classify_sheets(file_path) or {}
        except Exception as e:
            log.error("Sheet classification failed for %s: %s", file_path.name, e)
            classification = {}

    if not classification:
        return {}, None

    try:
        import openpyxl
        from sheet_classifier import effective_tier
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet_tier_map = {
            name: effective_tier(name, sheet_priority_tier(name), classification)
            for name in wb.sheetnames
        }
        wb.close()
        return classification, sheet_tier_map
    except Exception as e:
        log.error("Tier-map build failed for %s: %s", file_path.name, e)
        return classification, None


def _grounded_value(wb_values, stage1_tiers: dict, sheet: str, cell: str,
                    claimed, unit: str | None):
    """
    Ground a GPT hit against the actual workbook: the cell must be on a Stage-1
    sheet, must exist, must be non-empty, and its REAL value must loosely match
    what GPT claimed (×1 / ×100 / ÷100 for numbers — percent-representation
    slack only, NOT ×1000, so a fabricated "192,000" can't pass against a
    192,000,000 cell; containment for text; same-day for dates).

    Returns the ACTUAL cell value to use as the candidate (truth comes from the
    workbook, never from GPT's mouth), or None to reject the hit.
    """
    if stage1_tiers.get(sheet, 99) == 99:
        return None
    try:
        actual = wb_values[sheet][cell].value
    except Exception:
        return None
    if actual is None or (isinstance(actual, str) and not actual.strip()):
        return None

    if unit == "text":
        a, c = str(actual).strip().lower(), str(claimed).strip().lower()
        if not c or not a:
            return None
        return str(actual).strip() if (c in a or a in c) else None

    if unit == "date":
        da, dc = _to_date(actual), _to_date(claimed)
        return actual if (da and dc and da == dc) else None

    from metric_resolver import parse_numeric_value
    cv, ok = parse_numeric_value(claimed, unit)
    av = _as_num(actual)
    if not ok or av is None or not isinstance(cv, (int, float)):
        return None
    for eq in (cv, cv * 100.0, cv / 100.0):
        if abs(eq - av) <= 0.02 * max(abs(av), 1e-9):
            return actual
    return None


def _focused_gap_fill(
    file_path: Path,
    gaps: list[dict],
    records: dict[str, Any],
    stage1_tiers: dict[str, int],
    formula_cells: dict[str, set[str]],
) -> int:
    """
    One batched GPT call over ALL gap metrics, restricted to the Stage-1 sheets.
    Each returned hit is GROUNDED against the workbook (cell must exist on a
    whitelisted sheet and contain what GPT says it contains — the actual cell
    value is what gets used), then re-validated through resolve_metric so schema
    bounds still apply. Mutates `records` in place; returns fields filled.
    """
    # PRIMARY payload: the WHOLE-SHEET text of the Stage-1 sheets (summary /
    # inputs / returns), tier-ordered — the model reads the actual pages an
    # analyst reads, with table structure intact, instead of a flattened
    # label/value pair list. Every claimed hit is still grounded against the
    # real cell and re-validated through resolve_metric below — comprehension
    # got better; the trust machinery is unchanged.
    from workbook_orientation import render_sheets_text
    stage1_sheets = _select_read_sheets(stage1_tiers)
    cells_block = ""
    if stage1_sheets:
        try:
            cells_block = render_sheets_text(file_path, stage1_sheets, formula_cells)
        except Exception as e:
            log.error("Whole-sheet render failed for %s: %s", file_path.name, e)

    if not cells_block:
        # Fallback: the legacy labeled-pair payload.
        pairs = extract_raw_labeled_pairs(file_path, max_pairs=1200)
        pairs = [p for p in pairs if stage1_tiers.get(p.get("sheet"), 99) != 99]
        if not pairs:
            return 0
        for p in pairs:
            fset = formula_cells.get(p.get("sheet"))
            p["is_hardcoded"] = (fset is not None) and (p.get("cell") not in fset)
        cells_block = _pairs_block(pairs, stage1_tiers)

    result = _aam_gpt_call(gaps, cells_block)
    if not result:
        return 0

    import openpyxl
    try:
        wb_values = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        log.error("Gap-fill grounding open failed for %s: %s", file_path.name, e)
        return 0

    filled = 0
    for m in gaps:
        hit = result.get(m["metric_id"])
        if not hit:
            continue
        claimed = hit.get("value")
        if claimed in (None, "", "-", "—"):
            continue

        sheet, cell = hit.get("sheet") or "", hit.get("cell") or ""
        value = _grounded_value(
            wb_values, stage1_tiers, sheet, cell, claimed, m.get("unit")
        )
        if value is None:
            log.info(
                "Gap-fill hit REJECTED for %s: %s!%s claimed %r (off-scope, "
                "missing, or doesn't match the actual cell).",
                m["metric_id"], sheet, cell, claimed,
            )
            continue

        fset = formula_cells.get(sheet)
        candidate = {
            "metric_id":     m["metric_id"],
            "metric_name":   m["metric_name"],
            "value":         value,
            "source_file":   file_path.name,
            "sheet":         sheet,
            "sheet_tier":    stage1_tiers.get(sheet, sheet_priority_tier(sheet)),
            "label_cell":    cell,
            "value_cell":    cell,
            "matched_alias": "(AAM focused read)",
            "confidence":    "medium",
            "label_ratio":   1.0,
            "match_method":  "aam_focused",
            "is_hardcoded":  (fset is not None) and (cell not in fset),
        }
        prev_status = records[m["metric_name"]]["status"]
        rec = resolve_metric(m, [candidate])
        # Only accept the GPT value if it actually validated to something usable.
        # Reject "suspicious" too: that means the value failed schema-range
        # validation (e.g. GPT grabbed the $192M purchase-price cell for Total SF,
        # which is far outside the SF range). A blank the human can fill is better
        # than a confident-looking out-of-range number. The field keeps its prior
        # (missing/blank) state rather than absorbing the bad fill.
        if rec["status"] in ("missing", "suspicious"):
            continue
        rec.setdefault("validation_notes", []).insert(
            0,
            f"Filled by focused AAM GPT read (deterministic pass left it "
            f"'{prev_status}').",
        )
        rec["_via_aam_gpt"] = True
        records[m["metric_name"]] = rec
        filled += 1

    try:
        wb_values.close()
    except Exception:
        pass
    return filled


_AAM_SYSTEM_PROMPT = """\
You are a real estate analyst reading an underwriting workbook. You are given
the FULL CONTENT of the model's authoritative sheets — the summary / one-pager
/ inputs / assumptions / returns tabs an analyst reads first. Each sheet is
rendered row by row; every cell carries its A1 reference (e.g. C5=224100).
You are also given a short list of specific metrics to find. Read the sheets
the way an analyst would — headers, table structure, and context matter. For
each requested metric, return the single best cell: its value, the exact sheet
name, and the cell reference.

Rules:
- Use the value EXACTLY as it appears in the cell. NEVER compute, infer, scale,
  or adjust a value — your answer is checked against the actual cell and any
  mismatch is rejected.
- A trailing * on a number (e.g. C5=0.0817*) marks a HARD-CODED modeler input
  (not a formula). For deal assumptions — cap rates, interest rate / spread /
  cap, LTV, dates, unit counts — prefer a * cell over a computed one.
- PORTFOLIO MODELS: a deal may bundle several assets (two hotels, asset +
  parking). Per-asset columns or allocation blocks show SLICES of the deal —
  always pick the COMBINED / TOTAL deal-level cell, never one asset's share.
- Mind each sheet's units: a header like "$ in 000s" means cell values are
  thousands — still return the cell's literal value (the checker compares
  against the cell), never a rescaled one.
- If you cannot find a metric with reasonable confidence, OMIT it (do not guess).
- Confidence is "high", "medium", or "low".

Return ONLY JSON, no prose, no code fences:
{
  "<metric_id>": {"value": <number or string>, "sheet": "<exact sheet>",
                  "cell": "<A1>", "confidence": "high|medium|low"},
  ...
}
"""


def _select_read_sheets(stage1_tiers: dict[str, int]) -> list[str]:
    """
    Pick the sheets to read WHOLE for the focused GPT pass — the analyst's
    short stack: a few summary sheets, the inputs sheet(s), one more. Within
    each tier the NAME tier breaks ties (an explicit "One Pager" / "Executive
    Summary" name beats template tabs that content-classified into the same
    tier), mirroring the resolver's own tiebreak. Backfills to _MAX_READ_SHEETS
    from the remaining whitelisted sheets when a tier is thin.
    """
    by_name_rank = lambda n: (sheet_priority_tier(n), n)
    picked: list[str] = []
    for tier, quota in ((1, 4), (2, 2), (3, 2)):
        group = sorted((n for n, t in stage1_tiers.items() if t == tier), key=by_name_rank)
        picked.extend(group[:quota])
    if len(picked) < _MAX_READ_SHEETS:
        rest = sorted(
            (n for n, t in stage1_tiers.items() if t != 99 and n not in picked),
            key=lambda n: (stage1_tiers[n], *by_name_rank(n)),
        )
        picked.extend(rest[: _MAX_READ_SHEETS - len(picked)])
    return picked[:_MAX_READ_SHEETS]


def _pairs_block(pairs: list[dict], stage1_tiers: dict[str, int] | None) -> str:
    """Legacy fallback payload: flat labeled-cell lines, tier-ordered."""
    def _ptier(p):
        if stage1_tiers:
            return stage1_tiers.get(p.get("sheet"), p.get("sheet_tier", 99))
        return p.get("sheet_tier", 99)
    pairs_sorted = sorted(pairs, key=_ptier)[:_MAX_PAIRS]
    return "\n".join(
        f"{p['sheet']}!{p['cell']}  {p['label']} = {p['value']}"
        + (" [input]" if p.get("is_hardcoded") else "")
        for p in pairs_sorted
    )


def _aam_gpt_call(gaps: list[dict], cells_block: str) -> dict[str, dict]:
    """Single focused GPT call over the rendered sheet content.
    Returns { metric_id: {value, sheet, cell, confidence} }."""
    metric_lines = []
    for m in gaps:
        rng = f"[{m.get('range_min')}, {m.get('range_max')}]"
        metric_lines.append(
            f"- {m['metric_id']} ({m['metric_name']}): "
            f"{m.get('definition', '')[:120]} "
            f"| unit={m.get('unit', '?')} | valid range={rng}"
        )

    user_msg = (
        "METRICS TO FIND:\n" + "\n".join(metric_lines)
        + "\n\nWORKBOOK CONTENT (authoritative sheets):\n" + cells_block
    )

    try:
        from knowledge_store import with_active_rules
        system_content = with_active_rules(_AAM_SYSTEM_PROMPT, ["metric_resolution", "validation"])
        response = client.chat.completions.create(
            model=MODEL_FAST,
            temperature=0.0,
            messages=[
                {"role": "system", "content": system_content},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError as e:
        log.error("AAM focused read JSON parse failed: %s", e)
        return {}
    except Exception as e:
        log.error("AAM focused read API call failed: %s", e)
        return {}


def _as_num(v: Any) -> float | None:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_date(v: Any):
    import datetime as _dt
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return _dt.datetime.strptime(v[:19], fmt).date()
            except ValueError:
                continue
    return None


def _derived_hold_years(records: dict[str, Any]) -> float | None:
    """Years between Purchase Date and Exit Date, if both are present."""
    pd_rec, ed_rec = records.get("Purchase Date"), records.get("Exit Date")
    if not pd_rec or not ed_rec:
        return None
    a = _to_date(pd_rec.get("normalized_value") or pd_rec.get("raw_value"))
    b = _to_date(ed_rec.get("normalized_value") or ed_rec.get("raw_value"))
    if a and b:
        return abs((b - a).days) / 365.25
    return None


def _interpret_hold_value(v: float, derived: float | None) -> tuple[float, str] | None:
    """
    Decide whether a raw Hold Period value should be read as months and
    converted to years. Returns (years, note) if a conversion is warranted,
    else None (leave as-is).

      - With dates: pick the interpretation (years vs months) closest to the
        Purchase->Exit span. Disambiguates 12 / 18 / 24 correctly.
      - Without dates: fall back to the >24 month heuristic.
    """
    if derived is not None and derived > 0:
        fits_years = abs(v - derived)
        fits_months = abs(v / 12.0 - derived)
        if fits_months + 1e-9 < fits_years:
            years = round(v / 12.0, 1)
            return years, (
                f"Hold Period {v:g} read as MONTHS via date cross-check "
                f"(Purchase->Exit ~ {derived:.1f}y; {v:g}/12 ~ {years:.1f}y fits better "
                f"than {v:g}y)."
            )
        return None  # v already fits as years
    if 24 < v <= 360:
        years = round(v / 12.0, 1)
        return years, (
            f"Pattern fired: hold_period_gt_24_means_months. Normalized {v:g} "
            f"months -> {years:.1f} years (no dates available to cross-check)."
        )
    return None


def _normalize_aam(records: dict[str, Any]) -> None:
    """
    Representation-only normalization, scoped to AAM fields, applied in place.

    Currently handles the two transforms whose raw form misleads a verifier:
      - Hold Period stored in MONTHS -> years (the active
        `hold_period_gt_24_means_months` rule, mirrored from reconciliation).
      - Equity Multiple displayed as a multiplier ("1.40x"), not a percentage
        (the shared `ratio` formatter renders sub-1.5 ratios as %, which is
        correct for cap rates / LTV but wrong for an equity multiple).
    """
    # Hold Period: months -> years. Prefer a date cross-check (Purchase + Exit
    # dates) which disambiguates the ambiguous low values (12 / 18 / 24) the raw
    # >24 threshold can't; fall back to the threshold only when no dates exist.
    hp = records.get("Hold Period")
    if hp and hp.get("status") != "missing":
        v = _as_num(hp.get("normalized_value"))
        if v is None:
            v = _as_num(hp.get("raw_value"))
        if v is not None and v > 0:
            derived = _derived_hold_years(records)
            interp = _interpret_hold_value(v, derived)
            if interp:
                years, note = interp
                hp["normalized_value"] = years
                hp["display_value"] = f"{years:.1f} years"
                hp.setdefault("validation_notes", []).insert(0, note)

            # Sanity cross-check: if Purchase->Exit dates derive a hold the
            # extracted value disagrees with under BOTH the years AND the months
            # reading (even after conversion), the extracted cell is probably the
            # wrong one (e.g. a "Year 1" index, not the hold). Flag suspicious so
            # the human corrects it instead of trusting a wrong number. We do NOT
            # overwrite the value (dates may be unverified) — only surface it.
            if derived is not None and derived > 0:
                final_v = _as_num(hp.get("normalized_value")) or v
                best_err = min(
                    abs(final_v - derived),
                    abs(v - derived),
                    abs(v / 12.0 - derived),
                )
                if best_err / derived > 0.2:
                    hp["status"] = "suspicious"
                    hp.setdefault("validation_notes", []).insert(
                        0,
                        f"Hold Period {v:g} disagrees with the Purchase->Exit "
                        f"span (~{derived:.1f} years). The extracted cell may be "
                        f"wrong (e.g. a period index). Suggested: "
                        f"{derived:.1f} years."
                    )

    # Equity Multiple: always render as a multiplier.
    em = records.get("Equity Multiple")
    if em and em.get("status") != "missing":
        v = _as_num(em.get("normalized_value"))
        if v is not None:
            em["display_value"] = f"{v:.2f}x"


def _by_id(records: dict[str, Any], metric_id: str) -> dict[str, Any] | None:
    """Find an AAM record by metric_id (records are keyed by metric_name)."""
    for rec in records.values():
        if rec.get("metric_id") == metric_id:
            return rec
    return None


def _derive_noi_from_pricing(
    records: dict[str, Any],
    skip_ids: set[str] | None = None,
) -> None:
    """
    Drive NOI from the cap-rate identity (Price = NOI / Cap Rate), in place.

    An analyst reads NOI OUT of pricing, not the other way around: cell-matching
    grabs whichever NOI column it finds first (usually the exit/stabilized one,
    the higher number), which is why going-in and exit NOI collided on one cell.
    Deriving from price × cap is both more reliable and ties the appendix's NOI
    to the price and cap rate the human is verifying right beside it.

        Going-in NOI = Purchase Price × Going-in Cap Rate
        Exit NOI     = Exit Value     × Exit Cap Rate

    Best-effort and self-checking:
      - derives only when BOTH pricing inputs resolved to a usable value;
      - the COMPUTED NOI is range-checked against its own catalog schema, so a
        bad input (e.g. a cap that actually matched the NOI cell) yields an
        out-of-range product and is rejected rather than written;
      - a NOI whose id is in `skip_ids` is left untouched — used at confirm time
        when the human explicitly corrected the NOI itself (their value wins);
      - when a derived value disagrees with the extracted one by >10%, the note
        flags the extracted cell as the likely wrong period/column.
    """
    from metric_resolver import _format_display

    skip_ids = skip_ids or set()
    schema_by_id = {m["metric_id"]: m for m in load_metric_catalog()}

    # Only derive from a pricing input that actually validated. A 'suspicious'
    # value failed schema-range (e.g. BAC's going-in cap matched the $15.6M NOI
    # cell, not the 8.17% rate cell, before GPT fill) — feeding it would produce
    # a nonsense NOI. Better to leave the extracted NOI until the input is fixed.
    _USABLE = {"verified", "derived", "candidate_pool", "inferred"}

    def _num(rec: dict | None) -> float | None:
        if not rec or rec.get("status") not in _USABLE:
            return None
        v = _as_num(rec.get("normalized_value"))
        return v if (v is not None and v > 0) else None

    def _in_range(metric_id: str, value: float) -> bool:
        schema = schema_by_id.get(metric_id, {})
        lo, hi = schema.get("range_min"), schema.get("range_max")
        if lo is not None and value < lo:
            return False
        if hi is not None and value > hi:
            return False
        return True

    # (derived NOI id, price-input id, cap-input id, human-readable formula)
    rules = [
        ("net_operating_income_noi", "purchase_price", "going_in_cap_rate",
         "Purchase Price × Going-in Cap Rate"),
        ("exit_noi", "exit_value_terminal_value", "exit_cap_rate",
         "Exit Value × Exit Cap Rate"),
    ]
    for noi_id, price_id, cap_id, formula in rules:
        if noi_id in skip_ids:
            continue
        noi = _by_id(records, noi_id)
        if noi is None:
            continue
        price = _num(_by_id(records, price_id))
        cap = _num(_by_id(records, cap_id))
        if price is None or cap is None:
            continue  # missing a pricing input — leave the extracted NOI as-is

        derived = price * cap
        if not _in_range(noi_id, derived):
            continue  # nonsense product (bad input) — don't overwrite the NOI
        prev = _as_num(noi.get("normalized_value"))
        noi["raw_value"] = derived
        noi["normalized_value"] = derived
        noi["display_value"] = _format_display(derived, "USD", None)
        noi["status"] = "derived"
        noi["source_sheet"] = None
        noi["source_cell"] = formula  # shows the identity instead of a cell
        note = (f"Derived from pricing: {formula} = "
                f"{price:,.0f} × {cap:.4f} = {derived:,.0f}.")
        if prev is not None and abs(prev - derived) / max(derived, 1.0) > 0.10:
            note += (f" Extracted cell showed {prev:,.0f} — likely the wrong "
                     f"period/column; the pricing identity governs.")
        noi.setdefault("validation_notes", []).insert(0, note)


def _status_counts(records: dict[str, Any]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for r in records.values():
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    return counts
