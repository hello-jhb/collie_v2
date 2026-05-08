from pathlib import Path
from datetime import datetime
import json
import openpyxl


GOING_IN_CAP_RATE = 0.0575


def safe_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    return 0


def money(value):
    return round(float(value), 2)


def pct(value):
    return round(float(value) * 100, 2)


def find_row(ws, search_text, prefer_last=True):
    """Find row where any cell contains search_text. Prefer last match because financial statements may have header rows."""
    search_text = search_text.lower()
    matches = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and search_text in str(cell.value).lower():
                matches.append(cell.row)

    if not matches:
        return None

    return matches[-1] if prefer_last else matches[0]


def row_total(ws, row_num, start_col, end_col):
    """Sum numeric values across a row."""
    total = 0
    for col in range(start_col, end_col + 1):
        total += safe_number(ws.cell(row=row_num, column=col).value)
    return money(total)


def extract_acquisition_model(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    ws_assumption = wb["Assumption"]

    original_noi = ws_assumption["D7"].value
    going_in_cap_rate = ws_assumption["D8"].value
    purchase_price = ws_assumption["D9"].value
    ti_lc = ws_assumption["D10"].value
    closing_cost = ws_assumption["D11"].value
    total_basis = ws_assumption["D12"].value
    exit_cap_rate = ws_assumption["D15"].value
    forward_noi = ws_assumption["D16"].value
    exit_value = ws_assumption["D17"].value
    initial_debt = ws_assumption["D23"].value
    total_debt = ws_assumption["D25"].value

    # Pull IRRs from Proforma sheet if available
    ws_proforma = wb["Proforma"]
    
    unlevered_irr = ws_proforma["F43"].value
    levered_irr = ws_proforma["F61"].value

    print("DEBUG IRR:", unlevered_irr, levered_irr)


    return {
        "source_file": Path(file_path).name,
        "type": "acquisition_underwriting",
        "purchase_price": money(purchase_price),
        "original_noi": money(original_noi),
        "going_in_cap_rate": going_in_cap_rate,
        "implied_value_at_going_in_cap": money(original_noi / going_in_cap_rate),
        "ti_lc_budget": money(ti_lc),
        "closing_cost": money(closing_cost),
        "total_basis": money(total_basis),
        "initial_debt": money(initial_debt),
        "total_debt": money(total_debt),
        "exit_cap_rate": exit_cap_rate,
        "forward_noi": money(forward_noi),
        "exit_value": money(exit_value),
        "unlevered_irr": unlevered_irr,
        "levered_irr": levered_irr,
    }


def extract_financial_statement(file_path, year):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[str(year)]

    # Financial statements use monthly columns starting at D.
    # 2021 has Jan-Dec = D:O
    # 2022 has Jan-Sep = D:L and total in N.
    if year == 2021:
        start_col = 4
        end_col = 15
        months_count = 12
    else:
        start_col = 4
        end_col = 12
        months_count = 9

    revenue_row = find_row(ws, "TOTAL OPERATING REVENUE", prefer_last=True)
    opex_row = find_row(ws, "TOTAL OPERATING EXPENSES", prefer_last=True)
    noi_row = find_row(ws, "NET OPERATING INCOME", prefer_last=True)

    revenue = row_total(ws, revenue_row, start_col, end_col)
    opex = row_total(ws, opex_row, start_col, end_col)
    noi = row_total(ws, noi_row, start_col, end_col)

    # Expense detail rows
    expense_lines = {
        "repair_maintenance": "Total Repair & Maintenance Exp",
        "utilities": "Total Utilities Expense",
        "property_tax": "Total Property Tax Expense",
        "insurance": "Total Property Insurance Exp",
        "property_management": "Total Property Management Exp",
        "administrative": "Total Administrative Expense",
        "leasing_marketing": "Total Lease & Marketing Exp",
        "professional_services": "Total Professional Service Exp",
    }

    expenses = {}

    for key, label in expense_lines.items():
        row_num = find_row(ws, label)
        if row_num:
            expenses[key] = row_total(ws, row_num, start_col, end_col)
        else:
            expenses[key] = None

    annualized_noi = noi if months_count == 12 else money(noi / months_count * 12)
    implied_value = money(annualized_noi / GOING_IN_CAP_RATE)

    return {
        "source_file": Path(file_path).name,
        "type": "actual_financial_statement",
        "year": year,
        "period_months": months_count,
        "revenue": revenue,
        "opex": opex,
        "noi": noi,
        "annualized_noi": annualized_noi,
        "implied_value_at_going_in_cap": implied_value,
        "expenses": expenses,
    }


def extract_business_plan_2022(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    ws = wb["Proforma"]

    # Pull return metrics from 2022 Business Plan Proforma
    bp_unlevered_irr = ws["F43"].value
    bp_levered_irr = ws["F61"].value

    # In the BP model, Jan-Sep 2022 columns are identified by date row.
    date_row = 6

    target_cols = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=date_row, column=col).value
        if isinstance(value, datetime):
            if value.year == 2022 and value.month <= 9:
                target_cols.append(col)

    def sum_by_label(label):
        row_num = find_row(ws, label)
        if not row_num:
            return None

        total = 0
        for col in target_cols:
            total += safe_number(ws.cell(row=row_num, column=col).value)
        return money(total)

    revenue = sum_by_label("Effective Gross Revenue")
    opex = abs(sum_by_label("Total Operating Expenses"))
    noi = sum_by_label("Net Operating Income")

    annualized_noi = money(noi / len(target_cols) * 12)
    implied_value = money(annualized_noi / GOING_IN_CAP_RATE)

    expenses = {
        "repair_maintenance": abs(sum_by_label("Repair & Maintenance Expense")),
        "utilities": abs(sum_by_label("Utilities Expense")),
        "property_tax": abs(sum_by_label("Property Tax Expense")),
        "insurance": abs(sum_by_label("Property Insurance Expense")),
        "property_management": abs(sum_by_label("Property Management Expense")),
        "administrative": abs(sum_by_label("Administrative Expense")),
        "leasing_marketing": abs(sum_by_label("Leasing & Marketing Expense")),
        "professional_services": abs(sum_by_label("Professional Service Expense")),
    }

    return {
        "source_file": Path(file_path).name,
        "type": "business_plan",
        "year": 2022,
        "period": "Jan-Sep",
        "revenue": revenue,
        "opex": opex,
        "noi": noi,
        "annualized_noi": annualized_noi,
        "implied_value_at_going_in_cap": implied_value,
        "expenses": expenses,
        "bp_unlevered_irr": bp_unlevered_irr,
        "bp_levered_irr": bp_levered_irr,
    }


def calculate_variance(bp, actual):
    revenue_var = actual["revenue"] - bp["revenue"]
    opex_var = actual["opex"] - bp["opex"]
    noi_var = actual["noi"] - bp["noi"]

    expense_variances = {}

    for key, bp_value in bp["expenses"].items():
        actual_value = actual["expenses"].get(key)

        if bp_value is not None and actual_value is not None:
            expense_variances[key] = {
                "bp": money(bp_value),
                "actual": money(actual_value),
                "variance": money(actual_value - bp_value),
                "variance_pct": round((actual_value - bp_value) / bp_value * 100, 1)
                if bp_value != 0
                else None,
            }

    return {
        "revenue_variance": money(revenue_var),
        "revenue_variance_pct": round(revenue_var / bp["revenue"] * 100, 1),
        "opex_variance": money(opex_var),
        "opex_variance_pct": round(opex_var / bp["opex"] * 100, 1),
        "noi_variance": money(noi_var),
        "noi_variance_pct": round(noi_var / bp["noi"] * 100, 1),
        "expense_variances": expense_variances,
    }


def build_diagnosis(acquisition, actual_2021, bp_2022, actual_2022):
    variance_2022 = calculate_variance(bp_2022, actual_2022)

    original_noi = acquisition["original_noi"]
    actual_2021_noi = actual_2021["noi"]
    actual_2022_noi = actual_2022["annualized_noi"]

    value_change_2021 = actual_2021["implied_value_at_going_in_cap"] - acquisition["implied_value_at_going_in_cap"]
    value_change_2022 = actual_2022["implied_value_at_going_in_cap"] - acquisition["implied_value_at_going_in_cap"]

    sorted_expenses = sorted(
        variance_2022["expense_variances"].items(),
        key=lambda item: item[1]["variance"],
        reverse=True,
    )

    top_expense_leaks = sorted_expenses[:4]

    return {
        "summary": {
            "original_noi": original_noi,
            "actual_2021_noi": actual_2021_noi,
            "actual_2022_annualized_noi": actual_2022_noi,
            "bp_2022_ytd_noi": bp_2022["noi"],
            "actual_2022_ytd_noi": actual_2022["noi"],
            "acquisition_implied_value": acquisition["implied_value_at_going_in_cap"],
            "actual_2021_implied_value": actual_2021["implied_value_at_going_in_cap"],
            "actual_2022_implied_value": actual_2022["implied_value_at_going_in_cap"],
            "value_change_2021": money(value_change_2021),
            "value_change_2022": money(value_change_2022),
            "unlevered_irr": acquisition["unlevered_irr"],
            "levered_irr": acquisition["levered_irr"],
        },
        "variance_2022": variance_2022,
        "top_expense_leaks": top_expense_leaks,
        "diagnosis_text": (
            "2022 underperformance is primarily driven by operating margin compression. "
            "Revenue is only modestly below plan, while operating expenses are materially above plan. "
            "The system should distinguish recurring expense pressure, such as utilities and property tax, "
            "from potentially episodic items such as R&M or one-time bad debt."
        ),
    }


def run_extraction(upload_folder="uploads"):
    upload_folder = Path(upload_folder)

    acquisition_file = None
    bp_file = None
    fs_2021_file = None
    fs_2022_file = None

    for file in upload_folder.iterdir():
        name = file.name.lower()

        if "acquisition" in name or "underwriting" in name:
            acquisition_file = file
        elif "business plan" in name or "bp" in name:
            bp_file = file
        elif "2021" in name and "financial" in name:
            fs_2021_file = file
        elif "2022" in name and "financial" in name:
            fs_2022_file = file

    missing = []
    if not acquisition_file:
        missing.append("acquisition underwriting file")
    if not bp_file:
        missing.append("2022 business plan file")
    if not fs_2021_file:
        missing.append("2021 financial statement file")
    if not fs_2022_file:
        missing.append("2022 financial statement file")

    if missing:
        return {
            "status": "missing_files",
            "missing": missing,
        }

    acquisition = extract_acquisition_model(acquisition_file)
    bp_2022 = extract_business_plan_2022(bp_file)
    actual_2021 = extract_financial_statement(fs_2021_file, 2021)
    actual_2022 = extract_financial_statement(fs_2022_file, 2022)

    diagnosis = build_diagnosis(acquisition, actual_2021, bp_2022, actual_2022)

    result = {
        "status": "success",
        "acquisition": acquisition,
        "business_plan_2022": bp_2022,
        "actual_2021": actual_2021,
        "actual_2022": actual_2022,
        "diagnosis": diagnosis,
    }

    repository = Path("repository")
    repository.mkdir(exist_ok=True)

    with open(repository / "extracted_metrics.json", "w") as f:
        json.dump(result, f, indent=2, default=str)

    return result
